#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import base64
import json
import os
import re
import sys
import threading
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QObject, QThread, Qt, Signal
from PySide6.QtWidgets import (
    QApplication,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


ICSP_BASE = "https://icsp.scpgroup.com.cn"
ICSP_CLIENT_ID = "2a5c64fcf8cf475593350a6d11548711"
ICSP_SALT = "d0a8155e8e84e5832c3a908056737c2b"

PLAZA_CODE = "G002Z008C0030"
TENANT_ID = "10000"
ORG_TYPE_CODE = "10003"

MEMBER_QUERY_URL = ICSP_BASE + "/icsp-member/web/member/queryPageList"
QUERY_PAGE_SIZE = 10
MAX_QUERY_WORKERS = 100

RETURN_FIELD_FILE = "会员等级返回字段.txt"


def _read_text_with_fallback(path: str) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            with open(path, "r", encoding=encoding) as f:
                return f.read()
        except Exception:
            continue
    return ""


def load_return_fields() -> list[str]:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    target = os.path.join(base_dir, RETURN_FIELD_FILE)
    if not os.path.isfile(target):
        for name in os.listdir(base_dir):
            if name.endswith(".txt") and ("会员等级" in name or "返回字段" in name):
                target = os.path.join(base_dir, name)
                break
    if not os.path.isfile(target):
        return []

    raw = _read_text_with_fallback(target).strip()
    if not raw:
        return []

    # 文件可能是截断的 JSON 片段，用正则提取字段名更稳妥
    keys = re.findall(r'"([^"]+)"\s*:', raw)
    unique = []
    seen = set()
    for k in keys:
        if k not in seen:
            seen.add(k)
            unique.append(k)
    return unique


def normalize_mobile(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)  # 兼容 Excel 数字手机号
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 11:
        return digits[-11:]
    return digits


def is_valid_mobile(value: str) -> bool:
    return bool(re.fullmatch(r"1\d{10}", value))


def extract_rows_and_total(payload: dict) -> tuple[list[dict], int]:
    if not isinstance(payload, dict):
        return [], 0

    rows = []
    total = 0

    if isinstance(payload.get("rows"), list):
        rows = payload["rows"]
    elif isinstance(payload.get("list"), list):
        rows = payload["list"]
    elif isinstance(payload.get("data"), list):
        rows = payload["data"]
    elif isinstance(payload.get("data"), dict):
        nested = payload["data"]
        if isinstance(nested.get("rows"), list):
            rows = nested["rows"]
        elif isinstance(nested.get("list"), list):
            rows = nested["list"]
        elif isinstance(nested.get("records"), list):
            rows = nested["records"]
        total = int(nested.get("total") or nested.get("totalCount") or nested.get("totalSize") or 0)

    if not total:
        total = int(payload.get("total") or payload.get("totalCount") or payload.get("totalSize") or 0)
    if not total and rows:
        total = len(rows)
    return rows, total


def read_input_excel(path: str) -> tuple[list[str], list[list], int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    rows = [r for r in rows if any(cell is not None and str(cell).strip() != "" for cell in r)]
    if not rows:
        raise ValueError("导入文件没有可用数据")

    first_row = rows[0]
    first_row_text = [str(c).strip().lower() if c is not None else "" for c in first_row]
    header_keywords = ("手机号", "手机", "mobile", "mobileno", "phone")

    phone_col_idx = -1
    for idx, name in enumerate(first_row_text):
        if any(k in name for k in header_keywords):
            phone_col_idx = idx
            break

    if phone_col_idx >= 0:
        headers = [str(c).strip() if c is not None else f"列{idx + 1}" for idx, c in enumerate(first_row)]
        data_rows = rows[1:]
    else:
        # 无明确表头时，自动探测手机号最多的一列
        max_cols = max(len(r) for r in rows)
        best_idx = 0
        best_score = -1
        sample_size = min(len(rows), 200)
        for idx in range(max_cols):
            score = 0
            for r in rows[:sample_size]:
                cell = r[idx] if idx < len(r) else None
                if is_valid_mobile(normalize_mobile(cell)):
                    score += 1
            if score > best_score:
                best_score = score
                best_idx = idx
        phone_col_idx = best_idx
        headers = [f"列{i + 1}" for i in range(max_cols)]
        data_rows = rows

    if phone_col_idx < 0:
        raise ValueError("未识别到手机号列")

    normalized_rows = []
    for r in data_rows:
        row = list(r) + [None] * (len(headers) - len(r))
        normalized_rows.append(row[: len(headers)])

    return headers, normalized_rows, phone_col_idx


def export_result_excel(headers: list[str], rows: list[list], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "会员等级结果"

    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin

    for row in rows:
        ws.append(row)

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row_cells:
            cell.alignment = center
            cell.border = thin

    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        col = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_num in range(2, ws.max_row + 1):
            val = ws[f"{col}{row_num}"].value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[col].width = min(max_len * 1.2 + 4, 60)

    wb.save(output_path)


class WorkerSignals(QObject):
    log = Signal(str, str)
    success = Signal(str)
    error = Signal(str)
    stopped = Signal(str)
    finished = Signal()


class ICSPClient:
    def __init__(self, signals: WorkerSignals, thread_ref=None):
        self.signals = signals
        self.thread_ref = thread_ref
        self.session = requests.Session()
        self.user_info = {"userid": "", "usercode": "", "username": ""}
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36"
                ),
                "Accept": "application/json, text/plain, */*",
                "Origin": ICSP_BASE,
                "Referer": ICSP_BASE + "/login.html",
            }
        )

    def log(self, level: str, msg: str):
        self.signals.log.emit(level, msg)

    def check_stop(self):
        if self.thread_ref and getattr(self.thread_ref, "is_interrupted", False):
            raise InterruptedError("用户手动停止任务")

    @staticmethod
    def make_passwd(password: str) -> str:
        raw = (ICSP_SALT + password).encode("utf-8")
        return f"{base64.b64encode(raw).decode()}.{ICSP_SALT}"

    def login(self, username: str, password: str) -> bool:
        self.check_stop()
        form = {"clientId": ICSP_CLIENT_ID, "passwd": self.make_passwd(password), "user": username}
        headers = {"Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}
        self.log("INFO", f"[ICSP] 登录中: {username}")
        try:
            auth_resp = self.session.post(
                ICSP_BASE + "/icsp-permission/web/permission/sso/auth/authCode",
                data=form,
                headers=headers,
                allow_redirects=False,
                timeout=15,
            )
            self.check_stop()

            auth_code = ""
            if auth_resp.status_code == 302 and "Location" in auth_resp.headers:
                auth_code = auth_resp.headers["Location"].split("authCode=")[-1]
            elif auth_resp.status_code == 200:
                body = auth_resp.json()
                if body.get("success") and body.get("data"):
                    auth_code = str(body["data"])

            if not auth_code:
                self.log("ERROR", "未获取到 authCode")
                return False

            ts = str(int(time.time() * 1000))
            self.session.get(f"{ICSP_BASE}/auth.html?authCode={auth_code}", timeout=15)
            self.check_stop()
            self.session.get(
                f"{ICSP_BASE}/icsp-permission/web/wd/login/login/sso?_t={ts}&authCode={auth_code}",
                timeout=15,
            )
            self.check_stop()

            user_resp = self.session.get(f"{ICSP_BASE}/icsp-employee/web/login/query/v2?_t={ts}", timeout=15)
            if user_resp.status_code == 200:
                data = user_resp.json().get("data", {})
                self.user_info["userid"] = str(data.get("id", ""))
                self.user_info["usercode"] = str(data.get("loginCode", username))
                self.user_info["username"] = urllib.parse.quote(str(data.get("userName", "")))

            self.log("SUCCESS", "[ICSP] 登录成功")
            return True
        except InterruptedError:
            raise
        except Exception as exc:
            self.log("ERROR", f"[ICSP] 登录异常: {exc}")
            return False

    def api_headers(self) -> dict:
        return {
            "plazacode": PLAZA_CODE,
            "orgcode": PLAZA_CODE,
            "orgtypecode": ORG_TYPE_CODE,
            "tenantid": TENANT_ID,
            "groupcode": "G001",
            "internalid": "1",
            "vunioncode": "U001",
            "workingorgcode": PLAZA_CODE,
            "userid": self.user_info["userid"],
            "usercode": self.user_info["usercode"],
            "username": self.user_info["username"],
            "Content-Type": "application/json;charset=utf-8",
            "Referer": ICSP_BASE + "/scpg.html",
            "Accept": "*/*",
            "accept-language": "zh-CN",
        }

    def build_worker_session(self) -> requests.Session:
        sess = requests.Session()
        sess.headers.update(self.session.headers)
        sess.cookies.update(self.session.cookies)
        return sess

    def query_one_mobile(self, mobile: str, session: requests.Session) -> dict:
        self.check_stop()
        payload = {"pageNo": 1, "pageSize": QUERY_PAGE_SIZE, "tagIds": [], "mobileNo": mobile}
        resp = session.post(MEMBER_QUERY_URL, headers=self.api_headers(), json=payload, timeout=20)
        resp.raise_for_status()
        rows, _ = extract_rows_and_total(resp.json())
        if not rows:
            return {"levelName": "", "queryStatus": "NOT_FOUND"}
        first = rows[0]
        return {
            "levelName": first.get("levelName", ""),
            "queryStatus": "OK",
        }

    def query_member_levels_concurrent(self, mobiles: list[str], workers: int = MAX_QUERY_WORKERS) -> dict[str, dict]:
        result: dict[str, dict] = {}
        if not mobiles:
            return result

        workers = min(max(1, workers), len(mobiles))
        self.log("INFO", f"[会员等级] 准备并发查询: {len(mobiles)} 个手机号, 并发={workers}")

        local_data = threading.local()

        def task(mobile: str) -> tuple[str, dict]:
            self.check_stop()
            if not hasattr(local_data, "session"):
                local_data.session = self.build_worker_session()
            try:
                data = self.query_one_mobile(mobile, local_data.session)
                return mobile, data
            except InterruptedError:
                raise
            except Exception as exc:
                return mobile, {"levelName": "", "queryStatus": f"ERROR: {exc}"}

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(task, m): m for m in mobiles}
            done = 0
            total = len(futures)
            for future in as_completed(futures):
                self.check_stop()
                mobile = futures[future]
                try:
                    m, data = future.result()
                    result[m] = data
                except InterruptedError:
                    for f in futures:
                        f.cancel()
                    raise
                except Exception as exc:
                    result[mobile] = {"levelName": "", "queryStatus": f"ERROR: {exc}"}
                done += 1
                if done % 20 == 0 or done == total:
                    self.log("INFO", f"[会员等级] 查询进度: {done}/{total}")

        self.log("SUCCESS", f"[会员等级] 查询完成: {len(result)} 条")
        return result


class TaskThread(QThread):
    def __init__(self, username: str, password: str, input_file: str, signals: WorkerSignals):
        super().__init__()
        self.username = username
        self.password = password
        self.input_file = input_file
        self.signals = signals
        self.is_interrupted = False

    def stop(self):
        self.is_interrupted = True

    def run(self):
        try:
            self.signals.log.emit("INFO", f"读取文件: {self.input_file}")
            headers, rows, phone_col_idx = read_input_excel(self.input_file)
            if self.is_interrupted:
                raise InterruptedError("用户手动停止任务")

            mobiles_ordered = []
            unique_set = set()
            for r in rows:
                mobile = normalize_mobile(r[phone_col_idx] if phone_col_idx < len(r) else "")
                if is_valid_mobile(mobile):
                    mobiles_ordered.append(mobile)
                    unique_set.add(mobile)

            if not unique_set:
                self.signals.error.emit("未识别到有效手机号（需为11位手机号）")
                return

            self.signals.log.emit(
                "INFO",
                f"识别手机号: {len(mobiles_ordered)} 条, 去重后 {len(unique_set)} 条",
            )

            return_fields = load_return_fields()
            if return_fields:
                self.signals.log.emit("INFO", f"返回字段文件解析成功，共 {len(return_fields)} 个字段")
                if "levelName" not in return_fields:
                    self.signals.log.emit("WARN", "返回字段文件未包含 levelName，仍将按接口结果尝试获取")

            client = ICSPClient(self.signals, self)
            if not client.login(self.username, self.password):
                self.signals.error.emit("ICSP 登录失败，请检查账号密码")
                return

            level_map = client.query_member_levels_concurrent(sorted(unique_set), workers=MAX_QUERY_WORKERS)
            if self.is_interrupted:
                raise InterruptedError("用户手动停止任务")

            out_headers = headers + ["levelName", "queryStatus"]
            out_rows = []
            for r in rows:
                mobile = normalize_mobile(r[phone_col_idx] if phone_col_idx < len(r) else "")
                if is_valid_mobile(mobile):
                    info = level_map.get(mobile, {"levelName": "", "queryStatus": "NOT_FOUND"})
                    out_rows.append(list(r) + [info.get("levelName", ""), info.get("queryStatus", "")])
                else:
                    out_rows.append(list(r) + ["", "INVALID_MOBILE"])

            base_dir = os.path.dirname(os.path.abspath(self.input_file))
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(base_dir, f"会员等级查询结果_{ts}.xlsx")
            export_result_excel(out_headers, out_rows, output_path)

            self.signals.success.emit(
                f"导出完成: {output_path}\n"
                f"输入行数: {len(rows)}\n"
                f"有效手机号: {len(unique_set)}"
            )
        except InterruptedError:
            self.signals.stopped.emit("任务已停止")
        except Exception as exc:
            self.signals.error.emit(str(exc))
        finally:
            self.signals.finished.emit()


class ModernInput(QWidget):
    def __init__(self, label: str, is_password: bool = False):
        super().__init__()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)
        lbl = QLabel(label)
        lbl.setStyleSheet("color: #556070; font-weight: 600;")
        inp = QLineEdit()
        inp.setFixedHeight(34)
        inp.setStyleSheet("padding: 0 10px; border: 1px solid #D9DDE4; border-radius: 8px;")
        if is_password:
            inp.setEchoMode(QLineEdit.Password)
        layout.addWidget(lbl)
        layout.addWidget(inp)
        self.inp = inp

    def text(self) -> str:
        return self.inp.text().strip()


class FileInput(QWidget):
    def __init__(self, label: str, parent):
        super().__init__()
        self.parent_window = parent
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        lbl = QLabel(label)
        lbl.setStyleSheet("color: #556070; font-weight: 600;")
        layout.addWidget(lbl)

        row = QHBoxLayout()
        row.setSpacing(8)
        self.path_edit = QLineEdit()
        self.path_edit.setReadOnly(True)
        self.path_edit.setPlaceholderText("请选择包含手机号的 Excel 文件")
        self.path_edit.setFixedHeight(34)
        self.path_edit.setStyleSheet("padding: 0 10px; border: 1px solid #D9DDE4; border-radius: 8px;")

        self.btn = QPushButton("浏览")
        self.btn.setFixedHeight(34)
        self.btn.setCursor(Qt.PointingHandCursor)
        self.btn.setStyleSheet(
            "QPushButton {background: #0EA5E9; color: white; border-radius: 8px; font-weight: 600;}"
            "QPushButton:disabled {background: #7DD3FC;}"
        )
        self.btn.clicked.connect(self.choose_file)

        row.addWidget(self.path_edit)
        row.addWidget(self.btn)
        layout.addLayout(row)

    def choose_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.parent_window,
            "选择手机号Excel文件",
            "",
            "Excel Files (*.xlsx *.xlsm *.xltx *.xltm);;All Files (*.*)",
        )
        if file_path:
            self.path_edit.setText(file_path)

    def text(self) -> str:
        return self.path_edit.text().strip()


class LogPanel(QTextEdit):
    def __init__(self):
        super().__init__()
        self.setReadOnly(True)
        self.setStyleSheet("background: #1F2937; color: #D1D5DB; border-radius: 8px; padding: 8px;")

    def append_log(self, level: str, msg: str):
        color = {
            "SUCCESS": "#34D399",
            "WARN": "#FBBF24",
            "ERROR": "#F87171",
            "INFO": "#D1D5DB",
        }.get(level, "#D1D5DB")
        self.append(f'<span style="color:{color};">[{level}]</span> {msg}')


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("会员等级批量查询导出")
        self.setFixedSize(560, 640)
        self.worker = None

        self.start_btn_style = (
            "QPushButton {background: #2563EB; color: white; border-radius: 8px; font-weight: 700;}"
            "QPushButton:disabled {background: #93C5FD;}"
        )
        self.stop_btn_style = (
            "QPushButton {background: #DC2626; color: white; border-radius: 8px; font-weight: 700;}"
            "QPushButton:disabled {background: #FCA5A5;}"
        )

        root = QWidget()
        self.setCentralWidget(root)
        out = QVBoxLayout(root)
        out.setContentsMargins(20, 20, 20, 20)
        out.setSpacing(12)

        title = QLabel("ICSP 会员等级查询导出")
        title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1F2937;")
        out.addWidget(title)

        card = QFrame()
        card.setStyleSheet("background: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 12px;")
        form = QVBoxLayout(card)
        form.setContentsMargins(16, 16, 16, 16)
        form.setSpacing(12)

        row_account = QHBoxLayout()
        self.username = ModernInput("账号")
        self.password = ModernInput("密码", is_password=True)
        row_account.addWidget(self.username)
        row_account.addWidget(self.password)
        form.addLayout(row_account)

        self.file_input = FileInput("手机号Excel文件", self)
        form.addWidget(self.file_input)

        hint = QLabel("说明：导入Excel后将按手机号并发查询会员等级（levelName），默认100并发。")
        hint.setStyleSheet("color: #6B7280; font-size: 12px;")
        form.addWidget(hint)

        self.btn_run = QPushButton("开始查询并导出")
        self.btn_run.setFixedHeight(38)
        self.btn_run.setCursor(Qt.PointingHandCursor)
        self.btn_run.setStyleSheet(self.start_btn_style)
        self.btn_run.clicked.connect(self.start_task)
        form.addWidget(self.btn_run)

        out.addWidget(card)
        self.log_panel = LogPanel()
        out.addWidget(self.log_panel)

        self.signals = WorkerSignals()
        self.signals.log.connect(self.log_panel.append_log)
        self.signals.error.connect(self.on_error)
        self.signals.success.connect(self.on_success)
        self.signals.stopped.connect(self.on_stopped)
        self.signals.finished.connect(self.on_finished)

    def start_task(self):
        if self.worker and self.worker.isRunning():
            self.log_panel.append_log("WARN", "正在停止任务，请稍候...")
            self.worker.stop()
            self.btn_run.setText("停止中...")
            self.btn_run.setEnabled(False)
            return

        user = self.username.text()
        pwd = self.password.text()
        input_file = self.file_input.text()

        if not user or not pwd:
            QMessageBox.warning(self, "提示", "请输入账号和密码")
            return
        if not input_file:
            QMessageBox.warning(self, "提示", "请选择手机号Excel文件")
            return
        if not os.path.isfile(input_file):
            QMessageBox.warning(self, "提示", "所选文件不存在")
            return

        self.log_panel.clear()
        self.btn_run.setEnabled(True)
        self.btn_run.setText("停止运行")
        self.btn_run.setStyleSheet(self.stop_btn_style)
        self.log_panel.append_log("INFO", f"开始任务，输入文件: {input_file}")

        self.worker = TaskThread(user, pwd, input_file, self.signals)
        self.worker.start()

    def on_error(self, msg: str):
        self.log_panel.append_log("ERROR", msg)
        QMessageBox.critical(self, "失败", msg)

    def on_success(self, msg: str):
        self.log_panel.append_log("SUCCESS", msg.replace("\n", " | "))
        QMessageBox.information(self, "完成", msg)

    def on_stopped(self, msg: str):
        self.log_panel.append_log("WARN", msg)

    def on_finished(self):
        self.btn_run.setEnabled(True)
        self.btn_run.setText("开始查询并导出")
        self.btn_run.setStyleSheet(self.start_btn_style)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
