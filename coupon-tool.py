#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import requests
import time
import base64
import urllib.parse
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from PySide6.QtCore import Qt, Signal, QObject, QThread, QPointF, QDate
from PySide6.QtGui import QColor, QCursor
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QFrame, QTextEdit,
    QGraphicsDropShadowEffect, QDateEdit, QMessageBox
)

# ====== UI 配色 ======
COLOR_BG = "#F5F5F7"
COLOR_CARD = "#FFFFFF"
COLOR_PRIMARY = "#007AFF"
COLOR_STOP = "#FF3B30"
COLOR_ACCENT_BG = "#E5F0FF"
COLOR_TEXT_H1 = "#1D1D1F"
COLOR_TEXT_H2 = "#86868B"
COLOR_LOG_BG = "#282C34"
COLOR_LOG_TEXT = "#ABB2BF"
COLOR_BORDER = "#E5E5EA"

DIM_W, DIM_H = 480, 520
RAD_CARD = 20
RAD_INPUT = 12
FONT_FAMILY = "Microsoft YaHei UI"

GLOBAL_STYLES = f"""
    QWidget {{ color: {COLOR_TEXT_H1}; font-family: "{FONT_FAMILY}"; }}
    QMainWindow {{ background-color: {COLOR_BG}; }}
    QLineEdit, QDateEdit, QTextEdit {{
        background-color: {COLOR_CARD}; color: {COLOR_TEXT_H1};
        border: 1px solid {COLOR_BORDER}; border-radius: {RAD_INPUT}px;
        selection-background-color: {COLOR_PRIMARY}; selection-color: #FFFFFF;
    }}
    QLineEdit:disabled, QDateEdit:disabled {{ background-color: #F0F0F5; color: #AAAAAA; }}
    QLabel {{ background: transparent; }}
"""

# ====== 业务配置 ======
ICSP_CLIENT_ID = "2a5c64fcf8cf475593350a6d11548711"
ICSP_SALT = "d0a8155e8e84e5832c3a908056737c2b"
PLAZA_ID = "G002Z008C0030"
TENANT_ID = "10000"
ORG_TYPE_CODE = "10003"
PAGE_SIZE = 100
GRANT_PLAZA_BU_ID = "293"
COUPON_GROUP_TENANT_ID = 468
COUPON_GROUP_APP_BU_ID = 2070
COUPON_GROUP_OPEN_STATUS = "3"


# ====== 信号 ======
class WorkerSignals(QObject):
    log = Signal(str, str)
    finished = Signal()
    error = Signal(str)
    success = Signal(str)


# ====== ICSP 客户端 ======
class ICSPClient:
    def __init__(self, signals, thread_ref=None):
        self.session = requests.Session()
        self.base = "https://icsp.scpgroup.com.cn"
        self.signals = signals
        self.thread_ref = thread_ref
        self.client_id = ICSP_CLIENT_ID
        self.salt = ICSP_SALT
        self.user_info = {"userid": "", "usercode": "", "username": ""}
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/json, text/plain, */*",
            "Origin": self.base, "Referer": self.base + "/login.html"
        })

    def check_stop(self):
        if self.thread_ref and self.thread_ref.is_interrupted:
            raise InterruptedError("用户手动停止任务")

    def log(self, level, msg):
        self.signals.log.emit(level, msg)

    def make_passwd(self, pwd):
        combined = (self.salt + pwd).encode("utf-8")
        b64 = base64.b64encode(combined).decode()
        return f"{b64}.{self.salt}"

    def login(self, user, pwd):
        self.check_stop()
        passwd_value = self.make_passwd(pwd)
        form = {"clientId": self.client_id, "passwd": passwd_value, "user": user}
        headers = {"Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}
        self.log("INFO", f"[ICSP] 正在登录: {user}...")
        try:
            resp = self.session.post(
                f"{self.base}/icsp-permission/web/permission/sso/auth/authCode",
                data=form, headers=headers, allow_redirects=False, timeout=10
            )
            auth_code = ""
            if resp.status_code == 302 and "Location" in resp.headers:
                auth_code = resp.headers["Location"].split("authCode=")[-1]
            elif resp.status_code == 200:
                js = resp.json()
                if js.get("success") and js.get("data"):
                    auth_code = js["data"]
            if not auth_code:
                raise RuntimeError("AuthCode获取失败")

            ts = str(int(time.time() * 1000))
            self.session.get(f"{self.base}/auth.html?authCode={auth_code}", timeout=10)
            self.session.get(
                f"{self.base}/icsp-permission/web/wd/login/login/sso?_t={ts}&authCode={auth_code}",
                timeout=10
            )
            user_resp = self.session.get(
                f"{self.base}/icsp-employee/web/login/query/v2?_t={ts}", timeout=10
            )
            if user_resp.status_code == 200:
                u_data = user_resp.json().get("data", {})
                self.user_info["userid"] = str(u_data.get("id", ""))
                self.user_info["usercode"] = str(u_data.get("loginCode", user))
                self.user_info["username"] = urllib.parse.quote(u_data.get("userName", ""))
            self.log("SUCCESS", "[ICSP] 登录成功")
            return True
        except Exception as e:
            self.log("ERROR", f"[ICSP] 登录异常: {str(e)}")
            return False

    def get_api_headers(self, is_json=False):
        h = {
            "plazacode": PLAZA_ID, "orgcode": PLAZA_ID, "orgtypecode": ORG_TYPE_CODE,
            "tenantid": TENANT_ID, "groupcode": "G001", "internalid": "1",
            "vunioncode": "U001", "workingorgcode": PLAZA_ID,
            "userid": self.user_info["userid"], "usercode": self.user_info["usercode"],
            "username": self.user_info["username"]
        }
        if is_json:
            h["Content-Type"] = "application/json;chartset=utf-8"
        return h

    def extract_rows_smart(self, data):
        rows, total = [], 0
        if not data:
            return [], 0
        if isinstance(data, dict):
            if "status" in data and str(data["status"]) == "5000":
                return [], 0
            keys_to_check = ["rows", "data", "list", "result", "records", "content", "items", "resultList"]
            for key in keys_to_check:
                if key in data and isinstance(data[key], list):
                    rows = data[key]
                    break
            if not rows and "data" in data and isinstance(data["data"], dict):
                sub = data["data"]
                for key in keys_to_check:
                    if key in sub and isinstance(sub[key], list):
                        rows = sub[key]
                        break
                if "total" in sub:
                    total = sub["total"]
                elif "totalCount" in sub:
                    total = sub["totalCount"]
                elif "totalSize" in sub:
                    total = sub["totalSize"]
            if total == 0:
                if "total" in data:
                    total = data["total"]
                elif "totalCount" in data:
                    total = data["totalCount"]
                elif "totalSize" in data:
                    total = data["totalSize"]
        elif isinstance(data, list):
            rows = data
        return rows, total

    def page_signature(self, rows):
        signature = []
        for row in rows[:5]:
            if not isinstance(row, dict):
                signature.append(str(row))
                continue
            for key in ("id", "userCouponCode", "orderNo", "couponTemplateId", "couponTempalteId"):
                value = row.get(key)
                if value not in (None, ""):
                    signature.append(f"{key}:{value}")
                    break
            else:
                signature.append(str(sorted(row.items()))[:120])
        return tuple(signature)

    def fetch_paginated_coupon_rows(self, url, headers, payload, label):
        all_rows = []
        page = 1
        last_signature = None
        max_pages = 500

        while True:
            self.check_stop()
            request_payload = dict(payload)
            request_payload["pageNo"] = page
            request_payload["pageSize"] = PAGE_SIZE
            try:
                resp = self.session.post(
                    url,
                    headers=headers,
                    json=request_payload,
                    timeout=(10, 20)
                )
                resp.raise_for_status()
                rows, total = self.extract_rows_smart(resp.json())
            except Exception as e:
                self.log("WARN", f"{label} 第{page}页异常: {e}")
                break

            row_count = len(rows)
            progress = f"{label} 第{page}页: {row_count} 条"
            if total:
                progress += f" / 总{total}"
            self.log("INFO", progress)

            if not rows:
                break

            signature = self.page_signature(rows)
            if signature == last_signature:
                self.log("WARN", f"{label} 第{page}页与上一页重复，已停止翻页，避免死循环")
                break
            last_signature = signature

            all_rows.extend(rows)

            if total and len(all_rows) >= int(total):
                break
            if row_count < PAGE_SIZE:
                break

            page += 1
            if page > max_pages:
                self.log("WARN", f"{label} 超过 {max_pages} 页，已停止，避免长时间卡住")
                break

        self.log("INFO", f"{label} 查询完成: {len(all_rows)} 条")
        return all_rows

    # ---- 券查询: 获取券列表 ----
    def fetch_coupon_products(self, start, end):
        """调用商品搜索接口，分页获取所有券，过滤 payAmount != 0"""
        self.log("INFO", "正在查询券列表...")
        all_rows = []
        page = 1
        s_full = f"{start} 00:00:00"
        e_full = f"{end} 23:59:59"
        while True:
            self.check_stop()
            headers = self.get_api_headers(True)
            headers["orgname"] = urllib.parse.quote("深圳湾睿印RAIL IN")
            headers["orgtypename"] = urllib.parse.quote("广场")
            headers["groupname"] = urllib.parse.quote("印力商用置业有限公司")
            headers["plazaname"] = PLAZA_ID
            headers["code"] = PLAZA_ID
            headers["areacode"] = ""
            headers["areaname"] = "null"
            headers["storecode"] = "null"
            headers["storename"] = "null"
            headers["Referer"] = self.base + "/dsp.html"
            payload = {
                "activityLabel": "", "area": "", "deliveryType": "ALL",
                "pageIndex": page, "pageSize": PAGE_SIZE, "totalSize": "",
                "plazza": PLAZA_ID, "point": 0,
                "productCode": "", "productName": "",
                "productType": "ALL",
                "exchangeTimeBegin": s_full, "exchangeTimeEnd": e_full,
                "date": [s_full, e_full],
                "status": "ALL", "shelfState": "",
                "goodsTypes": [], "category": "", "releaseChannel": "",
                "rangeOrgCode": "", "belongOrgCode": "", "belongOrgName": "",
                "publishingPlatforms": ""
            }
            try:
                resp = self.session.post(
                    f"{self.base}/yinli-xapi-b/pmp/pointmall/platform/product/productSearch",
                    headers=headers, json=payload, timeout=15
                )
                rows, total = self.extract_rows_smart(resp.json())
            except Exception as e:
                self.log("WARN", f"券列表第{page}页异常: {e}")
                break
            if not rows:
                break
            all_rows.extend(rows)
            self.log("INFO", f"券列表已获取 {len(all_rows)} 条 (总{total})")
            if len(all_rows) >= total or len(rows) < PAGE_SIZE:
                break
            page += 1

        # 过滤 payAmount != 0
        filtered = [r for r in all_rows if r.get("payAmount") and float(r.get("payAmount", 0)) != 0]
        self.log("SUCCESS", f"券列表获取完成，有效券 {len(filtered)} 个")
        return filtered

    # ---- 发券查询 ----
    def fetch_coupon_grants(self, start, end, coupon_template_id=None):
        """按 couponTempalteId 分页获取发券记录"""
        all_rows = []
        page = 1
        while True:
            self.check_stop()
            headers = {
                "Content-Type": "application/json",
                "Referer": self.base + "/coupon.html"
            }
            payload = {
                "pageNo": page, "pageSize": PAGE_SIZE,
                "beginCreateTime": start, "endCreateTime": end,
                "grantPlazaBuId": GRANT_PLAZA_BU_ID
            }
            if coupon_template_id:
                payload["couponTempalteId"] = coupon_template_id
            try:
                resp = self.session.post(
                    f"{self.base}/icsp-coupon/web/user/coupon/list",
                    headers=headers, json=payload, timeout=15
                )
                rows, total = self.extract_rows_smart(resp.json())
            except Exception as e:
                self.log("WARN", f"发券查询第{page}页异常: {e}")
                break
            if not rows:
                break
            all_rows.extend(rows)
            if total and len(all_rows) >= total:
                break
            if len(rows) < PAGE_SIZE:
                break
            page += 1
        return all_rows

    # ---- 核销查询 ----
    def fetch_coupon_certs(self, start, end, coupon_template_id=None):
        """按 couponTemplateId 分页获取核销记录"""
        all_rows = []
        page = 1
        while True:
            self.check_stop()
            headers = {
                "Content-Type": "application/json",
                "Referer": self.base + "/coupon.html"
            }
            payload = {
                "pageNo": page, "pageSize": PAGE_SIZE,
                "beginCreateTime": start, "endCreateTime": end,
                "plazaBuId": GRANT_PLAZA_BU_ID
            }
            if coupon_template_id:
                payload["couponTemplateId"] = coupon_template_id
            try:
                resp = self.session.post(
                    f"{self.base}/icsp-coupon/web/user/coupon/certificate/list",
                    headers=headers, json=payload, timeout=15
                )
                rows, total = self.extract_rows_smart(resp.json())
            except Exception as e:
                self.log("WARN", f"核销查询第{page}页异常: {e}")
                break
            if not rows:
                break
            all_rows.extend(rows)
            if total and len(all_rows) >= total:
                break
            if len(rows) < PAGE_SIZE:
                break
            page += 1
        return all_rows


# ====== 后台线程 ======
class TaskThread(QThread):
    def __init__(self, icsp_u, icsp_p, s_date, e_date, signals):
        super().__init__()
        self.icsp_u, self.icsp_p = icsp_u, icsp_p
        self.s, self.e = s_date, e_date
        self.signals = signals
        self.is_interrupted = False

    def stop(self):
        self.is_interrupted = True

    def format_ts(self, value):
        """时间戳转字符串"""
        if value is None:
            return ""
        if isinstance(value, (int, float)) and value > 1000000000000:
            try:
                return datetime.fromtimestamp(value / 1000).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                pass
        return str(value)

    def safe_sheet_name(self, prefix, name):
        """生成合法 sheet 名称（最长31字符）"""
        full = f"{prefix}-{name}"
        if len(full) > 31:
            full = full[:31]
        # 替换 Excel 不允许的字符
        for ch in ['\\', '/', '*', '?', ':', '[', ']']:
            full = full.replace(ch, '_')
        return full

    def normalize_coupon_name(self, value):
        return " ".join(str(value or "").split())

    def extract_product_template_id(self, product):
        candidates = [
            product.get("couponTemplateId"),
            product.get("couponTempalteId"),
            product.get("templateId"),
            product.get("templateCode")
        ]
        coupon_result = product.get("couponResult")
        if isinstance(coupon_result, dict):
            candidates.extend([
                coupon_result.get("couponTemplateId"),
                coupon_result.get("couponTempalteId"),
                coupon_result.get("templateId"),
                coupon_result.get("templateCode")
            ])
        for candidate in candidates:
            if candidate not in (None, ""):
                return str(candidate).strip()
        return ""

    def build_record_map(self, products, rows, row_name_field, row_template_field):
        rows_by_template_id = defaultdict(list)
        rows_by_name = defaultdict(list)
        duplicate_name_codes = defaultdict(list)
        result = {}

        for row in rows:
            template_id = str(row.get(row_template_field, "") or "").strip()
            coupon_name = self.normalize_coupon_name(row.get(row_name_field, ""))
            if template_id:
                rows_by_template_id[template_id].append(row)
            if coupon_name:
                rows_by_name[coupon_name].append(row)

        for product in products:
            code = str(product.get("productCode", ""))
            coupon_name = self.normalize_coupon_name(product.get("productName", ""))
            template_id = self.extract_product_template_id(product)
            duplicate_name_codes[coupon_name].append(code)

            if template_id and template_id in rows_by_template_id:
                result[code] = list(rows_by_template_id[template_id])
                continue
            result[code] = list(rows_by_name.get(coupon_name, []))

        duplicate_names = [
            name for name, codes in duplicate_name_codes.items()
            if name and len(codes) > 1
        ]
        return result, duplicate_names

    def build_excel(self, products, grants_map, certs_map):
        """生成 Excel 文件"""
        wb = Workbook()
        wb.remove(wb.active)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal='center', vertical='center')
        data_align = Alignment(horizontal='center', vertical='center')
        sum_font = Font(bold=True, size=11)
        sum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        link_font = Font(color="0563C1", underline="single")

        def apply_style(ws, headers, rows_data):
            """给 sheet 应用统一样式"""
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = data_align
            if rows_data:
                for cell in ws[ws.max_row]:
                    cell.font = sum_font
                    cell.fill = sum_fill
            for i, h in enumerate(headers, 1):
                col_letter = get_column_letter(i)
                max_len = len(str(h))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=i, max_col=i):
                    for cell in row:
                        val_len = len(str(cell.value)) if cell.value is not None else 0
                        if val_len > max_len:
                            max_len = val_len
                ws.column_dimensions[col_letter].width = max_len * 1.5 + 4

        # 预计算 sheet 名称映射
        grant_sheet_names = {}
        cert_sheet_names = {}
        for p in products:
            name = str(p.get("productName", ""))
            code = str(p.get("productCode", ""))
            grant_sheet_names[code] = self.safe_sheet_name("发券明细", name)
            cert_sheet_names[code] = self.safe_sheet_name("核销明细", name)

        # 去重 sheet 名称（如果截断后重复，加序号）
        used_names = set()
        for code in grant_sheet_names:
            n = grant_sheet_names[code]
            if n in used_names:
                n = n[:28] + f"_{code[-2:]}"
            used_names.add(n)
            grant_sheet_names[code] = n
        for code in cert_sheet_names:
            n = cert_sheet_names[code]
            if n in used_names:
                n = n[:28] + f"_{code[-2:]}"
            used_names.add(n)
            cert_sheet_names[code] = n

        # ---- Sheet 1: 券总表 ----
        ws_main = wb.create_sheet("券总表")
        main_headers = ["券名称", "售价", "发券数量", "核销数量"]
        ws_main.append(main_headers)

        summary_rows = []
        for p in products:
            code = str(p.get("productCode", ""))
            name = str(p.get("productName", ""))
            price = p.get("payAmount", 0)
            grant_count = len(grants_map.get(code, []))
            cert_count = len(certs_map.get(code, []))
            summary_rows.append([name, price, grant_count, cert_count])

        for row_idx, row_data in enumerate(summary_rows, 2):
            ws_main.append(row_data)
            code = str(products[row_idx - 2].get("productCode", ""))
            # 发券数量超链接
            grant_sn = grant_sheet_names.get(code, "")
            cell_grant = ws_main.cell(row=row_idx, column=3)
            if grant_sn and cell_grant.value:
                cell_grant.hyperlink = f"#'{grant_sn}'!A1"
                cell_grant.font = link_font
            # 核销数量超链接
            cert_sn = cert_sheet_names.get(code, "")
            cell_cert = ws_main.cell(row=row_idx, column=4)
            if cert_sn and cell_cert.value:
                cell_cert.hyperlink = f"#'{cert_sn}'!A1"
                cell_cert.font = link_font

        # 汇总行
        if summary_rows:
            total_grant = sum(r[2] for r in summary_rows)
            total_cert = sum(r[3] for r in summary_rows)
            ws_main.append(["合计", "", total_grant, total_cert])

        apply_style(ws_main, main_headers, summary_rows)

        # ---- 发券明细 sheets ----
        for p in products:
            code = str(p.get("productCode", ""))
            sn = grant_sheet_names.get(code, "")
            grant_rows = grants_map.get(code, [])
            ws = wb.create_sheet(sn)

            # A1 返回链接
            g_headers = ["返回总表", "", "", "", ""]
            ws.append(g_headers)
            cell_back = ws.cell(row=1, column=1)
            cell_back.value = "返回总表"
            cell_back.hyperlink = "#'券总表'!A1"
            cell_back.font = link_font

            # 数据表头 (row 2)
            data_headers = ["发券时间", "券码", "会员", "手机号", "发放渠道"]
            ws.append(data_headers)
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            if not grant_rows:
                ws.append(["无数据"])
            else:
                for r in grant_rows:
                    ws.append([
                        self.format_ts(r.get("grantTime")),
                        str(r.get("userCouponCode", "")),
                        str(r.get("memberName", "")),
                        str(r.get("phone", "")),
                        str(r.get("grantModeDesc", ""))
                    ])
                # 汇总行
                ws.append([f"合计: {len(grant_rows)} 条", "", "", "", ""])

            # 样式 (从 row 3 开始是数据)
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=5):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = data_align
            if grant_rows:
                for cell in ws[ws.max_row]:
                    cell.font = sum_font
                    cell.fill = sum_fill
            for i in range(1, 6):
                col_letter = get_column_letter(i)
                max_len = len(str(data_headers[i - 1]))
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=i, max_col=i):
                    for cell in row:
                        val_len = len(str(cell.value)) if cell.value is not None else 0
                        if val_len > max_len:
                            max_len = val_len
                ws.column_dimensions[col_letter].width = max_len * 1.5 + 4

        # ---- 核销明细 sheets ----
        for p in products:
            code = str(p.get("productCode", ""))
            sn = cert_sheet_names.get(code, "")
            cert_rows = certs_map.get(code, [])
            price = float(p.get("payAmount", 0))
            ws = wb.create_sheet(sn)

            # A1 返回链接
            ws.append(["返回总表", "", ""])
            cell_back = ws.cell(row=1, column=1)
            cell_back.value = "返回总表"
            cell_back.hyperlink = "#'券总表'!A1"
            cell_back.font = link_font

            # 按 storeName 分组统计
            store_counts = defaultdict(int)
            for r in cert_rows:
                store = str(r.get("storeName", "未知门店"))
                store_counts[store] += 1

            data_headers = ["门店名称", "核销数量", "核销金额"]
            ws.append(data_headers)
            for cell in ws[2]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            store_data = sorted(store_counts.items(), key=lambda x: -x[1])
            if not store_data:
                ws.append(["无数据", "", ""])
            else:
                for store_name, count in store_data:
                    ws.append([store_name, count, round(count * price, 2)])
                total_count = sum(c for _, c in store_data)
                ws.append(["合计", total_count, round(total_count * price, 2)])

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=3):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = data_align
            if store_data:
                for cell in ws[ws.max_row]:
                    cell.font = sum_font
                    cell.fill = sum_fill
            for i in range(1, 4):
                col_letter = get_column_letter(i)
                max_len = len(str(data_headers[i - 1]))
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=i, max_col=i):
                    for cell in row:
                        val_len = len(str(cell.value)) if cell.value is not None else 0
                        if val_len > max_len:
                            max_len = val_len
                ws.column_dimensions[col_letter].width = max_len * 1.5 + 4

        return wb

    def run(self):
        try:
            s_full = f"{self.s} 00:00:00"
            e_full = f"{self.e} 23:59:59"

            icsp = ICSPClient(self.signals, self)
            if not icsp.login(self.icsp_u, self.icsp_p):
                self.signals.error.emit("ICSP 登录失败")
                return

            # 1. 获取券列表
            products = icsp.fetch_coupon_products(self.s, self.e)
            if not products:
                self.signals.log.emit("WARN", "未查询到有效券")
                self.signals.finished.emit()
                return

            # 2. 逐券查询发券和核销
            grants_map = {}
            certs_map = {}
            for i, p in enumerate(products):
                self.check_stop_safe()
                code = str(p.get("productCode", ""))
                name = str(p.get("productName", ""))
                self.signals.log.emit("INFO", f"[{i+1}/{len(products)}] 查询券: {name}")

                grants = icsp.fetch_coupon_grants(code, s_full, e_full)
                grants_map[code] = grants
                self.signals.log.emit("INFO", f"  发券: {len(grants)} 条")

                certs = icsp.fetch_coupon_certs(code, s_full, e_full)
                certs_map[code] = certs
                self.signals.log.emit("INFO", f"  核销: {len(certs)} 条")

            # 3. 生成 Excel
            self.signals.log.emit("INFO", "正在生成 Excel...")
            wb = self.build_excel(products, grants_map, certs_map)
            exe_dir = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))
            fname = os.path.join(exe_dir, f"券统计_{self.s}_{self.e}.xlsx")
            wb.save(fname)

            total_grants = sum(len(v) for v in grants_map.values())
            total_certs = sum(len(v) for v in certs_map.values())
            summary = (f"任务完成！文件: {fname}\n\n"
                       f"券种类: {len(products)} 个\n"
                       f"发券总数: {total_grants} 条\n"
                       f"核销总数: {total_certs} 条")
            self.signals.success.emit(summary)
            self.signals.log.emit("SUCCESS", "所有任务已完成")

        except InterruptedError:
            self.signals.log.emit("WARN", "任务已停止")
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.signals.error.emit(str(e))
        finally:
            self.signals.finished.emit()

    def check_stop_safe(self):
        if self.is_interrupted:
            raise InterruptedError("用户手动停止任务")


# ====== UI 组件 ======
class ModernInput(QWidget):
    def __init__(self, label, default="", is_password=False):
        super().__init__()
        l = QVBoxLayout(self)
        l.setContentsMargins(0, 0, 0, 0)
        l.setSpacing(5)
        self.lbl = QLabel(label)
        self.lbl.setStyleSheet(f"color:{COLOR_TEXT_H2};font-weight:bold;")
        self.inp = QLineEdit(default)
        self.inp.setFixedHeight(35)
        if is_password:
            self.inp.setEchoMode(QLineEdit.Password)
        self.inp.setStyleSheet(f"""
            QLineEdit {{
                background-color: {COLOR_CARD}; color: {COLOR_TEXT_H1};
                border: 1px solid {COLOR_BORDER}; border-radius: {RAD_INPUT}px; padding: 0 10px;
            }}
        """)
        l.addWidget(self.lbl)
        l.addWidget(self.inp)

    def text(self):
        return self.inp.text().strip()


class ModernDateInput(QWidget):
    def __init__(self, label):
        super().__init__()
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(5)
        self.lbl = QLabel(label)
        self.lbl.setStyleSheet(f"color:{COLOR_TEXT_H2};font-weight:bold;")
        self.layout.addWidget(self.lbl)
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setFixedHeight(35)
        self.date_edit.setStyleSheet(f"""
            QDateEdit {{
                background-color: {COLOR_CARD}; color: {COLOR_TEXT_H1};
                border: 1px solid {COLOR_BORDER}; border-radius: {RAD_INPUT}px; padding-left: 10px;
            }}
            QDateEdit::drop-down {{
                subcontrol-origin: padding; subcontrol-position: top right; width: 30px;
                border-left: 1px solid {COLOR_BORDER};
                border-top-right-radius: {RAD_INPUT}px; border-bottom-right-radius: {RAD_INPUT}px;
                background-color: {COLOR_ACCENT_BG};
            }}
            QDateEdit::down-arrow {{
                width: 10px; height: 10px; border: none; background: none; image: none;
                border-left: 2px solid {COLOR_PRIMARY}; border-bottom: 2px solid {COLOR_PRIMARY};
                transform: rotate(-45deg); margin-top: -3px;
            }}
        """)
        self.layout.addWidget(self.date_edit)

    def setDate(self, qdate):
        self.date_edit.setDate(qdate)

    def date(self):
        return self.date_edit.date()


class LogPanel(QTextEdit):
    def __init__(self):
        super().__init__()
        self.setReadOnly(True)
        self.setStyleSheet(
            f"background:{COLOR_LOG_BG};color:{COLOR_LOG_TEXT};border-radius:10px;padding:5px;border:none;")

    def append_log(self, level, msg):
        c = "#98C379" if level == "SUCCESS" else "#E5C07B" if level == "WARN" else "#E06C75" if level == "ERROR" else "#ABB2BF"
        self.append(f'<span style="color:{c}">[{level}]</span> {msg}')


# ====== 主窗口 ======
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(GLOBAL_STYLES)
        self.setWindowTitle("券统计工具")
        from PySide6.QtGui import QIcon
        icon_path = os.path.join(os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)), "HX.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.setFixedSize(DIM_W, DIM_H)
        self.worker = None

        cw = QWidget()
        self.setCentralWidget(cw)
        lay = QVBoxLayout(cw)
        lay.setContentsMargins(20, 30, 20, 20)

        # 标题
        h = QHBoxLayout()
        ico = QLabel("🎫")
        ico.setStyleSheet(f"background:{COLOR_ACCENT_BG};font-size:24px;border-radius:8px;padding:5px;")
        tit = QLabel("券统计工具")
        tit.setStyleSheet(f"font-size:18px;font-weight:bold;color:{COLOR_TEXT_H1};")
        h.addWidget(ico)
        h.addWidget(tit)
        h.addStretch()
        lay.addLayout(h)

        # 卡片
        card = QFrame()
        card.setStyleSheet(f"QFrame{{background:{COLOR_CARD};border-radius:{RAD_CARD}px;}}")
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setColor(QColor(0, 0, 0, 20))
        shadow.setOffset(QPointF(0, 4))
        card.setGraphicsEffect(shadow)
        cl = QVBoxLayout(card)
        cl.setContentsMargins(20, 20, 20, 20)
        cl.setSpacing(12)

        # 账号密码
        row1 = QHBoxLayout()
        self.u = ModernInput("ICSP账号")
        self.p = ModernInput("ICSP密码", is_password=True)
        row1.addWidget(self.u)
        row1.addWidget(self.p)
        cl.addLayout(row1)

        # 日期
        row2 = QHBoxLayout()
        self.s = ModernDateInput("开始日期")
        self.e = ModernDateInput("结束日期")
        now = QDate.currentDate()
        self.s.setDate(QDate(now.year(), now.month(), 1))
        self.e.setDate(now.addDays(-1))
        row2.addWidget(self.s)
        row2.addWidget(self.e)
        cl.addLayout(row2)

        # 按钮
        self.btn = QPushButton("🚀 开始查询")
        self.btn.setFixedHeight(40)
        self.btn.setCursor(QCursor(Qt.PointingHandCursor))
        self.btn.setStyleSheet(
            f"background:{COLOR_PRIMARY};color:white;border-radius:{RAD_INPUT}px;border:none;font-size:14px;font-weight:bold;")
        self.btn.clicked.connect(self.run)
        cl.addWidget(self.btn)

        lay.addWidget(card)

        # 日志
        self.log = LogPanel()
        lay.addWidget(self.log)

        # 信号
        self.sig = WorkerSignals()
        self.sig.log.connect(self.log.append_log)
        self.sig.finished.connect(self.on_finished)
        self.sig.success.connect(self.on_success)
        self.sig.error.connect(lambda m: self.log.append_log("ERROR", f"任务异常: {m}"))

    def run(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.btn.setText("正在停止...")
            self.btn.setEnabled(False)
            return
        u, p = self.u.text(), self.p.text()
        if not u or not p:
            QMessageBox.warning(self, "提示", "请填写 ICSP 账号和密码")
            return
        s_str = self.s.date().toString("yyyy-MM-dd")
        e_str = self.e.date().toString("yyyy-MM-dd")

        self.log.clear()
        self.btn.setText("停止运行")
        self.btn.setStyleSheet(
            f"background:{COLOR_STOP};color:white;border-radius:{RAD_INPUT}px;border:none;font-size:14px;font-weight:bold;")
        self.worker = TaskThread(u, p, s_str, e_str, self.sig)
        self.worker.start()

    def on_finished(self):
        self.btn.setText("🚀 开始查询")
        self.btn.setEnabled(True)
        self.btn.setStyleSheet(
            f"background:{COLOR_PRIMARY};color:white;border-radius:{RAD_INPUT}px;border:none;font-size:14px;font-weight:bold;")

    def on_success(self, summary):
        QMessageBox.information(self, "任务完成", summary)

def _icspclient_page_signature(self, rows):
    signature = []
    for row in rows[:5]:
        if not isinstance(row, dict):
            signature.append(str(row))
            continue
        for key in ("id", "userCouponCode", "orderNo", "couponTemplateId", "couponTempalteId"):
            value = row.get(key)
            if value not in (None, ""):
                signature.append(f"{key}:{value}")
                break
        else:
            signature.append(str(sorted(row.items()))[:120])
    return tuple(signature)


def _icspclient_fetch_paginated_coupon_rows(self, url, headers, payload, label):
    all_rows = []
    page = 1
    last_signature = None
    max_pages = 500

    while True:
        self.check_stop()
        request_payload = dict(payload)
        request_payload["pageNo"] = page
        request_payload["pageSize"] = PAGE_SIZE
        try:
            resp = self.session.post(
                url,
                headers=headers,
                json=request_payload,
                timeout=(10, 20)
            )
            resp.raise_for_status()
            data = resp.json()
            rows, total = self.extract_rows_smart(data)
        except Exception as e:
            self.log("WARN", f"{label} page {page} error: {e}")
            break

        row_count = len(rows)
        progress = f"{label} page {page}: {row_count} rows"
        if total:
            progress += f" / total {total}"
        self.log("INFO", progress)

        if not rows:
            if page == 1 and isinstance(data, dict):
                keys = ",".join(list(data.keys())[:8])
                summary = []
                for key in ("code", "status", "success", "msg", "message"):
                    if key in data:
                        summary.append(f"{key}={data.get(key)}")
                extra = f" | {'; '.join(summary)}" if summary else ""
                self.log("WARN", f"{label} first page empty, response keys: {keys}{extra}")
            break

        signature = self.page_signature(rows)
        if signature == last_signature:
            self.log("WARN", f"{label} page {page} repeated previous page, stop paging")
            break
        last_signature = signature

        all_rows.extend(rows)

        if total and len(all_rows) >= int(total):
            break
        if row_count < PAGE_SIZE:
            break

        page += 1
        if page > max_pages:
            self.log("WARN", f"{label} exceeded {max_pages} pages, stop paging")
            break

    self.log("INFO", f"{label} done: {len(all_rows)} rows")
    return all_rows


def _icspclient_fetch_coupon_grants(self, start, end, coupon_template_id=None):
    headers = {
        "Content-Type": "application/json",
        "Referer": self.base + "/coupon.html"
    }
    payload = {
        "beginCreateTime": start,
        "endCreateTime": end,
        "grantPlazaBuId": GRANT_PLAZA_BU_ID
    }
    if coupon_template_id:
        payload["couponTempalteId"] = coupon_template_id
    return self.fetch_paginated_coupon_rows(
        f"{self.base}/icsp-coupon/web/user/coupon/list",
        headers,
        payload,
        "grant"
    )


def _icspclient_fetch_coupon_grant_total(self, start, end, coupon_template_id):
    headers = {
        "Content-Type": "application/json",
        "Referer": self.base + "/coupon.html"
    }
    payload = {
        "beginCreateTime": start,
        "endCreateTime": end,
        "grantPlazaBuId": GRANT_PLAZA_BU_ID,
        "couponTempalteId": str(coupon_template_id),
        "pageNo": 1,
        "pageSize": 1
    }
    label = f"grant-total {coupon_template_id}"
    try:
        resp = self.session.post(
            f"{self.base}/icsp-coupon/web/user/coupon/list",
            headers=headers,
            json=payload,
            timeout=(10, 20)
        )
        resp.raise_for_status()
        data = resp.json()
        rows, total = self.extract_rows_smart(data)
    except Exception as e:
        self.log("WARN", f"{label} error: {e}")
        return 0

    total_value = total
    if not total_value and isinstance(data, dict):
        total_value = (
            data.get("total")
            or data.get("totalCount")
            or data.get("totalSize")
        )
        if not total_value and isinstance(data.get("data"), dict):
            total_value = (
                data["data"].get("total")
                or data["data"].get("totalCount")
                or data["data"].get("totalSize")
            )
    try:
        total_value = int(total_value or len(rows))
    except Exception:
        total_value = len(rows)
    self.log("INFO", f"{label}: {total_value}")
    return total_value


def _icspclient_fetch_coupon_certs(self, start, end, coupon_template_id=None):
    headers = {
        "Content-Type": "application/json",
        "Referer": self.base + "/coupon.html"
    }
    url = f"{self.base}/icsp-coupon/web/user/coupon/certificate/list"
    variants = [
        ("cert", {
            "beginCreateTime": start,
            "endCreateTime": end,
            "plazaBuId": GRANT_PLAZA_BU_ID
        }),
        ("cert-certTime", {
            "beginCertificateTime": start,
            "endCertificateTime": end,
            "plazaBuId": GRANT_PLAZA_BU_ID
        }),
        ("cert-certRange", {
            "certificateTimeBegin": start,
            "certificateTimeEnd": end,
            "plazaBuId": GRANT_PLAZA_BU_ID
        }),
        ("cert-grantPlaza", {
            "beginCreateTime": start,
            "endCreateTime": end,
            "grantPlazaBuId": GRANT_PLAZA_BU_ID
        }),
        ("cert-certTime-grantPlaza", {
            "beginCertificateTime": start,
            "endCertificateTime": end,
            "grantPlazaBuId": GRANT_PLAZA_BU_ID
        }),
        ("cert-certRange-grantPlaza", {
            "certificateTimeBegin": start,
            "certificateTimeEnd": end,
            "grantPlazaBuId": GRANT_PLAZA_BU_ID
        }),
    ]
    if coupon_template_id:
        for _, payload in variants:
            payload["couponTemplateId"] = coupon_template_id

    for label, payload in variants:
        rows = self.fetch_paginated_coupon_rows(url, headers, payload, label)
        if rows:
            if label != "cert":
                self.log("INFO", f"cert matched fallback variant: {label}")
            return rows
    return []


def _icspclient_fetch_coupon_group_user_rows(
    self,
    start,
    end,
    coupon_group_no=None,
    tenant_id=COUPON_GROUP_TENANT_ID,
    app_bu_id=COUPON_GROUP_APP_BU_ID,
):
    headers = {
        "Content-Type": "application/json",
        "Referer": self.base + "/coupon.html"
    }
    payload = {
        "grantPlazaBuId": str(GRANT_PLAZA_BU_ID),
        "tenantId": tenant_id,
        "appBuId": app_bu_id,
        "createStartTime": start,
        "createEndTime": end,
        "userCouponGroupCodeList": []
    }
    if coupon_group_no not in (None, ""):
        payload["couponGroupNo"] = str(coupon_group_no)
        label = f"coupon-group {coupon_group_no}"
    else:
        label = "coupon-group-all"
    return self.fetch_paginated_coupon_rows(
        f"{self.base}/icsp-coupon/web/user/coupon/group/list/of/user",
        headers,
        payload,
        label
    )


def _icspclient_fetch_coupon_certs_by_template(self, start, end, coupon_template_id):
    headers = {
        "Content-Type": "application/json",
        "Referer": self.base + "/coupon.html"
    }
    payload = {
        "beginCreateTime": start,
        "endCreateTime": end,
        "couponTemplateId": str(coupon_template_id)
    }
    label = f"cert-template {coupon_template_id}"
    return self.fetch_paginated_coupon_rows(
        f"{self.base}/icsp-coupon/web/user/coupon/certificate/list",
        headers,
        payload,
        label
    )


def _icspclient_fetch_coupon_certs_by_name(self, start, end, coupon_name):
    headers = {
        "Content-Type": "application/json",
        "Referer": self.base + "/coupon.html"
    }
    payload = {
        "beginCreateTime": start,
        "endCreateTime": end,
        "plazaBuId": str(GRANT_PLAZA_BU_ID),
        "couponName": str(coupon_name or "").strip()
    }
    label = f"cert-name {payload['couponName']}"
    return self.fetch_paginated_coupon_rows(
        f"{self.base}/icsp-coupon/web/user/coupon/certificate/list",
        headers,
        payload,
        label
    )


ICSPClient.page_signature = _icspclient_page_signature
ICSPClient.fetch_paginated_coupon_rows = _icspclient_fetch_paginated_coupon_rows
ICSPClient.fetch_coupon_grants = _icspclient_fetch_coupon_grants
ICSPClient.fetch_coupon_grant_total = _icspclient_fetch_coupon_grant_total
ICSPClient.fetch_coupon_certs = _icspclient_fetch_coupon_certs
ICSPClient.fetch_coupon_group_user_rows = _icspclient_fetch_coupon_group_user_rows
ICSPClient.fetch_coupon_certs_by_template = _icspclient_fetch_coupon_certs_by_template
ICSPClient.fetch_coupon_certs_by_name = _icspclient_fetch_coupon_certs_by_name


def _extract_delivery_types(product):
    raw_value = product.get("deliveryType")
    if raw_value in (None, ""):
        return []
    if isinstance(raw_value, (list, tuple, set)):
        values = raw_value
    else:
        values = [raw_value]
    result = []
    for item in values:
        text = str(item or "").strip()
        if text:
            result.append(text)
    return result


def _is_coupon_package_product(product):
    return "COUPON_PACKAGE" in _extract_delivery_types(product)


def _is_coupon_product(product):
    delivery_types = _extract_delivery_types(product)
    return "COUPON" in delivery_types and "COUPON_PACKAGE" not in delivery_types


def _extract_coupon_group_no(product):
    candidates = [
        product.get("couponGroupNo"),
        product.get("couponGroupCode"),
        product.get("groupNo")
    ]
    for key in ("couponResult", "couponGroupVO"):
        value = product.get(key)
        if isinstance(value, dict):
            candidates.extend([
                value.get("couponGroupNo"),
                value.get("couponGroupCode"),
                value.get("groupNo")
            ])
    for candidate in candidates:
        text = str(candidate or "").strip()
        if text:
            return text
    return ""


def _extract_spu_code(product):
    candidates = [
        product.get("spuCode"),
        product.get("productCode"),
        product.get("couponTemplateId"),
        product.get("couponTempalteId")
    ]
    for key in ("couponResult",):
        value = product.get(key)
        if isinstance(value, dict):
            candidates.extend([
                value.get("spuCode"),
                value.get("productCode"),
                value.get("couponTemplateId"),
                value.get("couponTempalteId")
            ])
    for candidate in candidates:
        text = str(candidate or "").strip()
        if text:
            return text
    return ""


def _extract_coupon_group_query_value(product, field_name, default_value):
    candidates = [product.get(field_name)]
    for key in ("couponResult", "couponGroupVO"):
        value = product.get(key)
        if isinstance(value, dict):
            candidates.append(value.get(field_name))
    for candidate in candidates:
        if candidate not in (None, ""):
            return candidate
    return default_value


def _normalize_coupon_group_cert_row(row):
    normalized = dict(row)
    normalized["_certSource"] = "coupon_group_status_3"
    if not normalized.get("storeName"):
        coupon_group_vo = normalized.get("couponGroupVO")
        coupon_group_title = ""
        if isinstance(coupon_group_vo, dict):
            coupon_group_title = str(coupon_group_vo.get("title", "") or "").strip()
        normalized["storeName"] = (
            str(normalized.get("grantPlazaName", "") or "").strip()
            or coupon_group_title
            or "COUPON_PACKAGE"
        )
    return normalized


def _dedupe_coupon_cert_rows(rows):
    deduped_rows = []
    seen_keys = set()
    first_rows_by_key = {}
    duplicate_count = 0
    duplicate_keys = []
    duplicate_details = []

    for row in rows:
        key = None
        for field in ("id", "userCouponCode", "couponCode", "certificateNo"):
            value = row.get(field)
            if value not in (None, ""):
                key = f"{field}:{value}"
                break

        if not key:
            deduped_rows.append(row)
            continue

        if key in seen_keys:
            duplicate_count += 1
            if len(duplicate_keys) < 5:
                duplicate_keys.append(key)
            if len(duplicate_details) < 5:
                duplicate_details.append({
                    "key": key,
                    "first": first_rows_by_key.get(key),
                    "duplicate": row,
                })
            continue

        seen_keys.add(key)
        first_rows_by_key[key] = row
        deduped_rows.append(row)

    return deduped_rows, duplicate_count, duplicate_keys, duplicate_details


def _extract_product_grant_count(product):
    candidates = [
        product.get("payedNum"),
        product.get("sendCount"),
        product.get("sendAmount"),
    ]
    effective_num = product.get("effectiveNum")
    refunded_num = product.get("refundedNum")
    if effective_num not in (None, "") or refunded_num not in (None, ""):
        try:
            candidates.append(int(effective_num or 0) + int(refunded_num or 0))
        except Exception:
            pass
    candidates.extend([
        product.get("soldNum"),
        product.get("payNum"),
    ])
    for candidate in candidates:
        if candidate in (None, ""):
            continue
        try:
            return int(float(candidate))
        except Exception:
            continue
    return 0


def _extract_coupon_group_match_names(task, row):
    names = []
    coupon_group_vo = row.get("couponGroupVO")
    if isinstance(coupon_group_vo, dict):
        title = task.normalize_coupon_name(coupon_group_vo.get("title", ""))
        if title:
            names.append(title)

    remark = task.normalize_coupon_name(row.get("remark", ""))
    if remark:
        names.append(remark)
        for sep in ("-", "－", "—"):
            if sep in remark:
                suffix = task.normalize_coupon_name(remark.split(sep, 1)[1])
                if suffix:
                    names.append(suffix)
                break

    deduped = []
    seen = set()
    for name in names:
        if name and name not in seen:
            seen.add(name)
            deduped.append(name)
    return deduped


def _query_coupon_package_cert_rows(task, icsp, products, start, end, certs_map):
    missing_group_products = []
    package_codes = set()

    for product in products:
        if not _is_coupon_package_product(product):
            continue

        _taskthread_check_stop_safe(task)
        code = str(product.get("productCode", ""))
        name = str(product.get("productName", ""))
        group_no = _extract_coupon_group_no(product)
        if not group_no:
            missing_group_products.append(name or code)
            continue

        tenant_id = _extract_coupon_group_query_value(
            product, "tenantId", COUPON_GROUP_TENANT_ID
        )
        app_bu_id = _extract_coupon_group_query_value(
            product, "appBuId", COUPON_GROUP_APP_BU_ID
        )

        task.signals.log.emit("INFO", f"query coupon package group {group_no} for {name}")
        rows = icsp.fetch_coupon_group_user_rows(
            start, end, group_no, tenant_id=tenant_id, app_bu_id=app_bu_id
        )
        open_rows = [
            _normalize_coupon_group_cert_row(row)
            for row in rows
            if str(row.get("status", "") or "").strip() == COUPON_GROUP_OPEN_STATUS
        ]
        certs_map[code] = open_rows
        package_codes.add(code)
        task.signals.log.emit(
            "INFO",
            f"coupon package open rows for {name}: {len(open_rows)}"
        )

    return package_codes, missing_group_products


def _scan_coupon_package_rows(task, icsp, products, start, end):
    package_products = [product for product in products if _is_coupon_package_product(product)]
    rows_by_code = {str(product.get("productCode", "")): [] for product in package_products}
    code_to_name = {}
    name_to_codes = defaultdict(set)

    for product in package_products:
        code = str(product.get("productCode", ""))
        name = task.normalize_coupon_name(product.get("productName", ""))
        code_to_name[code] = name or code
        if name:
            name_to_codes[name].add(code)

    if not package_products:
        return rows_by_code, {code: [] for code in rows_by_code}, []

    tenant_id = COUPON_GROUP_TENANT_ID
    app_bu_id = COUPON_GROUP_APP_BU_ID
    for product in package_products:
        tenant_id = _extract_coupon_group_query_value(product, "tenantId", tenant_id)
        app_bu_id = _extract_coupon_group_query_value(product, "appBuId", app_bu_id)
        if tenant_id and app_bu_id:
            break

    task.signals.log.emit("INFO", "query coupon package rows by global group scan")
    all_rows = icsp.fetch_coupon_group_user_rows(
        start, end, tenant_id=tenant_id, app_bu_id=app_bu_id
    )

    for row in all_rows:
        matched_codes = set()
        for name in _extract_coupon_group_match_names(task, row):
            matched_codes.update(name_to_codes.get(name, set()))
        if not matched_codes:
            continue
        normalized_row = _normalize_coupon_group_cert_row(row)
        for code in matched_codes:
            rows_by_code[code].append(normalized_row)

    cert_rows_by_code = {}
    missing_products = []
    for code, rows in rows_by_code.items():
        cert_rows_by_code[code] = [
            row for row in rows
            if str(row.get("status", "") or "").strip() == COUPON_GROUP_OPEN_STATUS
        ]
        if not rows:
            missing_products.append(code_to_name.get(code, code))

    return rows_by_code, cert_rows_by_code, missing_products


def _query_coupon_cert_rows_by_spu(task, icsp, products, start, end, certs_map):
    missing_spu_products = []
    coupon_codes = set()

    for product in products:
        if not _is_coupon_product(product):
            continue

        _taskthread_check_stop_safe(task)
        code = str(product.get("productCode", ""))
        name = str(product.get("productName", ""))
        spu_code = _extract_spu_code(product)
        if not spu_code:
            certs_map[code] = []
            missing_spu_products.append(name or code)
            continue

        task.signals.log.emit("INFO", f"query coupon cert by spuCode {spu_code} for {name}")
        rows = icsp.fetch_coupon_certs_by_template(start, end, spu_code)
        deduped = {}
        for row in rows:
            key = (
                row.get("id")
                or row.get("userCouponCode")
                or row.get("orderNo")
                or repr(row)
            )
            deduped[key] = row
        certs_map[code] = list(deduped.values())
        coupon_codes.add(code)
        task.signals.log.emit(
            "INFO",
            f"coupon cert rows for {name}: {len(certs_map[code])}"
        )

    return coupon_codes, missing_spu_products


def _query_coupon_cert_rows_by_template_id(task, icsp, products, start, end, certs_map):
    missing_template_products = []
    coupon_codes = set()

    for product in products:
        if not _is_coupon_product(product):
            continue

        _taskthread_check_stop_safe(task)
        code = str(product.get("productCode", ""))
        name = str(product.get("productName", "") or "").strip()
        template_id = task.extract_product_template_id(product)
        template_source = "couponTemplateId"
        if not template_id:
            template_id = _extract_spu_code(product)
            if template_id:
                template_source = "spuCode"
        if not template_id:
            certs_map[code] = []
            missing_template_products.append(name or code)
            continue

        task.signals.log.emit(
            "INFO",
            f"query coupon cert by {template_source} {template_id} for {name or code}"
        )
        rows = icsp.fetch_coupon_certs_by_template(start, end, template_id)
        deduped_rows, _duplicate_count, _duplicate_keys, _duplicate_details = _dedupe_coupon_cert_rows(rows)
        certs_map[code] = deduped_rows
        coupon_codes.add(code)
        task.signals.log.emit(
            "INFO",
            f"coupon cert rows for {name or code}: {len(certs_map[code])}"
        )

    return coupon_codes, missing_template_products


def _query_coupon_cert_rows_by_name(task, icsp, products, start, end, certs_map):
    missing_name_products = []
    coupon_codes = set()

    for product in products:
        if not _is_coupon_product(product):
            continue

        _taskthread_check_stop_safe(task)
        code = str(product.get("productCode", ""))
        name = str(product.get("productName", "") or "").strip()
        coupon_name = task.normalize_coupon_name(name)
        if not coupon_name:
            certs_map[code] = []
            missing_name_products.append(name or code)
            continue

        task.signals.log.emit("INFO", f"query coupon cert by name {coupon_name} for {name or code}")
        rows = icsp.fetch_coupon_certs_by_name(start, end, coupon_name)
        deduped_rows, _duplicate_count, _duplicate_keys, _duplicate_details = _dedupe_coupon_cert_rows(rows)
        certs_map[code] = deduped_rows
        coupon_codes.add(code)
        task.signals.log.emit(
            "INFO",
            f"coupon cert rows for {name or code}: {len(certs_map[code])}"
        )

    return coupon_codes, missing_name_products


def _query_coupon_grant_counts_by_template(task, icsp, products, start, end):
    grant_map = {}
    missing_template_products = []

    for product in products:
        code = str(product.get("productCode", ""))
        grant_map[code] = {"count": 0, "rows": []}
        if not _is_coupon_product(product):
            continue

        _taskthread_check_stop_safe(task)
        name = str(product.get("productName", ""))
        template_id = task.extract_product_template_id(product)
        template_source = "couponTempalteId"
        if not template_id:
            template_id = _extract_spu_code(product)
            if template_id:
                template_source = "spuCode"
        if not template_id:
            missing_template_products.append(name or code)
            fallback_count = _extract_product_grant_count(product)
            grant_map[code] = {"count": fallback_count, "rows": []}
            continue

        task.signals.log.emit(
            "INFO",
            f"query coupon grant total by {template_source} {template_id} for {name}"
        )
        total_count = icsp.fetch_coupon_grant_total(start, end, template_id)
        grant_map[code] = {"count": total_count, "rows": []}

    return grant_map, missing_template_products


def _scan_coupon_template_ids_from_grants(task, icsp, products, start, end):
    coupon_products = [product for product in products if _is_coupon_product(product)]
    template_ids_by_code = {str(product.get("productCode", "")): [] for product in coupon_products}
    name_to_codes = defaultdict(set)

    for product in coupon_products:
        code = str(product.get("productCode", ""))
        product_name = task.normalize_coupon_name(product.get("productName", ""))
        if product_name:
            name_to_codes[product_name].add(code)

    unresolved_codes = {code for code in template_ids_by_code}
    page = 1
    last_signature = None
    max_pages = 500
    headers = {
        "Content-Type": "application/json",
        "Referer": icsp.base + "/coupon.html"
    }
    url = f"{icsp.base}/icsp-coupon/web/user/coupon/list"

    while unresolved_codes and page <= max_pages:
        _taskthread_check_stop_safe(task)
        payload = {
            "beginCreateTime": start,
            "endCreateTime": end,
            "grantPlazaBuId": GRANT_PLAZA_BU_ID,
            "pageNo": page,
            "pageSize": PAGE_SIZE
        }
        try:
            resp = icsp.session.post(url, headers=headers, json=payload, timeout=(10, 20))
            resp.raise_for_status()
            data = resp.json()
            rows, total = icsp.extract_rows_smart(data)
        except Exception as e:
            task.signals.log.emit("WARN", f"grant-template-scan page {page} error: {e}")
            break

        row_count = len(rows)
        progress = f"grant-template-scan page {page}: {row_count} rows"
        if total:
            progress += f" / total {total}"
        task.signals.log.emit("INFO", progress)

        if not rows:
            break

        signature = icsp.page_signature(rows)
        if signature == last_signature:
            task.signals.log.emit("WARN", f"grant-template-scan page {page} repeated previous page, stop paging")
            break
        last_signature = signature

        for row in rows:
            coupon_name = task.normalize_coupon_name(row.get("couponName", ""))
            template_id = str(
                row.get("couponTempalteId")
                or row.get("couponTemplateId")
                or ""
            ).strip()
            if not coupon_name or not template_id:
                continue
            matched_codes = name_to_codes.get(coupon_name, set())
            if not matched_codes:
                continue
            for code in matched_codes:
                existing = template_ids_by_code.setdefault(code, [])
                if template_id not in existing:
                    existing.append(template_id)
                    unresolved_codes.discard(code)

        if total and page * PAGE_SIZE >= int(total):
            break
        if row_count < PAGE_SIZE:
            break

        page += 1

    missing_codes = sorted(code for code, values in template_ids_by_code.items() if not values)
    return template_ids_by_code, missing_codes, page - 1 if page > 1 else page


def _build_cert_map_from_grants(task, products, grants_map, cert_rows):
    certs_map = {str(product.get("productCode", "")): [] for product in products}
    template_to_codes = defaultdict(set)
    name_to_codes = defaultdict(set)

    for product in products:
        code = str(product.get("productCode", ""))
        product_name = task.normalize_coupon_name(product.get("productName", ""))
        if product_name:
            name_to_codes[product_name].add(code)

        for grant in grants_map.get(code, []):
            template_id = str(
                grant.get("couponTempalteId")
                or grant.get("couponTemplateId")
                or ""
            ).strip()
            grant_name = task.normalize_coupon_name(grant.get("couponName", ""))
            if template_id:
                template_to_codes[template_id].add(code)
            if grant_name:
                name_to_codes[grant_name].add(code)

    unmatched_count = 0
    ambiguous_count = 0

    for cert in cert_rows:
        matched_code = None
        template_id = str(
            cert.get("couponTemplateId")
            or cert.get("couponTempalteId")
            or ""
        ).strip()
        cert_name = task.normalize_coupon_name(cert.get("couponName", ""))

        template_codes = template_to_codes.get(template_id, set()) if template_id else set()
        if len(template_codes) == 1:
            matched_code = next(iter(template_codes))
        elif len(template_codes) > 1:
            ambiguous_count += 1
            continue

        if not matched_code and cert_name:
            name_codes = name_to_codes.get(cert_name, set())
            if len(name_codes) == 1:
                matched_code = next(iter(name_codes))
            elif len(name_codes) > 1:
                ambiguous_count += 1
                continue

        if matched_code:
            certs_map[matched_code].append(cert)
        else:
            unmatched_count += 1

    return certs_map, unmatched_count, ambiguous_count


def _collect_template_ids_by_product(products, grants_map):
    template_ids_by_code = {}
    for product in products:
        code = str(product.get("productCode", ""))
        template_ids = []
        seen = set()
        for grant in grants_map.get(code, []):
            template_id = str(
                grant.get("couponTempalteId")
                or grant.get("couponTemplateId")
                or ""
            ).strip()
            if template_id and template_id not in seen:
                seen.add(template_id)
                template_ids.append(template_id)
        template_ids_by_code[code] = template_ids
    return template_ids_by_code


def _query_certs_by_template_ids(task, icsp, products, template_ids_by_code, start, end):
    cert_cache = {}
    certs_map = {}
    missing_template_products = []

    for product in products:
        code = str(product.get("productCode", ""))
        name = str(product.get("productName", ""))
        template_ids = template_ids_by_code.get(code, [])
        if not template_ids:
            certs_map[code] = []
            missing_template_products.append(name or code)
            continue

        rows = []
        for template_id in template_ids:
            _taskthread_check_stop_safe(task)
            if template_id not in cert_cache:
                task.signals.log.emit("INFO", f"query cert by template {template_id} for {name}")
                cert_cache[template_id] = icsp.fetch_coupon_certs(start, end, template_id)
            rows.extend(cert_cache[template_id])

        deduped = {}
        for row in rows:
            key = (
                row.get("id")
                or row.get("userCouponCode")
                or row.get("orderNo")
                or repr(row)
            )
            deduped[key] = row
        certs_map[code] = list(deduped.values())

    return certs_map, cert_cache, missing_template_products


def _taskthread_run(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP 登录失败")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "未查询到有效券")
            return

        product_template_count = sum(1 for p in products if self.extract_product_template_id(p))
        if product_template_count == 0:
            self.signals.log.emit("WARN", "券列表未返回模板ID，将按券名称匹配发券和核销数据。")

        self.signals.log.emit("INFO", "正在查询全部发券明细...")
        all_grants = icsp.fetch_coupon_grants(s_full, e_full)
        self.signals.log.emit("INFO", f"发券总数: {len(all_grants)}")

        self.signals.log.emit("INFO", "正在查询全部核销明细...")
        all_certs = icsp.fetch_coupon_certs(s_full, e_full)
        self.signals.log.emit("INFO", f"核销总数: {len(all_certs)}")

        grants_map, grant_duplicate_names = self.build_record_map(
            products, all_grants, "couponName", "couponTempalteId"
        )
        template_ids_by_code = _collect_template_ids_by_product(products, grants_map)
        certs_map, cert_cache, missing_template_products = _query_certs_by_template_ids(
            self, icsp, products, template_ids_by_code, s_full, e_full
        )

        total_cert_rows = sum(len(rows) for rows in cert_cache.values())
        self.signals.log.emit("INFO", f"query cert by template total: {total_cert_rows}")

        duplicate_names = sorted(set(grant_duplicate_names))
        if duplicate_names:
            sample_names = "、".join(duplicate_names[:5])
            if len(duplicate_names) > 5:
                sample_names += " ..."
            self.signals.log.emit(
                "WARN",
                f"检测到重名券，当前按名称匹配，建议人工复核: {sample_names}"
            )

        if cert_unmatched_count:
            self.signals.log.emit("WARN", f"未匹配到券的核销记录: {cert_unmatched_count}")
        if cert_ambiguous_count:
            self.signals.log.emit("WARN", f"匹配到多个候选券的核销记录: {cert_ambiguous_count}")

        if missing_template_products:
            sample_names = "、".join(missing_template_products[:5])
            if len(missing_template_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing template ids for cert query: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            grant_count = len(grants_map.get(code, []))
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit(
                "INFO",
                f"[{i}/{len(products)}] {name} | 发券 {grant_count} | 核销 {cert_count}"
            )

        self.signals.log.emit("INFO", "正在生成 Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"券统计_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        total_grants = sum(len(v) for v in grants_map.values())
        total_certs = sum(len(v) for v in certs_map.values())
        summary = (
            f"任务完成，文件已生成: {fname}\n\n"
            f"券种类: {len(products)}\n"
            f"发券总数: {total_grants}\n"
            f"核销总数: {total_certs}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "所有任务已完成")

    except InterruptedError:
        self.signals.log.emit("WARN", "任务已停止")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()


def _taskthread_check_stop_safe(self):
    if self.is_interrupted:
        raise InterruptedError("用户手动停止任务")


def _taskthread_run_v2(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP 登录失败")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "未查询到有效券")
            return

        self.signals.log.emit("INFO", "正在查询全部发券明细...")
        all_grants = icsp.fetch_coupon_grants(s_full, e_full)
        self.signals.log.emit("INFO", f"发券总数: {len(all_grants)}")

        grants_map, grant_duplicate_names = self.build_record_map(
            products, all_grants, "couponName", "couponTempalteId"
        )
        template_ids_by_code = _collect_template_ids_by_product(products, grants_map)

        self.signals.log.emit("INFO", "正在按模板查询核销明细...")
        certs_map, cert_cache, missing_template_products = _query_certs_by_template_ids(
            self, icsp, products, template_ids_by_code, s_full, e_full
        )
        total_certs = sum(len(rows) for rows in cert_cache.values())
        self.signals.log.emit("INFO", f"核销总数: {total_certs}")

        duplicate_names = sorted(set(grant_duplicate_names))
        if duplicate_names:
            sample_names = "、".join(duplicate_names[:5])
            if len(duplicate_names) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"duplicate grant names detected: {sample_names}")

        if missing_template_products:
            sample_names = "、".join(missing_template_products[:5])
            if len(missing_template_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing template ids for cert query: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            grant_count = len(grants_map.get(code, []))
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit("INFO", f"[{i}/{len(products)}] {name} | 发券 {grant_count} | 核销 {cert_count}")

        self.signals.log.emit("INFO", "正在生成 Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"券统计_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        total_grants = sum(len(v) for v in grants_map.values())
        summary = (
            f"任务完成，文件已生成: {fname}\n\n"
            f"券种类: {len(products)}\n"
            f"发券总数: {total_grants}\n"
            f"核销总数: {sum(len(v) for v in certs_map.values())}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "所有任务已完成")

    except InterruptedError:
        self.signals.log.emit("WARN", "任务已停止")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()


def _taskthread_run_v3(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP login failed")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "no valid coupon products found")
            return

        self.signals.log.emit("INFO", "querying all grant rows...")
        all_grants = icsp.fetch_coupon_grants(s_full, e_full)
        self.signals.log.emit("INFO", f"grant total: {len(all_grants)}")

        grants_map, grant_duplicate_names = self.build_record_map(
            products, all_grants, "couponName", "couponTempalteId"
        )
        template_ids_by_code = _collect_template_ids_by_product(products, grants_map)

        self.signals.log.emit("INFO", "querying cert rows by template...")
        certs_map, cert_cache, missing_template_products = _query_certs_by_template_ids(
            self, icsp, products, template_ids_by_code, s_full, e_full
        )

        package_codes, missing_group_products = _query_coupon_package_cert_rows(
            self, icsp, products, s_full, e_full, certs_map
        )
        if package_codes:
            package_names = {
                str(product.get("productName", "")) or str(product.get("productCode", ""))
                for product in products
                if str(product.get("productCode", "")) in package_codes
            }
            missing_template_products = [
                name for name in missing_template_products
                if name not in package_names
            ]

        total_certs = sum(len(rows) for rows in certs_map.values())
        self.signals.log.emit("INFO", f"cert total: {total_certs}")

        duplicate_names = sorted(set(grant_duplicate_names))
        if duplicate_names:
            sample_names = ", ".join(duplicate_names[:5])
            if len(duplicate_names) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"duplicate grant names detected: {sample_names}")

        if missing_template_products:
            sample_names = ", ".join(missing_template_products[:5])
            if len(missing_template_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing template ids for cert query: {sample_names}")

        if missing_group_products:
            sample_names = ", ".join(missing_group_products[:5])
            if len(missing_group_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing couponGroupNo for coupon package query: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            grant_count = len(grants_map.get(code, []))
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit(
                "INFO",
                f"[{i}/{len(products)}] {name} | grant {grant_count} | cert {cert_count}"
            )

        self.signals.log.emit("INFO", "building Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"coupon_stats_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        total_grants = sum(len(v) for v in grants_map.values())
        summary = (
            f"task completed, file generated: {fname}\n\n"
            f"product count: {len(products)}\n"
            f"grant total: {total_grants}\n"
            f"cert total: {sum(len(v) for v in certs_map.values())}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "all tasks completed")

    except InterruptedError:
        self.signals.log.emit("WARN", "task stopped")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()


def _taskthread_build_excel_v2(self, products, grants_map, certs_map):
    wb = Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    data_align = Alignment(horizontal="center", vertical="center")
    sum_font = Font(bold=True, size=11)
    sum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    def auto_width(ws, start_row, headers):
        for i, header in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=i, max_col=i):
                for cell in row:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    if val_len > max_len:
                        max_len = val_len
            ws.column_dimensions[col_letter].width = max_len * 1.5 + 4

    def style_header_row(ws, row_no, col_count):
        for cell in ws[row_no][:col_count]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_rows(ws, min_row, max_col, total_row=False):
        for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, max_col=max_col):
            for cell in row:
                cell.border = thin_border
                cell.alignment = data_align
        if total_row and ws.max_row >= min_row:
            for cell in ws[ws.max_row][:max_col]:
                cell.font = sum_font
                cell.fill = sum_fill

    grant_sheet_names = {}
    cert_sheet_names = {}
    used_names = set()

    for product in products:
        code = str(product.get("productCode", ""))
        name = str(product.get("productName", ""))
        grant_name = self.safe_sheet_name("发券明细", name)
        cert_name = self.safe_sheet_name("核销明细", name)

        if grant_name in used_names:
            grant_name = grant_name[:28] + f"_{code[-2:]}"
        used_names.add(grant_name)
        grant_sheet_names[code] = grant_name

        if cert_name in used_names:
            cert_name = cert_name[:28] + f"_{code[-2:]}"
        used_names.add(cert_name)
        cert_sheet_names[code] = cert_name

    ws_main = wb.create_sheet("券总表")
    main_headers = ["券名称", "售价", "发券数量", "核销数量"]
    ws_main.append(main_headers)
    style_header_row(ws_main, 1, len(main_headers))

    summary_rows = []
    for product in products:
        code = str(product.get("productCode", ""))
        summary_rows.append([
            str(product.get("productName", "")),
            product.get("payAmount", 0),
            len(grants_map.get(code, [])),
            len(certs_map.get(code, []))
        ])

    for row_idx, row_data in enumerate(summary_rows, 2):
        ws_main.append(row_data)
        code = str(products[row_idx - 2].get("productCode", ""))
        grant_cell = ws_main.cell(row=row_idx, column=3)
        cert_cell = ws_main.cell(row=row_idx, column=4)
        if grant_cell.value:
            grant_cell.hyperlink = f"#'{grant_sheet_names[code]}'!A1"
            grant_cell.font = link_font
        if cert_cell.value:
            cert_cell.hyperlink = f"#'{cert_sheet_names[code]}'!A1"
            cert_cell.font = link_font

    if summary_rows:
        ws_main.append([
            "合计",
            "",
            sum(row[2] for row in summary_rows),
            sum(row[3] for row in summary_rows)
        ])

    style_data_rows(ws_main, 2, len(main_headers), total_row=bool(summary_rows))
    auto_width(ws_main, 2, main_headers)

    for product in products:
        code = str(product.get("productCode", ""))
        sheet_name = grant_sheet_names[code]
        grant_rows = grants_map.get(code, [])
        ws = wb.create_sheet(sheet_name)

        ws.append(["返回总表"])
        ws["A1"].hyperlink = "#'券总表'!A1"
        ws["A1"].font = link_font

        headers = ["发券时间", "券码", "会员", "手机号", "发放渠道"]
        ws.append(headers)
        style_header_row(ws, 2, len(headers))

        if grant_rows:
            for row in grant_rows:
                ws.append([
                    self.format_ts(row.get("grantTime") or row.get("createTime")),
                    str(row.get("userCouponCode", "")),
                    str(row.get("memberName", "")),
                    str(row.get("phone", "")),
                    str(row.get("grantModeDesc", "") or row.get("grantChannelName", ""))
                ])
            ws.append([f"合计: {len(grant_rows)}", "", "", "", ""])
        else:
            ws.append(["无数据", "", "", "", ""])

        style_data_rows(ws, 3, len(headers), total_row=bool(grant_rows))
        auto_width(ws, 2, headers)

    for product in products:
        code = str(product.get("productCode", ""))
        sheet_name = cert_sheet_names[code]
        cert_rows = certs_map.get(code, [])
        price = float(product.get("payAmount", 0) or 0)
        ws = wb.create_sheet(sheet_name)

        ws.append(["返回总表"])
        ws["A1"].hyperlink = "#'券总表'!A1"
        ws["A1"].font = link_font

        if _is_coupon_package_product(product):
            headers = ["统计项", "数量"]
            ws.append(headers)
            style_header_row(ws, 2, len(headers))
            ws.append(["status=3数量", len(cert_rows)])
            style_data_rows(ws, 3, len(headers), total_row=False)
            auto_width(ws, 2, headers)
            continue

        headers = ["门店名称", "核销数量", "核销金额"]
        ws.append(headers)
        style_header_row(ws, 2, len(headers))

        store_counts = defaultdict(int)
        for row in cert_rows:
            store_name = str(row.get("storeName", "") or "未知门店")
            store_counts[store_name] += 1

        if store_counts:
            store_rows = sorted(store_counts.items(), key=lambda item: (-item[1], item[0]))
            for store_name, count in store_rows:
                ws.append([store_name, count, round(count * price, 2)])
            total_count = sum(count for _, count in store_rows)
            ws.append(["合计", total_count, round(total_count * price, 2)])
        else:
            ws.append(["无数据", "", ""])

        style_data_rows(ws, 3, len(headers), total_row=bool(store_counts))
        auto_width(ws, 2, headers)

    return wb


def _taskthread_run_v4(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP login failed")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "no valid coupon products found")
            return

        self.signals.log.emit("INFO", "skip full grant query")
        grants_map = {str(product.get("productCode", "")): [] for product in products}
        certs_map = {str(product.get("productCode", "")): [] for product in products}

        coupon_codes, missing_coupon_name_products = _query_coupon_cert_rows_by_name(
            self, icsp, products, s_full, e_full, certs_map
        )
        package_codes, missing_group_products = _query_coupon_package_cert_rows(
            self, icsp, products, s_full, e_full, certs_map
        )

        total_certs = sum(len(rows) for rows in certs_map.values())
        self.signals.log.emit("INFO", f"cert total: {total_certs}")

        if missing_grant_template_products:
            sample_names = ", ".join(missing_grant_template_products[:5])
            if len(missing_grant_template_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"缂哄皯 couponTempalteId锛屽凡鍥為€€鍒板晢鍝佺粺璁″彂鍒告暟: {sample_names}")

        if missing_grant_template_products:
            sample_names = ", ".join(missing_grant_template_products[:5])
            if len(missing_grant_template_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"缂哄皯 couponTempalteId锛屽凡鍥為€€鍒板晢鍝佺粺璁″彂鍒告暟: {sample_names}")

        if missing_spu_products:
            sample_names = ", ".join(missing_spu_products[:5])
            if len(missing_spu_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing spuCode for coupon cert query: {sample_names}")

        if missing_group_products:
            sample_names = ", ".join(missing_group_products[:5])
            if len(missing_group_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing couponGroupNo for coupon package query: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            delivery_types = "/".join(_extract_delivery_types(product)) or "UNKNOWN"
            grant_count = len(grants_map.get(code, []))
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit(
                "INFO",
                f"[{i}/{len(products)}] {name} | type {delivery_types} | grant {grant_count} | cert {cert_count}"
            )

        self.signals.log.emit("INFO", "building Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"coupon_stats_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        summary = (
            f"task completed, file generated: {fname}\n\n"
            f"product count: {len(products)}\n"
            f"coupon cert products: {len(coupon_codes)}\n"
            f"coupon package products: {len(package_codes)}\n"
            f"grant total: skipped\n"
            f"cert total: {sum(len(v) for v in certs_map.values())}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "all tasks completed")

    except InterruptedError:
        self.signals.log.emit("WARN", "task stopped")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()


TaskThread.build_excel = _taskthread_build_excel_v2
def _taskthread_run_v5(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP login failed")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "no valid coupon products found")
            return

        grants_map = {str(product.get("productCode", "")): [] for product in products}
        certs_map = {str(product.get("productCode", "")): [] for product in products}

        coupon_products = [product for product in products if _is_coupon_product(product)]
        self.signals.log.emit("INFO", "scanning grant rows for coupon template ids...")
        template_ids_by_code, missing_coupon_codes, scanned_pages = _scan_coupon_template_ids_from_grants(
            self, icsp, coupon_products, s_full, e_full
        )
        self.signals.log.emit("INFO", f"grant-template-scan finished after {scanned_pages} page(s)")

        coupon_certs_map, cert_cache, missing_template_products = _query_certs_by_template_ids(
            self, icsp, coupon_products, template_ids_by_code, s_full, e_full
        )
        certs_map.update(coupon_certs_map)
        coupon_codes = {
            str(product.get("productCode", ""))
            for product in coupon_products
            if coupon_certs_map.get(str(product.get("productCode", ""))) is not None
        }

        package_codes, missing_group_products = _query_coupon_package_cert_rows(
            self, icsp, products, s_full, e_full, certs_map
        )

        total_certs = sum(len(rows) for rows in certs_map.values())
        self.signals.log.emit("INFO", f"cert total: {total_certs}")

        if missing_coupon_codes:
            code_set = set(missing_coupon_codes)
            sample_names = [
                str(product.get("productName", "")) or str(product.get("productCode", ""))
                for product in coupon_products
                if str(product.get("productCode", "")) in code_set
            ]
            sample_text = ", ".join(sample_names[:5])
            if len(sample_names) > 5:
                sample_text += " ..."
            self.signals.log.emit("WARN", f"grant-template-scan missing coupon templates: {sample_text}")

        if missing_template_products:
            sample_names = ", ".join(missing_template_products[:5])
            if len(missing_template_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing template ids for cert query: {sample_names}")

        if missing_group_products:
            sample_names = ", ".join(missing_group_products[:5])
            if len(missing_group_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"missing couponGroupNo for coupon package query: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            delivery_types = "/".join(_extract_delivery_types(product)) or "UNKNOWN"
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit(
                "INFO",
                f"[{i}/{len(products)}] {name} | type {delivery_types} | cert {cert_count}"
            )

        self.signals.log.emit("INFO", "building Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"coupon_stats_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        summary = (
            f"task completed, file generated: {fname}\n\n"
            f"product count: {len(products)}\n"
            f"coupon cert products: {len(coupon_codes)}\n"
            f"coupon package products: {len(package_codes)}\n"
            f"grant total: skipped export / scanned for templates only\n"
            f"cert total: {sum(len(v) for v in certs_map.values())}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "all tasks completed")

    except InterruptedError:
        self.signals.log.emit("WARN", "task stopped")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()


def _taskthread_build_excel_v3(self, products, grants_map, certs_map):
    wb = Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    data_align = Alignment(horizontal="center", vertical="center")
    sum_font = Font(bold=True, size=11)
    sum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")

    def split_grant_value(value):
        if isinstance(value, dict):
            return int(value.get("count", 0) or 0), list(value.get("rows", []) or [])
        rows = list(value or [])
        return len(rows), rows

    def auto_width(ws, start_row, headers):
        for i, header in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=i, max_col=i):
                for cell in row:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    if val_len > max_len:
                        max_len = val_len
            ws.column_dimensions[col_letter].width = max_len * 1.5 + 4

    def style_header_row(ws, row_no, col_count):
        for cell in ws[row_no][:col_count]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_rows(ws, min_row, max_col, total_row=False):
        for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, max_col=max_col):
            for cell in row:
                cell.border = thin_border
                cell.alignment = data_align
        if total_row and ws.max_row >= min_row:
            for cell in ws[ws.max_row][:max_col]:
                cell.font = sum_font
                cell.fill = sum_fill

    grant_sheet_names = {}
    cert_sheet_names = {}
    used_names = set()

    for product in products:
        code = str(product.get("productCode", ""))
        name = str(product.get("productName", ""))
        grant_name = self.safe_sheet_name("发券明细", name)
        cert_name = self.safe_sheet_name("核销明细", name)

        if grant_name in used_names:
            grant_name = grant_name[:28] + f"_{code[-2:]}"
        used_names.add(grant_name)
        grant_sheet_names[code] = grant_name

        if cert_name in used_names:
            cert_name = cert_name[:28] + f"_{code[-2:]}"
        used_names.add(cert_name)
        cert_sheet_names[code] = cert_name

    ws_main = wb.create_sheet("券总表")
    main_headers = ["券名称", "售价", "发券数量", "核销数量"]
    ws_main.append(main_headers)
    style_header_row(ws_main, 1, len(main_headers))

    summary_rows = []
    for product in products:
        code = str(product.get("productCode", ""))
        grant_count, _ = split_grant_value(grants_map.get(code, {}))
        cert_count = len(certs_map.get(code, []))
        summary_rows.append([
            str(product.get("productName", "")),
            product.get("payAmount", 0),
            grant_count,
            cert_count
        ])

    for row_idx, row_data in enumerate(summary_rows, 2):
        ws_main.append(row_data)
        code = str(products[row_idx - 2].get("productCode", ""))
        if row_data[2]:
            ws_main.cell(row=row_idx, column=3).hyperlink = f"#'{grant_sheet_names[code]}'!A1"
            ws_main.cell(row=row_idx, column=3).font = link_font
        if row_data[3]:
            ws_main.cell(row=row_idx, column=4).hyperlink = f"#'{cert_sheet_names[code]}'!A1"
            ws_main.cell(row=row_idx, column=4).font = link_font

    if summary_rows:
        ws_main.append([
            "合计",
            "",
            sum(row[2] for row in summary_rows),
            sum(row[3] for row in summary_rows)
        ])

    style_data_rows(ws_main, 2, len(main_headers), total_row=bool(summary_rows))
    auto_width(ws_main, 2, main_headers)

    for product in products:
        code = str(product.get("productCode", ""))
        grant_count, grant_rows = split_grant_value(grants_map.get(code, {}))
        ws = wb.create_sheet(grant_sheet_names[code])
        ws.append(["返回总表"])
        ws["A1"].hyperlink = "#'券总表'!A1"
        ws["A1"].font = link_font

        headers = ["发券时间", "券码", "会员", "手机号", "发放渠道"]
        ws.append(headers)
        style_header_row(ws, 2, len(headers))

        if grant_rows:
            for row in grant_rows:
                code_value = row.get("userCouponCode") or row.get("userCouponGroupCode") or ""
                channel_value = row.get("grantModeDesc") or row.get("grantChannelName") or ""
                ws.append([
                    self.format_ts(row.get("grantTime") or row.get("createTime") or row.get("openTime")),
                    str(code_value),
                    str(row.get("memberName", "")),
                    str(row.get("phone", "")),
                    str(channel_value)
                ])
            ws.append([f"合计: {grant_count}", "", "", "", ""])
        else:
            ws.append([f"仅统计数量，发券数: {grant_count}", "", "", "", ""])

        style_data_rows(ws, 3, len(headers), total_row=bool(grant_rows))
        auto_width(ws, 2, headers)

    for product in products:
        code = str(product.get("productCode", ""))
        cert_rows = certs_map.get(code, [])
        price = float(product.get("payAmount", 0) or 0)
        ws = wb.create_sheet(cert_sheet_names[code])
        ws.append(["返回总表"])
        ws["A1"].hyperlink = "#'券总表'!A1"
        ws["A1"].font = link_font

        if _is_coupon_package_product(product):
            headers = ["统计项", "数量"]
            ws.append(headers)
            style_header_row(ws, 2, len(headers))
            ws.append(["status=3数量", len(cert_rows)])
            style_data_rows(ws, 3, len(headers), total_row=False)
            auto_width(ws, 2, headers)
            continue

        headers = ["门店名称", "核销数量", "核销金额"]
        ws.append(headers)
        style_header_row(ws, 2, len(headers))

        store_counts = defaultdict(int)
        for row in cert_rows:
            store_counts[str(row.get("storeName", "") or "未知门店")] += 1

        if store_counts:
            store_rows = sorted(store_counts.items(), key=lambda item: (-item[1], item[0]))
            for store_name, count in store_rows:
                ws.append([store_name, count, round(count * price, 2)])
            total_count = sum(count for _, count in store_rows)
            ws.append(["合计", total_count, round(total_count * price, 2)])
        else:
            ws.append(["无数据", "", ""])

        style_data_rows(ws, 3, len(headers), total_row=bool(store_counts))
        auto_width(ws, 2, headers)

    return wb


def _taskthread_run_v6(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP 登录失败")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "未查询到有效券")
            return

        grants_map, missing_grant_template_products = _query_coupon_grant_counts_by_template(
            self, icsp, products, s_full, e_full
        )
        certs_map = {str(product.get("productCode", "")): [] for product in products}

        coupon_products = [product for product in products if _is_coupon_product(product)]
        self.signals.log.emit("INFO", "正在扫描发券记录以提取模板ID...")
        template_ids_by_code, missing_coupon_codes, scanned_pages = _scan_coupon_template_ids_from_grants(
            self, icsp, coupon_products, s_full, e_full
        )
        self.signals.log.emit("INFO", f"模板扫描完成，共扫描 {scanned_pages} 页")

        coupon_certs_map, cert_cache, missing_template_products = _query_certs_by_template_ids(
            self, icsp, coupon_products, template_ids_by_code, s_full, e_full
        )
        certs_map.update(coupon_certs_map)

        package_rows_map, package_cert_rows_map, missing_group_products = _scan_coupon_package_rows(
            self, icsp, products, s_full, e_full
        )
        package_codes = set()
        for code, rows in package_rows_map.items():
            if rows:
                package_codes.add(code)
            grants_map[code] = {"count": len(rows), "rows": rows}
        certs_map.update(package_cert_rows_map)

        total_certs = sum(len(rows) for rows in certs_map.values())
        self.signals.log.emit("INFO", f"核销总数: {total_certs}")

        if missing_coupon_codes:
            code_set = set(missing_coupon_codes)
            sample_names = [
                str(product.get("productName", "")) or str(product.get("productCode", ""))
                for product in coupon_products
                if str(product.get("productCode", "")) in code_set
            ]
            sample_text = ", ".join(sample_names[:5])
            if len(sample_names) > 5:
                sample_text += " ..."
            self.signals.log.emit("WARN", f"未从发券扫描中找到模板ID: {sample_text}")

        if missing_template_products:
            sample_names = ", ".join(missing_template_products[:5])
            if len(missing_template_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"缺少模板ID，无法查询核销: {sample_names}")

        if missing_group_products:
            sample_names = ", ".join(missing_group_products[:5])
            if len(missing_group_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"券包全局扫描未匹配到记录: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            delivery_types = "/".join(_extract_delivery_types(product)) or "UNKNOWN"
            grant_count = int(grants_map.get(code, {}).get("count", 0) or 0)
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit(
                "INFO",
                f"[{i}/{len(products)}] {name} | 类型 {delivery_types} | 发券 {grant_count} | 核销 {cert_count}"
            )

        self.signals.log.emit("INFO", "正在生成 Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"券统计_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        total_grants = sum(int(value.get("count", 0) or 0) for value in grants_map.values())
        summary = (
            f"任务完成，文件已生成: {fname}\n\n"
            f"券种类: {len(products)}\n"
            f"发券总数: {total_grants}\n"
            f"核销总数: {sum(len(v) for v in certs_map.values())}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "所有任务已完成")

    except InterruptedError:
        self.signals.log.emit("WARN", "任务已停止")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()


TaskThread.build_excel = _taskthread_build_excel_v3
def _taskthread_run_v7(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP 登录失败")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "未查询到有效券")
            return

        grants_map, missing_grant_template_products = _query_coupon_grant_counts_by_template(
            self, icsp, products, s_full, e_full
        )
        certs_map = {str(product.get("productCode", "")): [] for product in products}

        coupon_codes, missing_spu_products = _query_coupon_cert_rows_by_spu(
            self, icsp, products, s_full, e_full, certs_map
        )

        package_rows_map, package_cert_rows_map, missing_group_products = _scan_coupon_package_rows(
            self, icsp, products, s_full, e_full
        )
        package_codes = set()
        for code, rows in package_rows_map.items():
            if rows:
                package_codes.add(code)
            grants_map[code] = {"count": len(rows), "rows": rows}
        certs_map.update(package_cert_rows_map)

        total_certs = sum(len(rows) for rows in certs_map.values())
        self.signals.log.emit("INFO", f"核销总数: {total_certs}")

        if missing_spu_products:
            sample_names = ", ".join(missing_spu_products[:5])
            if len(missing_spu_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"缺少 spuCode，无法查询普通券核销: {sample_names}")

        if missing_group_products:
            sample_names = ", ".join(missing_group_products[:5])
            if len(missing_group_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"券包全局扫描未匹配到记录: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            delivery_types = "/".join(_extract_delivery_types(product)) or "UNKNOWN"
            grant_count = int(grants_map.get(code, {}).get("count", 0) or 0)
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit(
                "INFO",
                f"[{i}/{len(products)}] {name} | 类型 {delivery_types} | 发券 {grant_count} | 核销 {cert_count}"
            )

        self.signals.log.emit("INFO", "正在生成 Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"券统计_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        total_grants = sum(int(value.get("count", 0) or 0) for value in grants_map.values())
        summary = (
            f"任务完成，文件已生成: {fname}\n\n"
            f"券种类: {len(products)}\n"
            f"发券总数: {total_grants}\n"
            f"核销总数: {sum(len(v) for v in certs_map.values())}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "所有任务已完成")

    except InterruptedError:
        self.signals.log.emit("WARN", "任务已停止")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()

def _taskthread_run_v8(self):
    try:
        s_full = f"{self.s} 00:00:00"
        e_full = f"{self.e} 23:59:59"

        icsp = ICSPClient(self.signals, self)
        if not icsp.login(self.icsp_u, self.icsp_p):
            self.signals.error.emit("ICSP 登录失败")
            return

        products = icsp.fetch_coupon_products(self.s, self.e)
        if not products:
            self.signals.log.emit("WARN", "未查询到有效券")
            return

        grants_map, _missing_grant_template_products = _query_coupon_grant_counts_by_template(
            self, icsp, products, s_full, e_full
        )
        certs_map = {str(product.get("productCode", "")): [] for product in products}

        _coupon_codes, missing_template_products = _query_coupon_cert_rows_by_template_id(
            self, icsp, products, s_full, e_full, certs_map
        )

        package_rows_map, package_cert_rows_map, missing_group_products = _scan_coupon_package_rows(
            self, icsp, products, s_full, e_full
        )
        for code, rows in package_rows_map.items():
            grants_map[code] = {"count": len(rows), "rows": rows}
        certs_map.update(package_cert_rows_map)

        total_certs = sum(len(rows) for rows in certs_map.values())
        self.signals.log.emit("INFO", f"核销总数: {total_certs}")

        if missing_group_products:
            sample_names = ", ".join(missing_group_products[:5])
            if len(missing_group_products) > 5:
                sample_names += " ..."
            self.signals.log.emit("WARN", f"券包全局扫描未匹配到记录: {sample_names}")

        for i, product in enumerate(products, 1):
            _taskthread_check_stop_safe(self)
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            delivery_types = "/".join(_extract_delivery_types(product)) or "UNKNOWN"
            grant_count = int(grants_map.get(code, {}).get("count", 0) or 0)
            cert_count = len(certs_map.get(code, []))
            self.signals.log.emit(
                "INFO",
                f"[{i}/{len(products)}] {name} | 类型 {delivery_types} | 发券 {grant_count} | 核销 {cert_count}"
            )

        self.signals.log.emit("INFO", "正在生成 Excel...")
        wb = self.build_excel(products, grants_map, certs_map)
        exe_dir = os.path.dirname(
            sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
        )
        fname = os.path.join(exe_dir, f"券统计_{self.s}_{self.e}.xlsx")
        wb.save(fname)

        total_grants = sum(int(value.get("count", 0) or 0) for value in grants_map.values())
        summary = (
            f"任务完成，文件已生成: {fname}\n\n"
            f"券种类: {len(products)}\n"
            f"发券总数: {total_grants}\n"
            f"核销总数: {sum(len(v) for v in certs_map.values())}"
        )
        self.signals.success.emit(summary)
        self.signals.log.emit("SUCCESS", "所有任务已完成")

    except InterruptedError:
        self.signals.log.emit("WARN", "任务已停止")
    except Exception as e:
        import traceback
        traceback.print_exc()
        self.signals.error.emit(str(e))
    finally:
        self.signals.finished.emit()


TaskThread.run = _taskthread_run_v8
TaskThread.check_stop_safe = _taskthread_check_stop_safe


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())
