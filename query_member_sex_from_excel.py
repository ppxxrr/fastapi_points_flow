import argparse
import os
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

import requests
from openpyxl import load_workbook


API_URL = "https://icsp.scpgroup.com.cn/icsp-member/web/member/baseInfo"

DEFAULT_HEADERS = {
    "accept": "*/*",
    "accept-language": "zh-CN",
    "content-type": "application/json",
    "origin": "https://icsp.scpgroup.com.cn",
    "priority": "u=1, i",
    "referer": "https://icsp.scpgroup.com.cn/scpg.html",
    "sec-ch-ua": '"Not:A-Brand";v="99", "Google Chrome";v="145", "Chromium";v="145"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36"
    ),
}

DEFAULT_COOKIES = {
    "sso_sessionid": "cc535e28fd184551a7a085806bad5caf",
    "wdjsid": "b811a4e2-fe38-416b-95c3-25bc81b780cb",
}

THREAD_LOCAL = threading.local()


def normalize_member_id(value: Any) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            decimal_value = Decimal(text)
        except InvalidOperation:
            return text
    else:
        try:
            decimal_value = Decimal(str(value))
        except InvalidOperation:
            text = str(value).strip()
            return text or None

    if decimal_value == decimal_value.to_integral_value():
        return str(decimal_value.quantize(Decimal("1")))
    return format(decimal_value.normalize(), "f").rstrip("0").rstrip(".")


def get_column_index(header_row: list[Any], target_names: tuple[str, ...]) -> int:
    normalized = {
        (str(cell).strip().lower() if cell is not None else ""): index
        for index, cell in enumerate(header_row, start=1)
    }
    for name in target_names:
        if name.lower() in normalized:
            return normalized[name.lower()]
    raise ValueError(f"Column not found: {', '.join(target_names)}")


def extract_sex(response_json: dict[str, Any]) -> Any:
    for container_key in ("data", "result", "rows"):
        container = response_json.get(container_key)
        if isinstance(container, dict) and "sex" in container:
            return container.get("sex")
        if isinstance(container, list) and container:
            first_item = container[0]
            if isinstance(first_item, dict) and "sex" in first_item:
                return first_item.get("sex")
    if "sex" in response_json:
        return response_json.get("sex")
    return None


def build_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_with_sex{input_path.suffix}")


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(DEFAULT_HEADERS)
    session.cookies.update(DEFAULT_COOKIES)
    return session


def get_session() -> requests.Session:
    session = getattr(THREAD_LOCAL, "session", None)
    if session is None:
        session = build_session()
        THREAD_LOCAL.session = session
    return session


def fetch_sex(member_id: str, timeout: float, sleep_seconds: float) -> tuple[Any, str]:
    try:
        response = get_session().post(API_URL, json={"memberId": member_id}, timeout=timeout)
        response.raise_for_status()
        payload = response.json()
        return extract_sex(payload), "ok"
    except Exception as exc:
        return "", str(exc)
    finally:
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)


def print_progress(
    done: int,
    total: int,
    ok_count: int,
    fail_count: int,
    skip_count: int,
    start_time: float,
) -> None:
    if total <= 0:
        return

    elapsed = max(time.perf_counter() - start_time, 0.001)
    speed = done / elapsed
    remaining = max(total - done, 0) / speed if speed > 0 else 0
    percent = done / total * 100
    message = (
        f"\rProgress: {done}/{total} ({percent:5.1f}%)"
        f" | ok={ok_count} fail={fail_count} skip={skip_count}"
        f" | {speed:5.1f} rows/s | eta={remaining:5.1f}s"
    )
    print(message, end="", flush=True)


def process_excel(
    input_file: Path,
    output_file: Path,
    sheet_name: Optional[str],
    memberid_column_name: str,
    timeout: float,
    sleep_seconds: float,
    workers: int,
) -> None:
    workbook = load_workbook(input_file)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    header_values = [cell.value for cell in sheet[1]]
    memberid_col = get_column_index(header_values, (memberid_column_name, "memberid", "memberId"))

    sex_col = sheet.max_column + 1
    status_col = sheet.max_column + 2
    sheet.cell(row=1, column=sex_col, value="sex")
    sheet.cell(row=1, column=status_col, value="query_status")

    total_rows = max(sheet.max_row - 1, 0)
    tasks: list[tuple[int, str]] = []
    ok_count = 0
    fail_count = 0
    skip_count = 0
    done_count = 0
    start_time = time.perf_counter()

    for row in range(2, sheet.max_row + 1):
        raw_member_id = sheet.cell(row=row, column=memberid_col).value
        member_id = normalize_member_id(raw_member_id)
        if not member_id:
            sheet.cell(row=row, column=sex_col, value="")
            sheet.cell(row=row, column=status_col, value="memberid empty")
            done_count += 1
            skip_count += 1
            continue
        tasks.append((row, member_id))

    print_progress(done_count, total_rows, ok_count, fail_count, skip_count, start_time)

    max_workers = max(1, min(workers, len(tasks) or 1))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(fetch_sex, member_id, timeout, sleep_seconds): row
            for row, member_id in tasks
        }

        for future in as_completed(future_map):
            row = future_map[future]
            sex, status = future.result()
            sheet.cell(row=row, column=sex_col, value=sex)
            sheet.cell(row=row, column=status_col, value=status)

            done_count += 1
            if status == "ok":
                ok_count += 1
            else:
                fail_count += 1
            print_progress(done_count, total_rows, ok_count, fail_count, skip_count, start_time)

    if total_rows > 0:
        print()

    workbook.save(output_file)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read memberid values from Excel, query sex, and write the result back."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        default="a.xlsx",
        help="Input Excel file path. Default: a.xlsx in the current folder.",
    )
    parser.add_argument(
        "-o",
        "--output-file",
        help="Output Excel file path. Default: <input>_with_sex.xlsx",
    )
    parser.add_argument(
        "-s",
        "--sheet-name",
        help="Sheet name. Default: the active sheet.",
    )
    parser.add_argument(
        "-c",
        "--column-name",
        default="memberid",
        help="memberid column name. Default: memberid",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=15,
        help="Request timeout in seconds. Default: 15",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.0,
        help="Delay after each request in seconds. Default: 0",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=min(16, max(4, (os.cpu_count() or 4) * 2)),
        help="Number of concurrent workers. Default: auto",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_file = Path(args.input_file).expanduser().resolve()
    output_file = (
        Path(args.output_file).expanduser().resolve()
        if args.output_file
        else build_output_path(input_file)
    )

    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    process_excel(
        input_file=input_file,
        output_file=output_file,
        sheet_name=args.sheet_name,
        memberid_column_name=args.column_name,
        timeout=args.timeout,
        sleep_seconds=args.sleep,
        workers=args.workers,
    )
    print(f"Done. Output file: {output_file}")


if __name__ == "__main__":
    main()
