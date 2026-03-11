#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
import re
import csv
import requests
import time
import base64
import hashlib
import urllib.parse
import math
import json
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from PySide6.QtCore import Qt, Signal, QObject, QThread, QPointF, QDate, QEvent
from PySide6.QtGui import QColor, QPixmap, QCursor, QMouseEvent, QIcon
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QFrame, QTextEdit,
    QGraphicsDropShadowEffect, QDialog, QDateEdit, QSizePolicy,
    QMessageBox
)

# ====== 🎨 UI 核心配色配置 (强制亮色模式) ======
COLOR_BG = "#F5F5F7"
COLOR_CARD = "#FFFFFF"
COLOR_PRIMARY = "#007AFF"
COLOR_STOP = "#FF3B30"
COLOR_ACCENT_BG = "#E5F0FF"
COLOR_TEXT_H1 = "#1D1D1F"
COLOR_TEXT_H2 = "#86868B"
COLOR_LOG_BG = "#282C34"
COLOR_LOG_TEXT = "#ABB2BF"
COLOR_BORDER = "#E5E5EA"

DIM_W, DIM_H = 480, 850
RAD_CARD = 20
RAD_INPUT = 12
FONT_FAMILY = "Microsoft YaHei UI"

# 全局样式表
GLOBAL_STYLES = f"""
    QWidget {{
        color: {COLOR_TEXT_H1};
        font-family: "{FONT_FAMILY}";
    }}
    QMainWindow, QDialog {{
        background-color: {COLOR_BG};
    }}
    QLineEdit, QDateEdit, QTextEdit {{
        background-color: {COLOR_CARD};
        color: {COLOR_TEXT_H1};
        border: 1px solid {COLOR_BORDER};
        border-radius: {RAD_INPUT}px;
        selection-background-color: {COLOR_PRIMARY};
        selection-color: #FFFFFF;
    }}
    QLineEdit:disabled, QDateEdit:disabled {{
        background-color: #F0F0F5;
        color: #AAAAAA;
    }}
    QLabel {{
        background: transparent;
    }}
"""

# ====== 🛠️ 业务配置 ======
ICSP_CLIENT_ID = "2a5c64fcf8cf475593350a6d11548711"
ICSP_SALT = "d0a8155e8e84e5832c3a908056737c2b"
PLAZA_ID = "G002Z008C0030"
TENANT_ID = "10000"
ORG_TYPE_CODE = "10003"
PAGE_SIZE = 100
MAX_WORKERS = 10
SAFE_BATCH_SIZE = 5

PARKING_LOGIN_URL = "https://szwryportal.aibee.cn/union/portal/loginV2"
PARKING_CAPTCHA_URL = "https://szwryportal.aibee.cn/union/portal/getKaptcha"
PARKING_DATA_URL = "https://szwryportal.aibee.cn/parkingbi/api/backend/reservation/search_reservation"
PARKING_USER = "railinadmin"
PARKING_PWD_RAW = "pemdot-9sudxi-kAzcyw"

# ====== 📋 字段映射 ======
HEADER_PAY = {
    "tradeNo": "交易流水号", "outTradeNo": "商户订单号", "totalAmount": "交易金额(分)",
    "title": "商品说明", "tradeStatus": "交易状态", "payTime": "支付时间",
    "payReqNo": "支付请求号", "memberId": "会员ID", "merchantId": "商户号", "createTime": "创建时间"
}
HEADER_REFUND = {
    "refundNo": "退款单号", "outRefundNo": "商户退款单号", "outTradeNo": "原交易单号",
    "refundAmount": "退款金额(分)", "refundStatus": "退款状态", "refundTime": "退款时间",
    "refundCause": "退款原因", "createTime": "申请时间", "remark": "备注"
}
HEADER_MALL = {
    "orderCode": "商城订单号", "orderId": "订单ID", "spu": "商品名称", "amount": "实付金额(分)",
    "point": "消耗积分", "exchangeNum": "数量", "status": "订单状态", "mobilePhone": "会员手机",
    "exchangeTime": "兑换时间", "paySeqNo": "关联支付单号", "deliveryType": "交付类型", "remark": "备注"
}
HEADER_MALL_REFUND = {
    "refundNo": "退款单号", "refundSeqNo": "退款流水号", "orderNo": "关联订单号",
    "spu": "商品名称", "spuCode": "商品编码", "refundAmount": "退款金额(分)",
    "refundPoint": "退还积分", "refundStatus": "退款状态", "mobilePhone": "会员手机",
    "createTime": "申请时间", "successTime": "成功时间", "refundCause": "退款原因",
    "applicant": "申请人", "plazza": "所属广场", "bankType": "银行类型", "remark": "备注"
}
HEADER_PARKING = {
    "id": "ID", "createdAt": "创建时间", "site_id": "场站ID", "floor": "楼层",
    "reservation_number": "预约单号", "reservation_people": "预约人", "phone": "手机号",
    "member_type": "会员类型", "car_plate": "车牌号", "parking_space": "车位号",
    "order_time": "下单时间", "begin_time": "预约开始时间", "arrival_time": "入场时间",
    "current_state": "当前状态", "payment_amount": "支付金额"
}
HEADER_RENT = {
    "exchangeRecordNo": "兑换记录号", "orderCode": "订单号", "memberId": "会员ID",
    "phone": "手机号", "spu": "物品名称", "spuCode": "物品编码",
    "status": "状态", "depositPriceStr": "押金金额", "exchangeTime": "兑换时间",
    "payTime": "支付时间", "returnTime": "归还时间", "receiveAddress": "借用地点",
    "channelTradeNo": "渠道交易号", "tradeNo": "交易流水号", "payReqNo": "支付请求号"
}
# 活动报名记录 (已修改：新增 applyStatus)
HEADER_ACTIVITY = {
    "id": "ID", "activityName": "活动名称", "applyTime": "报名时间",
    "memberName": "会员姓名", "mobileNo": "手机号", "costCash": "实付金额",
    "sectionName": "活动场次", "sectionTimeStr": "场次时间",
    "activityOrderNo": "活动订单号", "applyCode": "核销码",
    "applyStatus": "报名状态",  # 新增字段
    "tradeOrderNo": "支付订单号(新增)"
}


# ====== 信号定义 ======
class WorkerSignals(QObject):
    log = Signal(str, str)
    finished = Signal()
    error = Signal(str)
    success = Signal(str)
    stopped = Signal()
    captcha_received = Signal(str, bytes)


# ====== ICSP 客户端 ======
class ICSPClient:
    def __init__(self, signals, thread_ref=None):
        self.session = requests.Session()
        self.base = "https://icsp.scpgroup.com.cn"
        self.signals = signals
        self.thread_ref = thread_ref
        self.client_id = ICSP_CLIENT_ID
        self.salt = ICSP_SALT
        self.user_info = {"userid": "", "usercode": "", "username": ""}
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/json, text/plain, */*",
            "Origin": self.base,
            "Referer": self.base + "/login.html"
        })

    def check_stop(self):
        if self.thread_ref and self.thread_ref.is_interrupted:
            raise InterruptedError("用户手动停止任务")

    def log(self, level, msg):
        self.signals.log.emit(level, msg)

    def make_passwd(self, pwd):
        combined = (self.salt + pwd).encode("utf-8")
        b64 = base64.b64encode(combined).decode()
        return f"{b64}.{self.salt}"

    def login(self, user, pwd):
        self.check_stop()
        passwd_value = self.make_passwd(pwd)
        form = {"clientId": self.client_id, "passwd": passwd_value, "user": user}
        headers = {"Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}
        self.log("INFO", f"[ICSP] 正在登录: {user}...")
        try:
            resp = self.session.post(
                f"{self.base}/icsp-permission/web/permission/sso/auth/authCode",
                data=form, headers=headers, allow_redirects=False, timeout=10
            )
            auth_code = ""
            if resp.status_code == 302 and "Location" in resp.headers:
                auth_code = resp.headers["Location"].split("authCode=")[-1]
            elif resp.status_code == 200:
                js = resp.json()
                if js.get("success") and js.get("data"): auth_code = js["data"]

            if not auth_code: raise RuntimeError("AuthCode获取失败")

            ts = str(int(time.time() * 1000))
            self.session.get(f"{self.base}/auth.html?authCode={auth_code}", timeout=10)
            self.session.get(f"{self.base}/icsp-permission/web/wd/login/login/sso?_t={ts}&authCode={auth_code}",
                             timeout=10)

            user_resp = self.session.get(f"{self.base}/icsp-employee/web/login/query/v2?_t={ts}", timeout=10)
            if user_resp.status_code == 200:
                u_data = user_resp.json().get("data", {})
                self.user_info["userid"] = str(u_data.get("id", ""))
                self.user_info["usercode"] = str(u_data.get("loginCode", user))
                self.user_info["username"] = urllib.parse.quote(u_data.get("userName", ""))

            self.log("SUCCESS", "[ICSP] 登录成功")
            return True
        except Exception as e:
            self.log("ERROR", f"[ICSP] 登录异常: {str(e)}")
            return False

    def get_api_headers(self, is_json=False):
        h = {
            "plazacode": PLAZA_ID, "orgcode": PLAZA_ID, "orgtypecode": ORG_TYPE_CODE,
            "tenantid": TENANT_ID, "groupcode": "G001", "internalid": "1",
            "vunioncode": "U001", "workingorgcode": PLAZA_ID,
            "userid": self.user_info["userid"], "usercode": self.user_info["usercode"],
            "username": self.user_info["username"]
        }
        if is_json: h["Content-Type"] = "application/json;charset=utf-8"
        return h

    def extract_rows_smart(self, data, context=""):
        rows, total = [], 0
        if not data: return [], 0
        if isinstance(data, dict):
            if "status" in data and str(data["status"]) == "5000":
                return [], 0

            # 1. 尝试直接在顶层查找常见列表键
            # 增加 resultList 支持
            keys_to_check = ["rows", "data", "list", "result", "records", "content", "items", "resultList"]

            for key in keys_to_check:
                if key in data and isinstance(data[key], list):
                    rows = data[key];
                    break

            # 2. 如果没找到，尝试进入 data 字段（嵌套结构）查找
            if not rows and "data" in data and isinstance(data["data"], dict):
                sub = data["data"]
                for key in keys_to_check:
                    if key in sub and isinstance(sub[key], list):
                        rows = sub[key];
                        break

                if "total" in sub:
                    total = sub["total"]
                elif "totalCount" in sub:
                    total = sub["totalCount"]
                elif "totalSize" in sub:
                    total = sub["totalSize"]

            # 如果内层没找到总数，检查外层
            if total == 0:
                if "total" in data:
                    total = data["total"]
                elif "totalCount" in data:
                    total = data["totalCount"]
                elif "totalSize" in data:
                    total = data["totalSize"]

        elif isinstance(data, list):
            rows = data
        return rows, total

    def fetch_data_hybrid(self, url, method, start, end, name, payload_template=None):
        all_rows = []
        lock = threading.Lock()
        self.log("INFO", f"启动抓取 [{name}]...")

        first_page_data, total = self._fetch_one_page(url, method, start, end, 1, 0, payload_template)
        if not first_page_data:
            self.log("WARN", f"[{name}] 无数据")
            return []

        all_rows.extend(first_page_data)

        if total > PAGE_SIZE:
            self.log("INFO", f"[{name}] 总数 {total}，启动全速并发...")
            self._fetch_known_total(url, method, start, end, name, payload_template, total, all_rows, lock)
        elif len(first_page_data) >= PAGE_SIZE:
            self.log("INFO", f"[{name}] 总数未知，启用安全批量并发(x{SAFE_BATCH_SIZE})...")
            self._fetch_safe_batch(url, method, start, end, name, payload_template, all_rows, lock)
        else:
            self.log("INFO", f"[{name}] 仅1页数据，完成")

        self.log("SUCCESS", f"[{name}] 抓取完成，共 {len(all_rows)} 条")
        return all_rows

    def _fetch_known_total(self, url, method, start, end, name, payload_template, total, all_rows, lock):
        tasks = []
        if method == "POST" or method == "GET_RENT":
            total_pages = math.ceil(total / PAGE_SIZE)
            for p in range(2, total_pages + 1):
                tasks.append((p, 0))
        else:
            offsets = range(PAGE_SIZE, total, PAGE_SIZE)
            for off in offsets:
                tasks.append((1, off))

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {
                executor.submit(self._fetch_one_page, url, method, start, end, p, off, payload_template): (p, off) for
                p, off in tasks}
            for future in as_completed(futures):
                rows, _ = future.result()
                if rows:
                    with lock:
                        all_rows.extend(rows)

    def _fetch_safe_batch(self, url, method, start, end, name, payload_template, all_rows, lock):
        current_page_base = 2
        while True:
            self.check_stop()
            futures_map = {}
            with ThreadPoolExecutor(max_workers=SAFE_BATCH_SIZE) as executor:
                for i in range(SAFE_BATCH_SIZE):
                    p_idx = current_page_base + i
                    off = (p_idx - 1) * PAGE_SIZE
                    if method == "POST" or method == "GET_RENT":
                        fut = executor.submit(self._fetch_one_page, url, method, start, end, p_idx, 0, payload_template)
                    else:
                        fut = executor.submit(self._fetch_one_page, url, method, start, end, 1, off, payload_template)
                    futures_map[fut] = p_idx

            batch_results = []
            for fut in as_completed(futures_map):
                p_idx = futures_map[fut]
                try:
                    rows, _ = fut.result()
                    batch_results.append((p_idx, rows))
                except Exception:
                    batch_results.append((p_idx, []))

            batch_results.sort(key=lambda x: x[0])
            should_stop = False
            for _, rows in batch_results:
                if rows:
                    with lock:
                        all_rows.extend(rows)
                    if len(rows) < PAGE_SIZE: should_stop = True
                else:
                    should_stop = True

            if should_stop: break
            current_page_base += SAFE_BATCH_SIZE
            self.log("INFO", f"[{name}] 已抓取 {len(all_rows)} 条...")
            if current_page_base > 1000: break

    def _fetch_one_page(self, url, method, start, end, page_index, offset, payload_template):
        timestr = str(int(time.time() * 1000))
        try:
            if method == "GET":
                s_simple = start.split(" ")[0]
                e_simple = end.split(" ")[0]
                params = {"plazaId": PLAZA_ID, "startDate": s_simple, "endDate": e_simple, "limit": PAGE_SIZE,
                          "offset": offset, "timestr": timestr}
                resp = self.session.get(url, headers=self.get_api_headers(), params=params, timeout=15)
                return self.extract_rows_smart(resp.json())
            elif method == "GET_RENT":
                # 租借服务订单 GET 请求专用
                headers = self.get_api_headers()
                headers["orgname"] = urllib.parse.quote("深圳湾睿印RAIL IN")
                headers["Referer"] = self.base + "/scpg.html"

                params = {
                    "goodsCode": "", "goodsName": "", "plazzaCode": PLAZA_ID,
                    "mobileNo": "", "id": "", "status": "", "userName": "", "exchangeRecordNo": "",
                    "totalSize": "", "channelTradeNo": "", "payReqNo": "", "tradeNo": "",
                    "createStartTime": start, "createEndTime": end,
                    "pageIndex": page_index, "pageSize": PAGE_SIZE, "timestr": timestr
                }
                resp = self.session.get(url, headers=headers, params=params, timeout=15)
                rows, total = self.extract_rows_smart(resp.json())
                return rows, total
            else:
                # POST 处理
                payload = payload_template.copy()
                is_mall_order = "recordSearch" in url
                is_activity = "applyList" in url

                headers = self.get_api_headers(True)

                if is_activity:
                    # 活动报名接口专用逻辑
                    payload["applyTimeBegin"] = f"{start} 00:00"
                    payload["applyTimeEnd"] = f"{end} 00:00"
                    # 清理通用字段
                    payload.pop("orderDates", None)
                    payload.pop("timeBegin", None)
                    payload.pop("timeEnd", None)
                    headers["Referer"] = "https://icsp.scpgroup.com.cn/adapter.html"
                else:
                    # 通用逻辑
                    s_val = start.split(" ")[0] if is_mall_order else start
                    e_val = end.split(" ")[0] if is_mall_order else end
                    payload["orderDates"] = [s_val, e_val]
                    payload["timeBegin"] = s_val
                    payload["timeEnd"] = e_val
                    if "refund" in url:
                        payload["applyDates"] = [start, end]

                payload["pageIndex"] = page_index
                payload["pageSize"] = PAGE_SIZE

                resp = self.session.post(url, headers=headers, json=payload, timeout=15)

                try:
                    res_json = resp.json()
                    return self.extract_rows_smart(res_json)
                except Exception:
                    self.log("WARN", f"解析失败: {resp.text[:200]}")
                    return [], 0

        except Exception as e:
            self.log("WARN", f"请求异常: {str(e)}")
            return [], 0

    # 新增：增强活动数据（查询详情获取支付单号）
    def enrich_activity_data(self, rows):
        if not rows: return
        self.log("INFO", f"正在查询 {len(rows)} 条活动详情以获取支付订单号...")

        def fetch_detail(row):
            apply_id = row.get("id")
            if not apply_id: return

            ts = str(int(time.time() * 1000))
            url = f"{self.base}/yinli-xapi-b/activity/v1/activity/applyDetail?applyId={apply_id}&timestr={ts}"
            try:
                resp = self.session.get(url, headers=self.get_api_headers(), timeout=10)
                d = resp.json()

                if "data" in d and isinstance(d["data"], dict):
                    data_obj = d["data"]
                else:
                    data_obj = d

                act_order = data_obj.get("activityOrder", {})
                trade_no = act_order.get("tradeOrderNo", "")

                row["tradeOrderNo"] = trade_no
            except Exception as e:
                pass

        with ThreadPoolExecutor(max_workers=MAX_WORKERS * 2) as executor:
            list(executor.map(fetch_detail, rows))

        self.log("SUCCESS", "支付订单号获取完成")


# ====== 停车系统客户端 (保持不变) ======
class ParkingClient:
    def __init__(self):
        self.session = requests.Session()
        self.signals = None
        self.base_headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/144.0.0.0 Safari/537.36 Edg/144.0.0.0',
            'Referer': 'https://szwryportal.aibee.cn/micro/parking/dashboard/detail/BackgroundReservation',
            'Origin': 'https://szwryportal.aibee.cn',
            'Content-Type': 'application/json;charset=UTF-8'
        }
        self.session.headers.update(self.base_headers)

    def set_signals(self, signals):
        self.signals = signals

    def log(self, level, msg):
        if self.signals:
            self.signals.log.emit(level, msg)
        else:
            print(f"[{level}] {msg}")

    def md5_encrypt(self, text):
        return hashlib.md5(text.encode('utf-8')).hexdigest()

    def fetch_captcha(self):
        try:
            resp = self.session.get(PARKING_CAPTCHA_URL, timeout=10)
            data = resp.json()
            return data.get('requestId'), base64.b64decode(data.get('data'))
        except Exception:
            return None, None

    def login(self, captcha_code, uuid):
        self.log("INFO", "[停车] 正在登录...")
        payload = {
            'username': PARKING_USER, 'password': self.md5_encrypt(PARKING_PWD_RAW),
            'kaptcha': captcha_code, 'code': captcha_code, 'uuid': uuid, 'requestId': uuid
        }
        try:
            resp = self.session.post(PARKING_LOGIN_URL, json=payload, timeout=10)
            res = resp.json()
            if res.get('status') == 200:
                token = res.get('data', {}).get('access_token')
                self.session.headers.update({
                    'Authorization': f'Bearer {token}',
                    'x-app-id': 'parkingbi', 'x-brand-id': 'railin', 'x-project-id': 'railin_shenzhen_vip',
                    'accept': 'application/json, text/plain, */*'
                })
                self.log("SUCCESS", "[停车] 登录成功")
                return True
            else:
                self.log("ERROR", f"[停车] 登录失败: {res.get('message')}")
                return False
        except Exception as e:
            self.log("ERROR", f"[停车] 登录异常: {e}");
            return False

    def _fetch_page(self, page_num, start_date, end_date):
        payload = {
            "entity_id": "", "page_offset": page_num, "page_size": PAGE_SIZE,
            "filter_body": {
                "start_order_time": start_date, "end_order_time": end_date,
                "reservation_number": "", "reservation_people": "", "phone": "",
                "car_plate": "", "parking_space": "", "start_arrival_time": "",
                "end_arrival_time": "", "order_status": "", "order_type": ""
            }
        }
        try:
            resp = self.session.post(PARKING_DATA_URL, json=payload, timeout=20)
            data_json = resp.json()
            if isinstance(data_json.get('data'), dict):
                d = data_json.get('data')
                return (d.get('list') or d.get('records') or []), d.get('total', 0)
            return [], 0
        except Exception:
            return [], 0

    def fetch_all_orders_parallel(self, start_date, end_date):
        s_full = f"{start_date} 00:00:00" if " " not in start_date else start_date
        e_full = f"{end_date} 23:59:59" if " " not in end_date else end_date

        self.log("INFO", f"启动并发抓取 [预约停车订单]...")

        rows, total = self._fetch_page(1, s_full, e_full)
        if not rows:
            self.log("WARN", "[停车] 无数据")
            return []

        all_records = list(rows)
        self.log("INFO", f"[停车] 首页获取 {len(rows)} 条，总数: {total}")

        if total <= PAGE_SIZE: return all_records

        total_pages = math.ceil(total / PAGE_SIZE)
        self.log("INFO", f"[停车] 启动并发: 共 {total_pages} 页")

        futures = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for p in range(2, total_pages + 1):
                futures.append(executor.submit(self._fetch_page, p, s_full, e_full))

            for future in as_completed(futures):
                r, _ = future.result()
                if r: all_records.extend(r)

        self.log("SUCCESS", f"[停车] 抓取完成，共 {len(all_records)} 条")
        return all_records


# ====== 数据清洗辅助函数 ======

def _extract_date(val):
    """从各种日期格式中提取 yyyy-MM-dd"""
    if val is None:
        return None
    if isinstance(val, (int, float)) and val > 1000000000000:
        try:
            return datetime.fromtimestamp(val / 1000).strftime('%Y-%m-%d')
        except Exception:
            return None
    s = str(val).strip()
    if len(s) >= 10:
        return s[:10]
    return None


def _date_range(start_str, end_str):
    """生成日期列表 [start, ..., end]"""
    s = datetime.strptime(start_str, "%Y-%m-%d")
    e = datetime.strptime(end_str, "%Y-%m-%d")
    dates = []
    cur = s
    while cur <= e:
        dates.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return dates


def _safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _detect_encoding(path):
    """检测文件编码，优先 utf-8-sig，失败则用 gbk"""
    for enc in ('utf-8-sig', 'gbk', 'gb18030'):
        try:
            with open(path, 'r', encoding=enc) as f:
                f.read(1024)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'utf-8-sig'


def parse_wechat_fund_csv(path, log_fn=None):
    """解析微信支付资金账单.csv，返回逐行 list[dict] 和按日+业务类型汇总、退款手续费汇总"""
    if not os.path.isfile(path):
        if log_fn:
            log_fn("WARN", f"未找到资金账单CSV: {path}，跳过")
        return [], {}, {}

    rows = []
    with open(path, 'r', encoding=_detect_encoding(path)) as f:
        lines = f.readlines()

    # 找到表头行（以 记账时间 开头）
    header_idx = None
    for i, line in enumerate(lines):
        stripped = line.strip().lstrip('`').strip()
        if stripped.startswith('记账时间'):
            header_idx = i
            break

    if header_idx is None:
        if log_fn:
            log_fn("WARN", "资金账单CSV格式异常：未找到表头行")
        return [], {}, {}

    # 用 csv reader 解析从表头开始的内容
    clean_lines = []
    for line in lines[header_idx:]:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        clean_lines.append(line)

    reader = csv.reader(clean_lines)
    headers = None
    for row_data in reader:
        # strip ` prefix from each cell
        cleaned = [c.strip().lstrip('`').strip() for c in row_data]
        if headers is None:
            headers = cleaned
            continue
        if len(cleaned) < len(headers):
            continue
        row_dict = dict(zip(headers, cleaned))
        rows.append(row_dict)

    # 按日+业务类型+收支类型汇总
    # 同时提取退款手续费
    daily = defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'refund_fee': 0.0}))
    refund_fee_daily = defaultdict(float)

    for r in rows:
        date = _extract_date(r.get('记账时间', ''))
        if not date:
            continue
        biz_type = r.get('业务类型', '').strip()
        income_type = r.get('收支类型', '').strip()
        amount = _safe_float(r.get('收支金额(元)', '0'))

        key = (biz_type, income_type)
        daily[date][key]['amount'] += amount

        # 从备注提取退款手续费
        remark = r.get('备注', '')
        fee_match = re.search(r'含手续费(\d+\.?\d*)元', remark)
        if fee_match:
            fee = float(fee_match.group(1))
            daily[date][key]['refund_fee'] += fee
            refund_fee_daily[date] += fee

    if log_fn:
        log_fn("SUCCESS", f"[资金账单] 解析完成，共 {len(rows)} 条记录")
    return rows, dict(daily), dict(refund_fee_daily)


def parse_wechat_trade_csv(path, log_fn=None):
    """解析微信支付交易订单-服务商.csv，返回按日汇总 dict"""
    if not os.path.isfile(path):
        if log_fn:
            log_fn("WARN", f"未找到交易订单CSV: {path}，跳过")
        return {}

    rows = []
    with open(path, 'r', encoding=_detect_encoding(path)) as f:
        lines = f.readlines()

    header_idx = None
    for i, line in enumerate(lines):
        stripped = line.strip().lstrip('`').strip()
        if stripped.startswith('交易时间'):
            header_idx = i
            break

    if header_idx is None:
        if log_fn:
            log_fn("WARN", "交易订单CSV格式异常：未找到表头行")
        return {}

    clean_lines = []
    for line in lines[header_idx:]:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        clean_lines.append(line)

    reader = csv.reader(clean_lines)
    headers = None
    for row_data in reader:
        cleaned = [c.strip().lstrip('`').strip() for c in row_data]
        if headers is None:
            headers = cleaned
            continue
        if len(cleaned) < len(headers):
            continue
        row_dict = dict(zip(headers, cleaned))
        rows.append(row_dict)

    # 跳过最后的汇总行（通常以"总计"开头或交易时间为空）
    filtered = [r for r in rows if r.get('交易时间', '').strip() and not r.get('交易时间', '').startswith('总')]

    daily = defaultdict(float)
    for r in filtered:
        date = _extract_date(r.get('交易时间', ''))
        if not date:
            continue
        # 取应结订单金额
        amount = _safe_float(r.get('应结订单金额', '0'))
        daily[date] += amount

    if log_fn:
        log_fn("SUCCESS", f"[交易订单-服务商] 解析完成，共 {len(filtered)} 条记录")
    return dict(daily)


def aggregate_pay_daily(rows):
    """支付订单按日汇总: payTime -> sum(totalAmount)/100，仅统计 TRADE_FINISHED 和 TRADE_SUCCESS"""
    daily = defaultdict(float)
    for r in rows:
        status = str(r.get('tradeStatus', '')).strip()
        if status not in ('TRADE_FINISHED', 'TRADE_SUCCESS'):
            continue
        date = _extract_date(r.get('payTime'))
        if not date:
            continue
        daily[date] += _safe_float(r.get('totalAmount', 0)) / 100
    return dict(daily)


def aggregate_refund_daily(rows):
    """退款订单按日汇总: refundTime -> sum(refundAmount)/100"""
    daily = defaultdict(float)
    for r in rows:
        date = _extract_date(r.get('refundTime'))
        if not date:
            continue
        daily[date] += _safe_float(r.get('refundAmount', 0)) / 100
    return dict(daily)


def aggregate_mall_orders(rows):
    """商城订单: 过滤amount>0, 按(日期,商品)汇总 和 按商品汇总"""
    filtered = [r for r in rows if _safe_float(r.get('amount', 0)) > 0 and r.get('status', '') != 'FAIL']
    daily_detail = defaultdict(lambda: {'qty': 0, 'amount': 0.0})
    product_summary = defaultdict(lambda: {'qty': 0, 'amount': 0.0})

    for r in filtered:
        date = _extract_date(r.get('exchangeTime'))
        if not date:
            continue
        spu = str(r.get('spu', '')).strip()
        qty = int(_safe_float(r.get('exchangeNum', 1)))
        amt = _safe_float(r.get('amount', 0)) / 100 * qty

        daily_detail[(date, spu)]['qty'] += qty
        daily_detail[(date, spu)]['amount'] += amt
        product_summary[spu]['qty'] += qty
        product_summary[spu]['amount'] += amt

    return dict(daily_detail), dict(product_summary)


def aggregate_mall_refunds(rows):
    """商城退款: 按(successTime日期, 商品)汇总 和 按商品汇总"""
    daily_detail = defaultdict(float)
    product_summary = defaultdict(float)

    for r in rows:
        date = _extract_date(r.get('successTime'))
        if not date:
            continue
        spu = str(r.get('spu', '')).strip()
        amt = _safe_float(r.get('refundAmount', 0)) / 100

        daily_detail[(date, spu)] += amt
        product_summary[spu] += amt

    return dict(daily_detail), dict(product_summary)


def aggregate_rent_daily(rows):
    """租借订单按日汇总: depositPriceStr 解析数字"""
    daily = defaultdict(float)
    for r in rows:
        pay_time = r.get('payTime')
        if not pay_time or str(pay_time).strip() == '':
            continue
        date = _extract_date(r.get('exchangeTime'))
        if not date:
            continue
        deposit_str = str(r.get('depositPriceStr', '0'))
        # 去掉"元"等非数字字符
        deposit_str = re.sub(r'[^\d.]', '', deposit_str)
        daily[date] += _safe_float(deposit_str)
    return dict(daily)


def aggregate_activity_daily(rows):
    """活动报名按日汇总: costCash"""
    daily = defaultdict(float)
    for r in rows:
        date = _extract_date(r.get('applyTime'))
        if not date:
            continue
        daily[date] += _safe_float(r.get('costCash', 0))
    return dict(daily)


def aggregate_parking_daily(rows):
    """停车订单按日汇总: payment_amount (None视为0)，仅统计当前状态为已完成"""
    daily = defaultdict(float)
    for r in rows:
        if r.get('current_state', '') != '已完成':
            continue
        date = _extract_date(r.get('order_time'))
        if not date:
            continue
        daily[date] += _safe_float(r.get('payment_amount', 0))
    return dict(daily)


# ====== 3. 后台线程 ======
class TaskThread(QThread):
    def __init__(self, icsp_u, icsp_p, s_date, e_date, parking_captcha, parking_uuid, parking_client, signals):
        super().__init__()
        self.icsp_u, self.icsp_p = icsp_u, icsp_p
        self.s, self.e = s_date, e_date
        self.p_captcha, self.p_uuid = parking_captcha, parking_uuid
        self.p_client = parking_client
        self.signals = signals
        self.is_interrupted = False

    def stop(self):
        self.is_interrupted = True

    def format_cell(self, key, value):
        if value is None: return ""
        if ("Time" in key or "time" in key) and isinstance(value, (int, float)) and value > 1000000000000:
            try:
                return datetime.fromtimestamp(value / 1000).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                pass
        return str(value)

    def save_sheet(self, wb, sheet_name, data, header_map):
        ws = wb.create_sheet(sheet_name)
        db_keys = list(header_map.keys())
        cn_headers = list(header_map.values())
        ws.append(cn_headers)
        if not data:
            ws.append(["无数据"])
            return
        for row_dict in data:
            row_val = []
            for k in db_keys: row_val.append(self.format_cell(k, row_dict.get(k)))
            ws.append(row_val)
        for i, col in enumerate(ws.columns):
            width = 16
            if i < len(cn_headers) and (
                    "名称" in cn_headers[i] or "单号" in cn_headers[i] or "时间" in cn_headers[i]): width = 24
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

    def save_clean_sheet(self, wb, sheet_name, headers, rows):
        """通用：写入清洗后的 sheet，含汇总行、边框、自动列宽"""
        ws = wb.create_sheet(sheet_name)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal='center', vertical='center')
        data_align = Alignment(horizontal='center', vertical='center')
        sum_font = Font(bold=True, size=11)
        sum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # 写表头
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 写数据行
        if not rows:
            ws.append(["无数据"])
        else:
            for r in rows:
                ws.append(r)
            # 汇总行
            sum_row = ["合计"]
            for col_idx in range(1, len(headers)):
                vals = []
                for r in rows:
                    v = r[col_idx] if col_idx < len(r) else None
                    if isinstance(v, (int, float)):
                        vals.append(v)
                if vals:
                    sum_row.append(round(sum(vals), 2))
                else:
                    sum_row.append("")
            ws.append(sum_row)

        # 设置数据行和汇总行样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = data_align
        # 汇总行加粗+底色
        if rows:
            for cell in ws[ws.max_row]:
                cell.font = sum_font
                cell.fill = sum_fill

        # 自动列宽
        for i, header in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            max_len = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=i, max_col=i):
                for cell in row:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    if val_len > max_len:
                        max_len = val_len
            ws.column_dimensions[col_letter].width = max_len * 1.5 + 4

    def run(self):
        try:
            parking_rows = []
            icsp_data = {}

            s_full = f"{self.s} 00:00:00"
            e_full = f"{self.e} 23:59:59"

            if self.p_captcha and self.p_uuid:
                self.p_client.set_signals(self.signals)
                if self.p_client.login(self.p_captcha, self.p_uuid):
                    parking_rows = self.p_client.fetch_all_orders_parallel(s_full, e_full)
            else:
                self.signals.log.emit("WARN", "未输入验证码，跳过停车系统")

            if self.icsp_u and self.icsp_p:
                icsp_client = ICSPClient(self.signals, self)
                if icsp_client.login(self.icsp_u, self.icsp_p):
                    icsp_data["pay"] = icsp_client.fetch_data_hybrid(
                        f"{icsp_client.base}/yinli-xapi-b/pay/pay/bp/orders", "GET", s_full, e_full, "支付订单"
                    )
                    icsp_data["refund"] = icsp_client.fetch_data_hybrid(
                        f"{icsp_client.base}/yinli-xapi-b/pay/pay/bp/refund/orders", "GET", s_full, e_full, "支付退款"
                    )
                    mall_pl = {"orgCode": PLAZA_ID, "orderStatus": "ALL", "orderStatusArr": ["ALL"],
                               "orgTypeCode": ORG_TYPE_CODE, "orderNoType": 1, "payChannel": "ALL",
                               "deliveryType": "ALL"}
                    icsp_data["mall"] = icsp_client.fetch_data_hybrid(
                        f"{icsp_client.base}/yinli-xapi-b/pmp/pointmall/platform/exchange/recordSearch", "POST", s_full,
                        e_full, "商城订单", mall_pl
                    )
                    mall_ref_pl = {"orgCode": PLAZA_ID, "mobilePhone": "", "refundStatus": "", "productCode": "",
                                   "productName": "", "refundNo": "", "exchangeRecordNo": "", "refundSeqNo": ""}
                    icsp_data["mall_refund"] = icsp_client.fetch_data_hybrid(
                        f"{icsp_client.base}/yinli-xapi-b/pmp/pointmall/platform/refund/list", "POST", s_full, e_full,
                        "商城退款", mall_ref_pl
                    )
                    icsp_data["rent"] = icsp_client.fetch_data_hybrid(
                        f"{icsp_client.base}/yinli-xapi-b/pmp/pointmall/rent/searchRentRecord", "GET_RENT", s_full,
                        e_full, "租借服务订单"
                    )
                    activity_pl = {"orgCode": PLAZA_ID, "orgTypeCode": "10003"}
                    activity_rows = icsp_client.fetch_data_hybrid(
                        f"{icsp_client.base}/yinli-xapi-b/activity/v1/activity/applyList", "POST", self.s,
                        self.e, "摩客云活动报名记录", activity_pl
                    )
                    if activity_rows:
                        icsp_client.enrich_activity_data(activity_rows)
                    icsp_data["activity"] = activity_rows
                else:
                    self.signals.log.emit("WARN", "ICSP登录失败")

            # ====== 数据清洗与汇总 ======
            self.signals.log.emit("INFO", "开始数据清洗与汇总...")
            log_fn = lambda lvl, msg: self.signals.log.emit(lvl, msg)

            # 聚合各数据源
            pay_daily = aggregate_pay_daily(icsp_data.get("pay", []))
            refund_daily = aggregate_refund_daily(icsp_data.get("refund", []))
            mall_daily_detail, mall_product_summary = aggregate_mall_orders(icsp_data.get("mall", []))
            mall_ref_daily_detail, mall_ref_product_summary = aggregate_mall_refunds(icsp_data.get("mall_refund", []))
            rent_daily = aggregate_rent_daily(icsp_data.get("rent", []))
            activity_daily = aggregate_activity_daily(icsp_data.get("activity", []))
            parking_daily = aggregate_parking_daily(parking_rows)

            # CSV 解析
            script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            fund_csv_path = os.path.join(script_dir, "微信支付资金账单.csv")
            trade_csv_path = os.path.join(script_dir, "微信支付交易订单-服务商.csv")

            fund_rows_raw, fund_daily, refund_fee_daily = parse_wechat_fund_csv(fund_csv_path, log_fn)
            if not fund_daily:
                fund_daily = {}
            if not refund_fee_daily:
                refund_fee_daily = {}
            wechat_trade_daily = parse_wechat_trade_csv(trade_csv_path, log_fn)

            # 日期范围
            dates = _date_range(self.s, self.e)

            # ====== 生成 Excel ======
            fname = f"印享星对账表_{self.s}_{self.e}.xlsx"
            self.signals.log.emit("INFO", f"正在生成Excel: {fname}...")

            wb = Workbook()
            wb.remove(wb.active)

            # --- 辅助：从 fund_daily 提取特定业务类型的日金额 ---
            def fund_lookup(date, biz_keyword, income_keyword=None):
                if date not in fund_daily:
                    return 0.0
                total = 0.0
                for (biz_type, income_type), info in fund_daily[date].items():
                    if biz_keyword in biz_type:
                        if income_keyword is None or income_keyword in income_type:
                            total += info['amount']
                return total

            # --- 汇总商城退款按日 ---
            mall_ref_by_date = defaultdict(float)
            for (d, spu), amt in mall_ref_daily_detail.items():
                mall_ref_by_date[d] += amt

            # --- 汇总商城订单按日 ---
            mall_by_date = defaultdict(float)
            for (d, spu), info in mall_daily_detail.items():
                mall_by_date[d] += info['amount']

            # --- Sheet 12: 订单总表 ---
            order_summary_rows = []
            for d in dates:
                mall_amt = round(mall_by_date.get(d, 0), 2)
                rent_amt = round(rent_daily.get(d, 0), 2)
                act_amt = round(activity_daily.get(d, 0), 2)
                subtotal = round(mall_amt + rent_amt + act_amt, 2)
                ref_amt = round(mall_ref_by_date.get(d, 0), 2)
                total = round(subtotal - ref_amt, 2)
                yxx_total = round(pay_daily.get(d, 0) - refund_daily.get(d, 0), 2)
                diff = round(total - yxx_total, 2)
                if mall_amt or rent_amt or act_amt or ref_amt or yxx_total:
                    order_summary_rows.append([d, mall_amt, rent_amt, act_amt, subtotal, ref_amt, total, diff])
            self.save_clean_sheet(wb, "订单总表",
                                 ["日期", "商城订单", "押金订单", "活动订单", "合计", "商城押金退款", "总计", "差异"],
                                 order_summary_rows)

            # --- Sheet 13: 对账总表 ---
            recon_rows = []
            for d in dates:
                yxx_pay = round(pay_daily.get(d, 0), 2)
                yxx_ref = round(refund_daily.get(d, 0), 2)
                yxx_total = round(yxx_pay - yxx_ref, 2)
                scan_fine = None  # 扫码罚款留空
                scan_fee = round(parking_daily.get(d, 0) + wechat_trade_daily.get(d, 0), 2)
                wx_income = round(fund_lookup(d, '交易', '收入'), 2)
                wx_fee = round(fund_lookup(d, '手续费'), 2)
                wx_refund = round(fund_lookup(d, '退款'), 2)
                net_income = round(wx_income - wx_fee - wx_refund, 2)
                withdraw = round(fund_lookup(d, '提现'), 2)
                ref_fee = round(refund_fee_daily.get(d, 0), 2)
                scan_fine_val = 0 if scan_fine is None else scan_fine
                diff = round(yxx_total + scan_fine_val + scan_fee - wx_fee - (net_income - ref_fee), 2)
                remark = None

                has_data = any([yxx_pay, yxx_ref, scan_fee, wx_income, wx_fee, wx_refund, withdraw, ref_fee])
                if has_data:
                    recon_rows.append([
                        d, yxx_pay, yxx_ref, yxx_total, scan_fine, scan_fee,
                        wx_income, wx_fee, wx_refund, net_income, withdraw, ref_fee, diff, remark
                    ])
            self.save_clean_sheet(wb, "对账总表",
                                 ["日期", "印享星支付", "印享星退款", "印享星合计", "扫码罚款", "扫码收费",
                                  "微信支付到账", "手续费", "退款", "到账净额", "提现", "退款手续费", "差异", "备注"],
                                 recon_rows)

            # 调整 sheet 顺序：对账总表放最前面
            desired_order = ["对账总表", "订单总表"]
            sheet_names = wb.sheetnames
            for idx, name in enumerate(desired_order):
                if name in sheet_names:
                    wb.move_sheet(name, offset=idx - sheet_names.index(name))
                    sheet_names = wb.sheetnames  # refresh after move

            wb.save(fname)

            summary = (f"任务完成！文件: {fname}\n\n"
                       f"停车订单: {len(parking_rows)} 条\n"
                       f"支付订单: {len(icsp_data.get('pay', []))} 条\n"
                       f"支付退款: {len(icsp_data.get('refund', []))} 条\n"
                       f"商城订单: {len(icsp_data.get('mall', []))} 条\n"
                       f"商城退款: {len(icsp_data.get('mall_refund', []))} 条\n"
                       f"租借订单: {len(icsp_data.get('rent', []))} 条\n"
                       f"活动报名: {len(icsp_data.get('activity', []))} 条")

            self.signals.success.emit(summary)
            self.signals.log.emit("SUCCESS", "所有任务已完成")

        except InterruptedError:
            self.signals.log.emit("WARN", "任务停止")
            self.signals.stopped.emit()
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.signals.error.emit(str(e))
        finally:
            self.signals.finished.emit()


class CaptchaThread(QThread):
    def __init__(self, client, signals):
        super().__init__()
        self.client = client
        self.signals = signals

    def run(self):
        try:
            uuid, img_bytes = self.client.fetch_captcha()
            if uuid and img_bytes:
                self.signals.captcha_received.emit(uuid, img_bytes)
            else:
                self.signals.log.emit("WARN", "验证码加载失败")
        except Exception:
            self.signals.log.emit("ERROR", "验证码网络异常")


class ClickableLabel(QLabel):
    clicked = Signal()

    def mousePressEvent(self, e): self.clicked.emit(); super().mousePressEvent(e)


# ====== UI 组件 ======

class ModernInput(QWidget):
    def __init__(self, label, default="", is_password=False):
        super().__init__()
        l = QVBoxLayout(self);
        l.setContentsMargins(0, 0, 0, 0);
        l.setSpacing(5)
        self.lbl = QLabel(label);
        self.lbl.setStyleSheet(f"color:{COLOR_TEXT_H2};font-weight:bold;")
        self.inp = QLineEdit(default);
        self.inp.setFixedHeight(35)
        if is_password: self.inp.setEchoMode(QLineEdit.Password)
        self.inp.setStyleSheet(f"""
            QLineEdit {{
                background-color: {COLOR_CARD};
                color: {COLOR_TEXT_H1};
                border: 1px solid {COLOR_BORDER};
                border-radius: {RAD_INPUT}px;
                padding: 0 10px;
            }}
        """)
        l.addWidget(self.lbl);
        l.addWidget(self.inp)

    def text(self): return self.inp.text().strip()


class ModernDateInput(QWidget):
    def __init__(self, label):
        super().__init__()
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(5)

        self.lbl = QLabel(label);
        self.lbl.setStyleSheet(f"color:{COLOR_TEXT_H2};font-weight:bold;")
        self.layout.addWidget(self.lbl)

        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("yyyy-MM-dd")
        self.date_edit.setFixedHeight(35)

        self.date_edit.setStyleSheet(f"""
            QDateEdit {{
                background-color: {COLOR_CARD};
                color: {COLOR_TEXT_H1};
                border: 1px solid {COLOR_BORDER};
                border-radius: {RAD_INPUT}px;
                padding-left: 10px;
            }}
            QDateEdit::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 30px;
                border-left: 1px solid {COLOR_BORDER};
                border-top-right-radius: {RAD_INPUT}px;
                border-bottom-right-radius: {RAD_INPUT}px;
                background-color: {COLOR_ACCENT_BG}; 
            }}
            QDateEdit::down-arrow {{
                width: 10px;
                height: 10px;
                border: none;
                background: none;
                image: none;
                border-left: 2px solid {COLOR_PRIMARY};
                border-bottom: 2px solid {COLOR_PRIMARY};
                transform: rotate(-45deg);
                margin-top: -3px;
            }}
        """)

        self.layout.addWidget(self.date_edit)

    def setDate(self, qdate): self.date_edit.setDate(qdate)

    def date(self): return self.date_edit.date()


class LogPanel(QTextEdit):
    def __init__(self):
        super().__init__();
        self.setReadOnly(True)
        self.setStyleSheet(
            f"background:{COLOR_LOG_BG};color:{COLOR_LOG_TEXT};border-radius:10px;padding:5px;border:none;")

    def append_log(self, level, msg):
        c = "#98C379" if level == "SUCCESS" else "#E5C07B" if level == "WARN" else "#E06C75" if level == "ERROR" else "#ABB2BF"
        self.append(f'<span style="color:{c}">[{level}]</span> {msg}')


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(GLOBAL_STYLES)
        self.setWindowTitle("印享星对账助手 (20260225V8)")
        self.setFixedSize(DIM_W, DIM_H)
        self.worker = None
        self.parking_uuid = ""
        self.parking_client = ParkingClient()

        cw = QWidget();
        self.setCentralWidget(cw);
        lay = QVBoxLayout(cw)
        lay.setContentsMargins(20, 30, 20, 20)

        h = QHBoxLayout()
        ico = QLabel("🗓️");
        ico.setStyleSheet(f"background:{COLOR_ACCENT_BG};font-size:24px;border-radius:8px;padding:5px;")
        tit = QLabel("印享星数据提取");
        tit.setStyleSheet(f"font-size:18px;font-weight:bold;color:{COLOR_TEXT_H1};")
        h.addWidget(ico);
        h.addWidget(tit);
        h.addStretch();
        lay.addLayout(h)

        card = QFrame();
        card.setStyleSheet(f"QFrame{{background:{COLOR_CARD};border-radius:{RAD_CARD}px;}}")
        shadow = QGraphicsDropShadowEffect();
        shadow.setBlurRadius(20);
        shadow.setColor(QColor(0, 0, 0, 20));
        shadow.setOffset(QPointF(0, 4))
        card.setGraphicsEffect(shadow)
        cl = QVBoxLayout(card);
        cl.setContentsMargins(20, 20, 20, 20);
        cl.setSpacing(12)

        # 1. 账号密码行
        row1 = QHBoxLayout()
        self.u = ModernInput("ICSP账号")
        self.p = ModernInput("ICSP密码", is_password=True)
        row1.addWidget(self.u);
        row1.addWidget(self.p)
        cl.addLayout(row1)

        # 2. 日期选择行
        row2 = QHBoxLayout()
        self.s = ModernDateInput("开始日期")
        self.e = ModernDateInput("结束日期")
        now = QDate.currentDate()
        self.s.setDate(QDate(now.year(), now.month(), 1))
        self.e.setDate(now.addDays(-1))
        row2.addWidget(self.s);
        row2.addWidget(self.e)
        cl.addLayout(row2)

        # 3. 验证码和按钮行
        lbl_cap = QLabel("停车验证码 / 操作");
        lbl_cap.setStyleSheet(f"color:{COLOR_TEXT_H2};font-weight:bold;font-size:11px;")
        cl.addWidget(lbl_cap)

        cap_layout = QHBoxLayout()
        cap_layout.setSpacing(10)

        self.img_label = ClickableLabel()
        self.img_label.setFixedSize(120, 40)
        self.img_label.setStyleSheet("background:#EEE; border-radius:5px; border:1px solid #CCC;")
        self.img_label.setScaledContents(True)
        self.img_label.setCursor(QCursor(Qt.PointingHandCursor))
        self.img_label.clicked.connect(self.refresh_captcha)

        self.inp_cap = QLineEdit();
        self.inp_cap.setPlaceholderText("验证码")
        self.inp_cap.setFixedWidth(120)
        self.inp_cap.setFixedHeight(40);
        self.inp_cap.setStyleSheet(f"""
            background-color: {COLOR_CARD};
            color: {COLOR_TEXT_H1};
            border:1px solid {COLOR_BORDER};
            border-radius:{RAD_INPUT}px;
            padding:0 10px;
        """)

        self.btn = QPushButton("🚀 开始抓取")
        self.btn.setFixedHeight(40);
        self.btn.setCursor(Qt.PointingHandCursor)
        self.btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn.setStyleSheet(
            f"background:{COLOR_PRIMARY};color:white;border-radius:{RAD_INPUT}px;font-weight:bold;font-size:15px;border:none;")
        self.btn.clicked.connect(self.run);

        cap_layout.addWidget(self.img_label);
        cap_layout.addWidget(self.inp_cap);
        cap_layout.addWidget(self.btn)
        cl.addLayout(cap_layout)

        lay.addWidget(card)
        self.log = LogPanel();
        lay.addWidget(self.log)

        self.sig = WorkerSignals()
        self.sig.log.connect(self.log.append_log)
        self.sig.finished.connect(self.on_finished)
        self.sig.success.connect(self.on_success)
        self.sig.error.connect(lambda m: self.log.append_log("ERROR", f"任务异常: {m}"))
        self.sig.captcha_received.connect(self.update_captcha)
        self.refresh_captcha()

    def refresh_captcha(self):
        self.log.append_log("INFO", "刷新验证码...")
        self.img_label.setText("...")
        self.captcha_thread = CaptchaThread(self.parking_client, self.sig)
        self.captcha_thread.start()

    def update_captcha(self, uuid, img_bytes):
        self.parking_uuid = uuid;
        pix = QPixmap();
        pix.loadFromData(img_bytes);
        self.img_label.setPixmap(pix)

    def run(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.btn.setText("正在停止...")
            self.btn.setEnabled(False)
            return
        u, p = self.u.text(), self.p.text()
        s_str = self.s.date().toString("yyyy-MM-dd")
        e_str = self.e.date().toString("yyyy-MM-dd")
        p_code = self.inp_cap.text().strip()

        if not u and not p and not p_code:
            QMessageBox.warning(self, "提示", "请至少填写 ICSP 账号密码或停车验证码")
            return

        self.log.clear();
        self.btn.setText("停止运行");
        self.btn.setStyleSheet(f"background:{COLOR_STOP};color:white;border-radius:{RAD_INPUT}px;border:none;")
        self.worker = TaskThread(u, p, s_str, e_str, p_code, self.parking_uuid, self.parking_client, self.sig)
        self.worker.start()

    def on_finished(self):
        self.btn.setText("🚀 开始抓取")
        self.btn.setEnabled(True)
        self.btn.setStyleSheet(f"background:{COLOR_PRIMARY};color:white;border-radius:{RAD_INPUT}px;border:none;")
        self.refresh_captcha()

    def on_success(self, summary):
        QMessageBox.information(self, "任务完成", summary)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec())