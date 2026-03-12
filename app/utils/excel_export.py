from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_FIELD_FILES = ("point_flow_fields.json", "point_flow_fields.txt")
PROJECT_ROOT = Path(__file__).resolve().parents[2]


def _read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return path.read_text(encoding=encoding)
        except Exception:
            continue
    return ""


def load_fields_from_sample(base_dir: str | Path | None = None) -> list[str]:
    root = Path(base_dir) if base_dir is not None else PROJECT_ROOT
    candidates = [root / name for name in DEFAULT_FIELD_FILES]

    try:
        for path in sorted(root.iterdir()):
            if path.is_file() and path.suffix.lower() in {".json", ".txt"}:
                stem = path.stem.lower()
                if "flow" in stem or "field" in stem:
                    candidates.append(path)
    except Exception:
        pass

    for sample_path in candidates:
        if not sample_path.is_file():
            continue

        try:
            raw = _read_text_with_fallback(sample_path).strip()
            if not raw:
                continue

            sample = json.loads(raw.rstrip(","))
            if isinstance(sample, dict):
                return list(sample.keys())
            if isinstance(sample, list) and all(isinstance(item, str) for item in sample):
                return sample
        except json.JSONDecodeError:
            lines = [line.strip().strip(",") for line in raw.splitlines() if line.strip()]
            if lines:
                return lines
        except Exception:
            continue

    return []


def format_cell_value(key: str, value):
    if value is None:
        return ""
    if key == "consumeAmount":
        try:
            return round(float(value) / 100, 2)
        except Exception:
            return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if "time" in key.lower() and isinstance(value, (int, float)) and value > 10**12:
        try:
            return datetime.fromtimestamp(value / 1000).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)
    return value


def export_to_excel(
    rows: list[dict],
    fields: Sequence[str],
    start_date: str,
    end_date: str,
    output_dir: str | Path,
    file_tag: str | None = None,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "points_flow"

    field_list = list(fields)
    if not field_list:
        keys: set[str] = set()
        for row in rows:
            keys.update(row.keys())
        field_list = sorted(keys)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    worksheet.append(field_list)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    for row in rows:
        worksheet.append([format_cell_value(key, row.get(key)) for key in field_list])

    for row_cells in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(field_list),
    ):
        for cell in row_cells:
            cell.border = thin
            cell.alignment = center

    worksheet.freeze_panes = "A2"
    for col_idx, header in enumerate(field_list, start=1):
        column = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_num in range(2, worksheet.max_row + 1):
            value = worksheet[f"{column}{row_num}"].value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        worksheet.column_dimensions[column].width = min(max_len * 1.2 + 4, 60)

    export_root = Path(output_dir)
    export_root.mkdir(parents=True, exist_ok=True)

    suffix = f"_{file_tag}" if file_tag else ""
    file_name = f"points_flow_{start_date}_{end_date}{suffix}.xlsx"
    output_path = export_root / file_name
    workbook.save(output_path)
    return output_path
