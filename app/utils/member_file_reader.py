from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


MOBILE_HEADER_CANDIDATES = ("手机号", "手机号码", "mobile", "mobile_no", "mobileno", "phone")


@dataclass(slots=True)
class MobileFileReadResult:
    file_path: str
    total_rows: int
    valid_rows: int
    unique_mobiles: list[str]
    invalid_rows: list[dict[str, str]]


def normalize_mobile(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 11:
        return digits[-11:]
    return digits


def is_valid_mobile(value: str) -> bool:
    return bool(re.fullmatch(r"1\d{10}", value))


def read_mobile_list_from_file(file_path: str | Path, mobile_column: str | None = None) -> MobileFileReadResult:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_mobile_list_from_csv(path, mobile_column=mobile_column)
    return _read_mobile_list_from_excel(path, mobile_column=mobile_column)


def _read_mobile_list_from_excel(path: Path, mobile_column: str | None = None) -> MobileFileReadResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    workbook.close()
    return _build_result(path, rows, mobile_column=mobile_column)


def _read_mobile_list_from_csv(path: Path, mobile_column: str | None = None) -> MobileFileReadResult:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            with open(path, "r", encoding=encoding, newline="") as fp:
                rows = list(csv.reader(fp))
            return _build_result(path, rows, mobile_column=mobile_column)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Cannot decode CSV file: {path}")


def _build_result(path: Path, rows: list[list[object]], mobile_column: str | None = None) -> MobileFileReadResult:
    filtered_rows = [list(row) for row in rows if any(cell is not None and str(cell).strip() != "" for cell in row)]
    if not filtered_rows:
        raise ValueError(f"No usable rows found in {path}")

    header_row = [str(cell).strip() if cell is not None else "" for cell in filtered_rows[0]]
    mobile_index = _resolve_mobile_column(header_row, filtered_rows, mobile_column=mobile_column)
    has_header = _looks_like_header(header_row, mobile_index, mobile_column=mobile_column)
    data_rows = filtered_rows[1:] if has_header else filtered_rows

    unique_mobiles: list[str] = []
    seen_mobiles: set[str] = set()
    invalid_rows: list[dict[str, str]] = []
    valid_rows = 0

    for row_number, row in enumerate(data_rows, start=2 if has_header else 1):
        raw_value = row[mobile_index] if mobile_index < len(row) else ""
        mobile = normalize_mobile(raw_value)
        if is_valid_mobile(mobile):
            valid_rows += 1
            if mobile not in seen_mobiles:
                seen_mobiles.add(mobile)
                unique_mobiles.append(mobile)
            continue
        invalid_rows.append({"row_no": str(row_number), "raw_value": "" if raw_value is None else str(raw_value)})

    return MobileFileReadResult(
        file_path=str(path),
        total_rows=len(data_rows),
        valid_rows=valid_rows,
        unique_mobiles=unique_mobiles,
        invalid_rows=invalid_rows,
    )


def _resolve_mobile_column(
    header_row: list[str],
    rows: list[list[object]],
    *,
    mobile_column: str | None,
) -> int:
    if mobile_column:
        normalized_target = mobile_column.strip().lower()
        for index, header in enumerate(header_row):
            if header.strip().lower() == normalized_target:
                return index
        raise ValueError(f"Mobile column {mobile_column!r} was not found.")

    for candidate in MOBILE_HEADER_CANDIDATES:
        for index, header in enumerate(header_row):
            if candidate.lower() == header.strip().lower():
                return index

    max_cols = max(len(row) for row in rows)
    best_index = 0
    best_score = -1
    sample_rows = rows[: min(len(rows), 200)]
    for index in range(max_cols):
        score = 0
        for row in sample_rows:
            raw_value = row[index] if index < len(row) else ""
            if is_valid_mobile(normalize_mobile(raw_value)):
                score += 1
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _looks_like_header(header_row: list[str], mobile_index: int, *, mobile_column: str | None) -> bool:
    if mobile_column:
        return True
    if mobile_index < len(header_row):
        header_text = header_row[mobile_index].strip().lower()
        if any(candidate.lower() == header_text for candidate in MOBILE_HEADER_CANDIDATES):
            return True
    return False
