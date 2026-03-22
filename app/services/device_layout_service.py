from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.device_layout import DeviceLayoutPoint


ROOT_DIR = Path(__file__).resolve().parents[2]
MAP_DIR = ROOT_DIR / "map"
PATROL_TEMPLATE_PATH = MAP_DIR / "巡更点位模板.xlsx"
RATIO_STEP = Decimal("0.000001")


@dataclass(frozen=True, slots=True)
class DevicePointTypeMeta:
    key: str
    label: str


@dataclass(frozen=True, slots=True)
class FloorMeta:
    code: str
    label: str
    image_name: str
    image_width: int
    image_height: int


POINT_TYPES = [
    DevicePointTypeMeta(key="guide", label="导视点位"),
    DevicePointTypeMeta(key="passenger", label="客流点位"),
    DevicePointTypeMeta(key="wifi", label="WiFi点位"),
    DevicePointTypeMeta(key="patrol", label="巡更点位"),
]
POINT_TYPE_META_BY_KEY = {item.key: item for item in POINT_TYPES}
POINT_TYPE_KEY_BY_LABEL = {
    "导视点位": "guide",
    "客流点位": "passenger",
    "WiFi点位": "wifi",
    "wifi点位": "wifi",
    "巡更点位": "patrol",
}

FLOORS = [
    FloorMeta(code="L5", label="L5", image_name="L5.png", image_width=1024, image_height=1536),
    FloorMeta(code="L4", label="L4", image_name="L4.png", image_width=1024, image_height=1536),
    FloorMeta(code="L3", label="L3", image_name="L3.png", image_width=1024, image_height=1536),
    FloorMeta(code="L2", label="L2", image_name="L2.png", image_width=1024, image_height=1536),
    FloorMeta(code="L1", label="L1", image_name="L1.png", image_width=1024, image_height=1536),
    FloorMeta(code="B1", label="B1", image_name="B1.png", image_width=1536, image_height=1024),
    FloorMeta(code="B2", label="B2", image_name="B2.png", image_width=1536, image_height=1024),
]
FLOOR_META_BY_CODE = {item.code: item for item in FLOORS}
FLOOR_CODE_ALIASES = {
    "5F": "L5",
    "4F": "L4",
    "3F": "L3",
    "2F": "L2",
    "1F": "L1",
}


@dataclass(slots=True)
class DeviceLayoutImportSummary:
    point_type: str
    total_rows: int = 0
    created_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
def normalize_point_type(value: str | None) -> str:
    text = str(value or "").strip()
    if not text:
        raise ValueError("点位类型不能为空")
    lowered = text.lower()
    if lowered in POINT_TYPE_META_BY_KEY:
        return lowered
    normalized = POINT_TYPE_KEY_BY_LABEL.get(text.replace(" ", "")) or POINT_TYPE_KEY_BY_LABEL.get(lowered.replace(" ", ""))
    if normalized:
        return normalized
    raise ValueError(f"不支持的点位类型：{text}")


def normalize_floor_code(value: str | None) -> str:
    text = str(value or "").strip().upper()
    if not text:
        raise ValueError("楼层不能为空")
    text = FLOOR_CODE_ALIASES.get(text, text)
    if text in FLOOR_META_BY_CODE:
        return text
    raise ValueError(f"不支持的楼层：{value}")


def ratio_to_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def quantize_ratio(value: float | Decimal | None) -> Decimal | None:
    if value is None:
        return None
    decimal_value = Decimal(str(value))
    if decimal_value < 0 or decimal_value > 1:
        raise ValueError("坐标值必须在 0 到 1 之间")
    return decimal_value.quantize(RATIO_STEP, rounding=ROUND_HALF_UP)


def parse_coordinate(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    is_percent = text.endswith("%")
    if is_percent:
        text = text[:-1].strip()
    try:
        decimal_value = Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"无法识别坐标值：{value}") from exc
    if is_percent or decimal_value > 1:
        decimal_value = decimal_value / Decimal("100")
    if decimal_value < 0 or decimal_value > 1:
        raise ValueError(f"坐标值超出范围：{value}")
    return decimal_value.quantize(RATIO_STEP, rounding=ROUND_HALF_UP)


def serialize_point(point: DeviceLayoutPoint) -> dict[str, Any]:
    return {
        "point_type": point.point_type,
        "point_code": point.point_code,
        "point_name": point.point_name,
        "floor_code": point.floor_code,
        "x_ratio": ratio_to_float(point.x_ratio),
        "y_ratio": ratio_to_float(point.y_ratio),
    }


def resolve_map_path(floor_code: str) -> Path:
    normalized_floor = normalize_floor_code(floor_code)
    floor = FLOOR_META_BY_CODE[normalized_floor]
    path = MAP_DIR / floor.image_name
    if not path.is_file():
        raise FileNotFoundError(path)
    return path


def _header_style(ws) -> None:
    border = Border(
        left=Side(style="thin", color="D9E2F1"),
        right=Side(style="thin", color="D9E2F1"),
        top=Side(style="thin", color="D9E2F1"),
        bottom=Side(style="thin", color="D9E2F1"),
    )
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = alignment

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

    for index, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(index)].width = min(max(max_length * 1.25 + 4, 12), 28)


def _extract_rows_from_template(template_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(template_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(cell or "").strip() for cell in next(rows_iter, [])]
    header_index = {header: index for index, header in enumerate(headers)}
    required = ["编码", "点位名称", "楼层"]
    if any(item not in header_index for item in required):
        raise RuntimeError(f"点位模板缺少必要列：{', '.join(required)}")

    rows: list[dict[str, str]] = []
    for values in rows_iter:
        code = str(values[header_index["编码"]] or "").strip()
        name = str(values[header_index["点位名称"]] or "").strip()
        floor = str(values[header_index["楼层"]] or "").strip()
        if not code or not name or not floor:
            continue
        rows.append({"point_code": code, "point_name": name, "floor_code": normalize_floor_code(floor)})
    return rows


class DeviceLayoutService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def build_config(self) -> dict[str, Any]:
        return {
            "default_point_type": "patrol",
            "default_floor_code": FLOORS[0].code,
            "point_types": [{"key": item.key, "label": item.label} for item in POINT_TYPES],
            "floors": [
                {
                    "code": item.code,
                    "label": item.label,
                    "image_width": item.image_width,
                    "image_height": item.image_height,
                }
                for item in FLOORS
            ],
        }

    def ensure_seeded(self, point_type: str) -> None:
        normalized_point_type = normalize_point_type(point_type)
        if normalized_point_type != "patrol" or not PATROL_TEMPLATE_PATH.is_file():
            return

        existing_count = int(
            self.db.scalar(
                select(func.count()).select_from(DeviceLayoutPoint).where(DeviceLayoutPoint.point_type == normalized_point_type)
            )
            or 0
        )
        if existing_count > 0:
            return

        rows = _extract_rows_from_template(PATROL_TEMPLATE_PATH)
        for row in rows:
            self.db.add(
                DeviceLayoutPoint(
                    point_type=normalized_point_type,
                    point_code=row["point_code"],
                    point_name=row["point_name"],
                    floor_code=row["floor_code"],
                    source_file=str(PATROL_TEMPLATE_PATH.resolve()),
                )
            )
        self.db.commit()

    def list_points(self, point_type: str) -> list[dict[str, Any]]:
        normalized_point_type = normalize_point_type(point_type)
        self.ensure_seeded(normalized_point_type)
        points = self.db.scalars(
            select(DeviceLayoutPoint)
            .where(DeviceLayoutPoint.point_type == normalized_point_type)
            .order_by(DeviceLayoutPoint.floor_code, DeviceLayoutPoint.point_name, DeviceLayoutPoint.point_code)
        ).all()
        return [serialize_point(item) for item in points]

    def save_points(self, point_type: str, items: list[dict[str, Any]]) -> int:
        normalized_point_type = normalize_point_type(point_type)
        self.ensure_seeded(normalized_point_type)

        point_codes = sorted({str(item.get("point_code") or "").strip() for item in items if item.get("point_code")})
        existing = {
            item.point_code: item
            for item in self.db.scalars(
                select(DeviceLayoutPoint).where(
                    DeviceLayoutPoint.point_type == normalized_point_type,
                    DeviceLayoutPoint.point_code.in_(point_codes),
                )
            ).all()
        } if point_codes else {}

        saved_count = 0
        for item in items:
            point_code = str(item.get("point_code") or "").strip()
            if not point_code:
                continue
            point_name = str(item.get("point_name") or point_code).strip() or point_code
            floor_code = normalize_floor_code(item.get("floor_code"))
            x_ratio = quantize_ratio(item.get("x_ratio"))
            y_ratio = quantize_ratio(item.get("y_ratio"))

            point = existing.get(point_code)
            if point is None:
                point = DeviceLayoutPoint(
                    point_type=normalized_point_type,
                    point_code=point_code,
                    point_name=point_name,
                    floor_code=floor_code,
                )
                self.db.add(point)
                existing[point_code] = point
            else:
                point.point_name = point_name
                point.floor_code = floor_code

            point.x_ratio = x_ratio
            point.y_ratio = y_ratio
            saved_count += 1

        self.db.commit()
        return saved_count

    def import_points(
        self,
        *,
        file_bytes: bytes,
        file_name: str,
        selected_point_type: str | None,
    ) -> DeviceLayoutImportSummary:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [str(cell or "").strip() for cell in next(rows_iter, [])]
        header_index = {header: index for index, header in enumerate(headers)}

        required_headers = ["编码", "点位名称", "楼层"]
        if any(item not in header_index for item in required_headers):
            raise RuntimeError(f"导入模板缺少必要列：{', '.join(required_headers)}")

        normalized_selected_type = normalize_point_type(selected_point_type or "patrol")
        self.ensure_seeded(normalized_selected_type)
        summary = DeviceLayoutImportSummary(point_type=normalized_selected_type)

        point_type_column = next((item for item in ("点位类型", "point_type", "类型") if item in header_index), None)
        x_column = next((item for item in ("坐标X", "x_ratio", "X", "坐标x") if item in header_index), None)
        y_column = next((item for item in ("坐标Y", "y_ratio", "Y", "坐标y") if item in header_index), None)

        cache: dict[tuple[str, str], DeviceLayoutPoint] = {}

        def get_existing(normalized_type: str, point_code: str) -> DeviceLayoutPoint | None:
            key = (normalized_type, point_code)
            if key in cache:
                return cache[key]
            point = self.db.scalar(
                select(DeviceLayoutPoint).where(
                    DeviceLayoutPoint.point_type == normalized_type,
                    DeviceLayoutPoint.point_code == point_code,
                )
            )
            if point is not None:
                cache[key] = point
            return point

        for values in rows_iter:
            if values is None or not any(value not in (None, "") for value in values):
                continue
            summary.total_rows += 1
            code = str(values[header_index["编码"]] or "").strip()
            name = str(values[header_index["点位名称"]] or "").strip()
            floor = str(values[header_index["楼层"]] or "").strip()
            if not code or not name or not floor:
                summary.skipped_count += 1
                continue

            row_type = normalized_selected_type
            if point_type_column is not None and values[header_index[point_type_column]] not in (None, ""):
                row_type = normalize_point_type(str(values[header_index[point_type_column]]))
            self.ensure_seeded(row_type)

            floor_code = normalize_floor_code(floor)
            point = get_existing(row_type, code)
            x_ratio = parse_coordinate(values[header_index[x_column]]) if x_column is not None else None
            y_ratio = parse_coordinate(values[header_index[y_column]]) if y_column is not None else None

            if point is None:
                point = DeviceLayoutPoint(
                    point_type=row_type,
                    point_code=code,
                    point_name=name,
                    floor_code=floor_code,
                    x_ratio=x_ratio,
                    y_ratio=y_ratio,
                    source_file=file_name,
                )
                self.db.add(point)
                cache[(row_type, code)] = point
                summary.created_count += 1
            else:
                point.point_name = name
                point.floor_code = floor_code
                if x_column is not None:
                    point.x_ratio = x_ratio
                if y_column is not None:
                    point.y_ratio = y_ratio
                point.source_file = file_name
                summary.updated_count += 1

        self.db.commit()
        return summary

    def build_export_workbook(self, point_type: str) -> Workbook:
        normalized_point_type = normalize_point_type(point_type)
        self.ensure_seeded(normalized_point_type)
        points = self.db.scalars(
            select(DeviceLayoutPoint)
            .where(DeviceLayoutPoint.point_type == normalized_point_type)
            .order_by(DeviceLayoutPoint.floor_code, DeviceLayoutPoint.point_name, DeviceLayoutPoint.point_code)
        ).all()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "点位坐标"
        sheet.append(["编码", "点位名称", "楼层", "点位类型", "坐标X", "坐标Y"])
        for point in points:
            sheet.append(
                [
                    point.point_code,
                    point.point_name,
                    point.floor_code,
                    POINT_TYPE_META_BY_KEY[point.point_type].label,
                    ratio_to_float(point.x_ratio),
                    ratio_to_float(point.y_ratio),
                ]
            )
        _header_style(sheet)
        return workbook


def build_workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
