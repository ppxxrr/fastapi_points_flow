from __future__ import annotations

import base64
import csv
import hashlib
import math
import re
import urllib.parse
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.db.session import SessionLocal
from app.services.bi_analytics_service import BiAnalyticsService
from app.services.icsp_client import ICSP_BASE, ICSPClient, ORG_TYPE_CODE, PLAZA_CODE, TENANT_ID
from app.services.parking_captcha_store import (
    export_captcha_session_cookies,
    parking_captcha_challenge_store,
)


PAGE_SIZE = 100
GRANT_PLAZA_BU_ID = "293"
FIXED_RECONCILIATION_USERNAME = "h-pengxr01"
COUPON_GROUP_TENANT_ID = 468
COUPON_GROUP_APP_BU_ID = 2070
COUPON_GROUP_OPEN_STATUS = "3"

PARKING_LOGIN_URL = "https://szwryportal.aibee.cn/union/portal/loginV2"
PARKING_CAPTCHA_URL = "https://szwryportal.aibee.cn/union/portal/getKaptcha"
PARKING_DATA_URL = "https://szwryportal.aibee.cn/parkingbi/api/backend/reservation/search_reservation"
PARKING_USER = "railinadmin"
PARKING_PWD_RAW = "pemdot-9sudxi-kAzcyw"
UPLOAD_FUND_FILENAME = "wechat_fund.csv"
UPLOAD_TRADE_FILENAME = "wechat_trade.csv"

PARKING_HEADER_MAP = {
    "id": "ID",
    "createdAt": "创建时间",
    "site_id": "场站ID",
    "floor": "楼层",
    "reservation_number": "预约单号",
    "reservation_people": "预约人",
    "phone": "手机号",
    "member_type": "会员类型",
    "car_plate": "车牌号",
    "parking_space": "车位号",
    "order_time": "下单时间",
    "begin_time": "预约开始时间",
    "arrival_time": "入场时间",
    "current_state": "当前状态",
    "payment_amount": "支付金额",
}


@dataclass(slots=True)
class ReconciliationJobResult:
    output_file: Path
    result_count: int


@dataclass(slots=True)
class ParkingProjectConfig:
    key: str
    label: str
    plaza_id: str
    rent_org_name: str
    project_id: str
    referer: str
    enable_parking: bool


PARKING_PROJECTS: dict[str, ParkingProjectConfig] = {
    "railin_shenzhen_vip": ParkingProjectConfig(
        key="railin_shenzhen_vip",
        label="睿印",
        plaza_id="G002Z008C0030",
        rent_org_name="深圳湾睿印RAIL IN",
        project_id="railin_shenzhen_vip",
        referer="https://szwryportal.aibee.cn/micro/parking/dashboard/detail/BackgroundReservation",
        enable_parking=True,
    ),
    "railinli": ParkingProjectConfig(
        key="railinli",
        label="睿印里",
        plaza_id="G013Z001C0034",
        rent_org_name="睿印里RAIL INLI",
        project_id="railinli",
        referer="https://szwryportal.aibee.cn/micro/parking/dashboard/detail/BackgroundReservation",
        enable_parking=False,
    )
}


def list_parking_projects() -> list[ParkingProjectConfig]:
    return list(PARKING_PROJECTS.values())


def get_parking_project(project_key: str | None) -> ParkingProjectConfig:
    key = (project_key or "railin_shenzhen_vip").strip()
    project = PARKING_PROJECTS.get(key)
    if project is None:
        raise RuntimeError(f"Unsupported parking project: {key}")
    return project


def _thin_border() -> Border:
    side = Side(style="thin", color="D9E2F1")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_sheet(ws) -> None:
    border = _thin_border()
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    data_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = border

    for index, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[get_column_letter(index)].width = min(max(max_length * 1.35 + 4, 12), 42)


def _extract_rows_and_total(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], int]:
    if not isinstance(payload, dict):
        return [], 0

    rows: list[dict[str, Any]] = []
    total = 0
    keys_to_check = ("rows", "data", "list", "result", "records", "content", "items", "resultList")

    for key in keys_to_check:
        if isinstance(payload.get(key), list):
            rows = payload[key]
            break

    nested = payload.get("data")
    if not rows and isinstance(nested, dict):
        for key in keys_to_check:
            if isinstance(nested.get(key), list):
                rows = nested[key]
                break
        total = int(nested.get("total") or nested.get("totalCount") or nested.get("totalSize") or 0)

    if not total:
        total = int(payload.get("total") or payload.get("totalCount") or payload.get("totalSize") or 0)
    if not total and rows:
        total = len(rows)
    return rows, total


def _format_export_datetime(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value)
    return text.replace("T", " ")[:19] if "T" in text else text


def _assert_fixed_account(auth_state: dict[str, Any]) -> str:
    login_username = str(auth_state.get("login_username", "")).strip()
    if login_username.lower() != FIXED_RECONCILIATION_USERNAME.lower():
        raise RuntimeError(f"请使用固定账号 {FIXED_RECONCILIATION_USERNAME} 登录后再执行该功能。")
    return login_username


def _safe_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    text = str(value).strip().replace(",", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if text in ("", ".", "-", "-."):
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def _extract_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip().replace("/", "-")
    match = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    return match.group(1) if match else ""


def _detect_encoding(path: Path) -> str:
    for enc in ("utf-8-sig", "gbk", "gb18030"):
        try:
            path.read_text(encoding=enc)[:1024]
            return enc
        except Exception:
            continue
    return "utf-8-sig"


def resolve_uploaded_wechat_csvs(upload_root: str | Path, upload_session_id: str | None) -> tuple[Path, Path]:
    if not upload_session_id:
        raise RuntimeError("请先上传微信支付资金账单和交易订单 CSV。")
    root = Path(upload_root).resolve()
    session_dir = (root / upload_session_id).resolve()
    if root not in session_dir.parents or not session_dir.is_dir():
        raise RuntimeError("微信账单上传会话不存在或已失效。")
    fund_path = session_dir / UPLOAD_FUND_FILENAME
    trade_path = session_dir / UPLOAD_TRADE_FILENAME
    if not fund_path.is_file() or not trade_path.is_file():
        raise RuntimeError("未找到完整的微信账单 CSV 文件。")
    return fund_path, trade_path


def parse_wechat_fund_csv(path: Path, log_fn=None):
    if not path.is_file():
        raise RuntimeError("未找到微信支付资金账单 CSV 文件。")
    lines = path.read_text(encoding=_detect_encoding(path)).splitlines()
    header_idx = next((i for i, line in enumerate(lines) if line.strip().lstrip("`").strip().startswith("记账时间")), None)
    if header_idx is None:
        raise RuntimeError("微信支付资金账单 CSV 格式异常，未找到表头。")
    clean_lines = [line.strip() for line in lines[header_idx:] if line.strip() and not line.strip().startswith("#")]
    reader = csv.reader(clean_lines)
    headers = None
    rows = []
    for row_data in reader:
        cleaned = [c.strip().lstrip("`").strip() for c in row_data]
        if headers is None:
            headers = cleaned
            continue
        if len(cleaned) < len(headers):
            continue
        rows.append(dict(zip(headers, cleaned)))

    daily = defaultdict(lambda: defaultdict(lambda: {"amount": 0.0, "refund_fee": 0.0}))
    refund_fee_daily = defaultdict(float)
    for row in rows:
        biz_date = _extract_date(row.get("记账时间", ""))
        if not biz_date:
            continue
        biz_type = str(row.get("业务类型", "")).strip()
        income_type = str(row.get("收支类型", "")).strip()
        amount = _safe_float(row.get("收支金额(元)", "0"))
        daily[biz_date][(biz_type, income_type)]["amount"] += amount
        remark = str(row.get("备注", ""))
        fee_match = re.search(r"含手续费(\d+\.?\d*)元", remark)
        if fee_match:
            fee = float(fee_match.group(1))
            daily[biz_date][(biz_type, income_type)]["refund_fee"] += fee
            refund_fee_daily[biz_date] += fee

    if log_fn:
        log_fn("SUCCESS", f"[微信支付资金账单] 解析完成，共 {len(rows)} 条记录")
    return rows, dict(daily), dict(refund_fee_daily)


def parse_wechat_trade_csv(path: Path, log_fn=None):
    if not path.is_file():
        raise RuntimeError("未找到微信支付交易订单 CSV 文件。")
    lines = path.read_text(encoding=_detect_encoding(path)).splitlines()
    header_idx = next((i for i, line in enumerate(lines) if line.strip().lstrip("`").strip().startswith("交易时间")), None)
    if header_idx is None:
        raise RuntimeError("微信支付交易订单 CSV 格式异常，未找到表头。")
    clean_lines = [line.strip() for line in lines[header_idx:] if line.strip() and not line.strip().startswith("#")]
    reader = csv.reader(clean_lines)
    headers = None
    rows = []
    for row_data in reader:
        cleaned = [c.strip().lstrip("`").strip() for c in row_data]
        if headers is None:
            headers = cleaned
            continue
        if len(cleaned) < len(headers):
            continue
        rows.append(dict(zip(headers, cleaned)))

    filtered = [r for r in rows if r.get("交易时间", "").strip() and not r.get("交易时间", "").startswith("总")]
    daily = defaultdict(float)
    for row in filtered:
        biz_date = _extract_date(row.get("交易时间", ""))
        if not biz_date:
            continue
        daily[biz_date] += _safe_float(row.get("应结订单金额", "0"))
    if log_fn:
        log_fn("SUCCESS", f"[微信支付交易订单] 解析完成，共 {len(filtered)} 条记录")
    return dict(daily)


class CouponToolExporter:
    def __init__(self, auth_state: dict[str, Any], logger=None):
        self.logger = logger
        self.login_username = _assert_fixed_account(auth_state)
        self.client = ICSPClient.from_auth_state(auth_state, logger=logger)
        if not self.client.validate_authenticated_session(self.login_username):
            raise RuntimeError("当前登录态已失效，请重新以固定账号登录后重试。")

    def log(self, level: str, message: str) -> None:
        if self.logger:
            self.logger(level, message)

    @staticmethod
    def format_ts(value: Any) -> str:
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def safe_sheet_name(prefix: str, name: str) -> str:
        full = f"{prefix}-{name}"
        if len(full) > 31:
            full = full[:31]
        for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
            full = full.replace(ch, "_")
        return full

    @staticmethod
    def normalize_coupon_name(value: Any) -> str:
        return " ".join(str(value or "").split())

    @staticmethod
    def extract_product_template_id(product: dict[str, Any]) -> str:
        candidates = [
            product.get("couponTemplateId"),
            product.get("couponTempalteId"),
            product.get("templateId"),
            product.get("templateCode"),
        ]
        coupon_result = product.get("couponResult")
        if isinstance(coupon_result, dict):
            candidates.extend(
                [
                    coupon_result.get("couponTemplateId"),
                    coupon_result.get("couponTempalteId"),
                    coupon_result.get("templateId"),
                    coupon_result.get("templateCode"),
                ]
            )
        for candidate in candidates:
            if candidate not in (None, ""):
                return str(candidate).strip()
        return ""

    @staticmethod
    def page_signature(rows: list[dict[str, Any]]) -> tuple[str, ...]:
        signature: list[str] = []
        for row in rows[:5]:
            if not isinstance(row, dict):
                signature.append(str(row))
                continue
            for key in ("id", "userCouponCode", "orderNo", "couponTemplateId", "couponTempalteId"):
                value = row.get(key)
                if value not in (None, ""):
                    signature.append(f"{key}:{value}")
                    break
            else:
                signature.append(str(sorted(row.items()))[:120])
        return tuple(signature)

    def _api_headers(self) -> dict[str, str]:
        return {
            "plazacode": PLAZA_CODE,
            "orgcode": PLAZA_CODE,
            "orgtypecode": ORG_TYPE_CODE,
            "tenantid": TENANT_ID,
            "groupcode": "G001",
            "internalid": "1",
            "vunioncode": "U001",
            "workingorgcode": PLAZA_CODE,
            "userid": str(self.client.user_info.get("userid") or self.client.user_info.get("userId") or ""),
            "usercode": str(self.client.user_info.get("usercode") or ""),
            "username": str(self.client.user_info.get("username") or ""),
            "Content-Type": "application/json;charset=utf-8",
        }

    def _coupon_headers(self) -> dict[str, str]:
        return {"Content-Type": "application/json", "Referer": ICSP_BASE + "/coupon.html"}

    def _product_headers(self) -> dict[str, str]:
        headers = self._api_headers()
        headers.update(
            {
                "orgname": urllib.parse.quote("深圳湾睿印RAIL IN"),
                "orgtypename": urllib.parse.quote("广场"),
                "groupname": urllib.parse.quote("印力商用置业有限公司"),
                "plazaname": PLAZA_CODE,
                "code": PLAZA_CODE,
                "areacode": "",
                "areaname": "null",
                "storecode": "null",
                "storename": "null",
                "Referer": ICSP_BASE + "/dsp.html",
            }
        )
        return headers

    def _post_json(self, url: str, payload: dict[str, Any], *, headers: dict[str, str]) -> tuple[list[dict[str, Any]], int]:
        response = self.client.session.post(url, headers=headers, json=payload, timeout=(10, 30))
        response.raise_for_status()
        return _extract_rows_and_total(response.json())

    def _paginate(
        self,
        url: str,
        payload: dict[str, Any],
        *,
        headers: dict[str, str],
        label: str,
        page_key: str = "pageNo",
        page_size_key: str = "pageSize",
    ) -> list[dict[str, Any]]:
        page = 1
        all_rows: list[dict[str, Any]] = []
        last_signature: tuple[str, ...] | None = None
        while True:
            request_payload = dict(payload)
            request_payload[page_key] = page
            request_payload[page_size_key] = PAGE_SIZE
            rows, total = self._post_json(url, request_payload, headers=headers)
            if not rows:
                break
            signature = self.page_signature(rows)
            if signature == last_signature:
                self.log("WARN", f"{label} 第 {page} 页与上一页重复，已停止翻页")
                break
            last_signature = signature
            all_rows.extend(rows)
            self.log("INFO", f"{label} 第 {page} 页：{len(rows)} 条，累计 {len(all_rows)} 条")
            if total and len(all_rows) >= total:
                break
            if len(rows) < PAGE_SIZE:
                break
            page += 1
            if page > 500:
                self.log("WARN", f"{label} 超过 500 页，停止翻页")
                break
        return all_rows

    def fetch_coupon_products(self, start_date: str, end_date: str) -> list[dict[str, Any]]:
        rows = self._paginate(
            ICSP_BASE + "/yinli-xapi-b/pmp/pointmall/platform/product/productSearch",
            {
                "activityLabel": "",
                "area": "",
                "deliveryType": "ALL",
                "plazza": PLAZA_CODE,
                "point": 0,
                "productCode": "",
                "productName": "",
                "productType": "ALL",
                "exchangeTimeBegin": f"{start_date} 00:00:00",
                "exchangeTimeEnd": f"{end_date} 23:59:59",
                "date": [f"{start_date} 00:00:00", f"{end_date} 23:59:59"],
                "status": "ALL",
                "shelfState": "",
                "goodsTypes": [],
                "category": "",
                "releaseChannel": "",
                "rangeOrgCode": "",
                "belongOrgCode": "",
                "belongOrgName": "",
                "publishingPlatforms": "",
            },
            headers=self._product_headers(),
            label="券商品查询",
            page_key="pageIndex",
        )
        return [row for row in rows if float(row.get("payAmount") or 0) != 0]

    def fetch_coupon_grant_total(self, start_date: str, end_date: str, coupon_template_id: str) -> int:
        response = self.client.session.post(
            ICSP_BASE + "/icsp-coupon/web/user/coupon/list",
            headers=self._coupon_headers(),
            json={
                "beginCreateTime": start_date,
                "endCreateTime": end_date,
                "grantPlazaBuId": GRANT_PLAZA_BU_ID,
                "couponTempalteId": str(coupon_template_id),
                "pageNo": 1,
                "pageSize": 1,
            },
            timeout=(10, 20),
        )
        response.raise_for_status()
        payload = response.json()
        rows, total = _extract_rows_and_total(payload)
        nested = payload.get("data") if isinstance(payload, dict) else None
        total_value = total or payload.get("total") or payload.get("totalCount") or payload.get("totalSize")
        if not total_value and isinstance(nested, dict):
            total_value = nested.get("total") or nested.get("totalCount") or nested.get("totalSize")
        try:
            return int(total_value or len(rows))
        except Exception:
            return len(rows)

    def fetch_coupon_certs_by_template(self, start_date: str, end_date: str, coupon_template_id: str) -> list[dict[str, Any]]:
        return self._paginate(
            ICSP_BASE + "/icsp-coupon/web/user/coupon/certificate/list",
            {
                "beginCreateTime": start_date,
                "endCreateTime": end_date,
                "couponTemplateId": str(coupon_template_id),
            },
            headers=self._coupon_headers(),
            label=f"核销模板 {coupon_template_id}",
        )

    def fetch_coupon_group_user_rows(
        self,
        start_date: str,
        end_date: str,
        coupon_group_no: str | None = None,
        *,
        tenant_id: int = COUPON_GROUP_TENANT_ID,
        app_bu_id: int = COUPON_GROUP_APP_BU_ID,
    ) -> list[dict[str, Any]]:
        payload: dict[str, Any] = {
            "grantPlazaBuId": str(GRANT_PLAZA_BU_ID),
            "tenantId": tenant_id,
            "appBuId": app_bu_id,
            "createStartTime": start_date,
            "createEndTime": end_date,
            "userCouponGroupCodeList": [],
        }
        label = "券包全量扫描"
        if coupon_group_no:
            payload["couponGroupNo"] = str(coupon_group_no)
            label = f"券包 {coupon_group_no}"
        return self._paginate(
            ICSP_BASE + "/icsp-coupon/web/user/coupon/group/list/of/user",
            payload,
            headers=self._coupon_headers(),
            label=label,
        )

    @staticmethod
    def _extract_delivery_types(product: dict[str, Any]) -> list[str]:
        raw_value = product.get("deliveryType")
        if raw_value in (None, ""):
            return []
        values = raw_value if isinstance(raw_value, (list, tuple, set)) else [raw_value]
        return [str(item).strip() for item in values if str(item or "").strip()]

    @classmethod
    def _is_coupon_package_product(cls, product: dict[str, Any]) -> bool:
        return "COUPON_PACKAGE" in cls._extract_delivery_types(product)

    @classmethod
    def _is_coupon_product(cls, product: dict[str, Any]) -> bool:
        delivery_types = cls._extract_delivery_types(product)
        return "COUPON" in delivery_types and "COUPON_PACKAGE" not in delivery_types

    @staticmethod
    def _extract_coupon_group_no(product: dict[str, Any]) -> str:
        candidates: list[Any] = [product.get("couponGroupNo"), product.get("couponGroupCode"), product.get("groupNo")]
        for key in ("couponResult", "couponGroupVO"):
            value = product.get(key)
            if isinstance(value, dict):
                candidates.extend([value.get("couponGroupNo"), value.get("couponGroupCode"), value.get("groupNo")])
        for candidate in candidates:
            text = str(candidate or "").strip()
            if text:
                return text
        return ""

    @staticmethod
    def _extract_spu_code(product: dict[str, Any]) -> str:
        candidates: list[Any] = [
            product.get("spuCode"),
            product.get("productCode"),
            product.get("couponTemplateId"),
            product.get("couponTempalteId"),
        ]
        coupon_result = product.get("couponResult")
        if isinstance(coupon_result, dict):
            candidates.extend(
                [
                    coupon_result.get("spuCode"),
                    coupon_result.get("productCode"),
                    coupon_result.get("couponTemplateId"),
                    coupon_result.get("couponTempalteId"),
                ]
            )
        for candidate in candidates:
            text = str(candidate or "").strip()
            if text:
                return text
        return ""

    @staticmethod
    def _extract_coupon_group_query_value(product: dict[str, Any], field_name: str, default_value: Any) -> Any:
        candidates = [product.get(field_name)]
        for key in ("couponResult", "couponGroupVO"):
            value = product.get(key)
            if isinstance(value, dict):
                candidates.append(value.get(field_name))
        for candidate in candidates:
            if candidate not in (None, ""):
                return candidate
        return default_value

    @staticmethod
    def _normalize_coupon_group_cert_row(row: dict[str, Any]) -> dict[str, Any]:
        normalized = dict(row)
        normalized["_certSource"] = "coupon_group_status_3"
        if not normalized.get("storeName"):
            coupon_group_vo = normalized.get("couponGroupVO")
            coupon_group_title = ""
            if isinstance(coupon_group_vo, dict):
                coupon_group_title = str(coupon_group_vo.get("title", "") or "").strip()
            normalized["storeName"] = str(normalized.get("grantPlazaName", "") or "").strip() or coupon_group_title or "COUPON_PACKAGE"
        return normalized

    @staticmethod
    def _dedupe_coupon_cert_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        deduped_rows: list[dict[str, Any]] = []
        seen_keys: set[str] = set()
        for row in rows:
            key: str | None = None
            for field in ("id", "userCouponCode", "couponCode", "certificateNo"):
                value = row.get(field)
                if value not in (None, ""):
                    key = f"{field}:{value}"
                    break
            if not key:
                deduped_rows.append(row)
                continue
            if key in seen_keys:
                continue
            seen_keys.add(key)
            deduped_rows.append(row)
        return deduped_rows

    @staticmethod
    def _extract_product_grant_count(product: dict[str, Any]) -> int:
        candidates = [product.get("payedNum"), product.get("sendCount"), product.get("sendAmount")]
        effective_num = product.get("effectiveNum")
        refunded_num = product.get("refundedNum")
        if effective_num not in (None, "") or refunded_num not in (None, ""):
            try:
                candidates.append(int(effective_num or 0) + int(refunded_num or 0))
            except Exception:
                pass
        candidates.extend([product.get("soldNum"), product.get("payNum")])
        for candidate in candidates:
            if candidate in (None, ""):
                continue
            try:
                return int(float(candidate))
            except Exception:
                continue
        return 0

    def _extract_coupon_group_match_names(self, row: dict[str, Any]) -> list[str]:
        names: list[str] = []
        coupon_group_vo = row.get("couponGroupVO")
        if isinstance(coupon_group_vo, dict):
            title = self.normalize_coupon_name(coupon_group_vo.get("title", ""))
            if title:
                names.append(title)
        remark = self.normalize_coupon_name(row.get("remark", ""))
        if remark:
            names.append(remark)
            for sep in ("-", "－", "—"):
                if sep in remark:
                    suffix = self.normalize_coupon_name(remark.split(sep, 1)[1])
                    if suffix:
                        names.append(suffix)
                    break
        deduped: list[str] = []
        seen: set[str] = set()
        for name in names:
            if name and name not in seen:
                seen.add(name)
                deduped.append(name)
        return deduped

    def _query_coupon_grant_counts_by_template(
        self, products: list[dict[str, Any]], start_date: str, end_date: str
    ) -> tuple[dict[str, dict[str, Any]], list[str]]:
        grant_map: dict[str, dict[str, Any]] = {}
        missing_template_products: list[str] = []
        for product in products:
            code = str(product.get("productCode", ""))
            grant_map[code] = {"count": 0, "rows": []}
            if not self._is_coupon_product(product):
                continue
            name = str(product.get("productName", ""))
            template_id = self.extract_product_template_id(product)
            template_source = "couponTempalteId"
            if not template_id:
                template_id = self._extract_spu_code(product)
                if template_id:
                    template_source = "spuCode"
            if not template_id:
                missing_template_products.append(name or code)
                grant_map[code] = {"count": self._extract_product_grant_count(product), "rows": []}
                continue
            self.log("INFO", f"按 {template_source}={template_id} 查询发券总数：{name or code}")
            grant_map[code] = {"count": self.fetch_coupon_grant_total(start_date, end_date, template_id), "rows": []}
        return grant_map, missing_template_products

    def _query_coupon_cert_rows_by_template_id(
        self,
        products: list[dict[str, Any]],
        start_date: str,
        end_date: str,
        certs_map: dict[str, list[dict[str, Any]]],
    ) -> tuple[set[str], list[str]]:
        missing_template_products: list[str] = []
        coupon_codes: set[str] = set()
        for product in products:
            if not self._is_coupon_product(product):
                continue
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", "") or "").strip()
            template_id = self.extract_product_template_id(product)
            template_source = "couponTemplateId"
            if not template_id:
                template_id = self._extract_spu_code(product)
                if template_id:
                    template_source = "spuCode"
            if not template_id:
                certs_map[code] = []
                missing_template_products.append(name or code)
                continue
            self.log("INFO", f"按 {template_source}={template_id} 查询核销：{name or code}")
            certs_map[code] = self._dedupe_coupon_cert_rows(self.fetch_coupon_certs_by_template(start_date, end_date, template_id))
            coupon_codes.add(code)
            self.log("INFO", f"核销记录：{name or code} {len(certs_map[code])} 条")
        return coupon_codes, missing_template_products

    def _scan_coupon_package_rows(
        self, products: list[dict[str, Any]], start_date: str, end_date: str
    ) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]], list[str]]:
        package_products = [product for product in products if self._is_coupon_package_product(product)]
        rows_by_code = {str(product.get("productCode", "")): [] for product in package_products}
        code_to_name: dict[str, str] = {}
        name_to_codes: dict[str, set[str]] = defaultdict(set)
        for product in package_products:
            code = str(product.get("productCode", ""))
            name = self.normalize_coupon_name(product.get("productName", ""))
            code_to_name[code] = name or code
            if name:
                name_to_codes[name].add(code)
        if not package_products:
            return rows_by_code, {code: [] for code in rows_by_code}, []
        tenant_id = COUPON_GROUP_TENANT_ID
        app_bu_id = COUPON_GROUP_APP_BU_ID
        for product in package_products:
            tenant_id = self._extract_coupon_group_query_value(product, "tenantId", tenant_id)
            app_bu_id = self._extract_coupon_group_query_value(product, "appBuId", app_bu_id)
            if tenant_id and app_bu_id:
                break
        self.log("INFO", "开始全局扫描券包记录")
        all_rows = self.fetch_coupon_group_user_rows(start_date, end_date, tenant_id=tenant_id, app_bu_id=app_bu_id)
        for row in all_rows:
            matched_codes: set[str] = set()
            for name in self._extract_coupon_group_match_names(row):
                matched_codes.update(name_to_codes.get(name, set()))
            if not matched_codes:
                continue
            normalized_row = self._normalize_coupon_group_cert_row(row)
            for code in matched_codes:
                rows_by_code[code].append(normalized_row)
        cert_rows_by_code: dict[str, list[dict[str, Any]]] = {}
        missing_products: list[str] = []
        for code, rows in rows_by_code.items():
            cert_rows_by_code[code] = [
                row for row in rows if str(row.get("status", "") or "").strip() == COUPON_GROUP_OPEN_STATUS
            ]
            if not rows:
                missing_products.append(code_to_name.get(code, code))
        return rows_by_code, cert_rows_by_code, missing_products

    @staticmethod
    def _split_grant_value(value: Any) -> tuple[int, list[dict[str, Any]]]:
        if isinstance(value, dict):
            return int(value.get("count", 0) or 0), list(value.get("rows", []) or [])
        rows = list(value or [])
        return len(rows), rows

    def _build_excel(
        self,
        products: list[dict[str, Any]],
        grants_map: dict[str, dict[str, Any] | list[dict[str, Any]]],
        certs_map: dict[str, list[dict[str, Any]]],
    ) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)
        grant_sheet_names: dict[str, str] = {}
        cert_sheet_names: dict[str, str] = {}
        used_names: set[str] = set()

        for product in products:
            code = str(product.get("productCode", ""))
            name = str(product.get("productName", ""))
            grant_name = self.safe_sheet_name("发券明细", name)
            cert_name = self.safe_sheet_name("核销明细", name)
            if grant_name in used_names:
                grant_name = grant_name[:28] + f"_{code[-2:]}"
            used_names.add(grant_name)
            grant_sheet_names[code] = grant_name
            if cert_name in used_names:
                cert_name = cert_name[:28] + f"_{code[-2:]}"
            used_names.add(cert_name)
            cert_sheet_names[code] = cert_name

        ws_main = wb.create_sheet("券总表")
        ws_main.append(["券名称", "售价", "发券数量", "核销数量"])
        summary_rows: list[list[Any]] = []
        for product in products:
            code = str(product.get("productCode", ""))
            grant_count, _ = self._split_grant_value(grants_map.get(code, {}))
            cert_count = len(certs_map.get(code, []))
            summary_rows.append([str(product.get("productName", "")), product.get("payAmount", 0), grant_count, cert_count])

        for row_idx, row_data in enumerate(summary_rows, start=2):
            ws_main.append(row_data)
            code = str(products[row_idx - 2].get("productCode", ""))
            if row_data[2]:
                ws_main.cell(row=row_idx, column=3).hyperlink = f"#'{grant_sheet_names[code]}'!A1"
                ws_main.cell(row=row_idx, column=3).font = Font(color="0563C1", underline="single")
            if row_data[3]:
                ws_main.cell(row=row_idx, column=4).hyperlink = f"#'{cert_sheet_names[code]}'!A1"
                ws_main.cell(row=row_idx, column=4).font = Font(color="0563C1", underline="single")
        if summary_rows:
            ws_main.append(["合计", "", sum(int(row[2]) for row in summary_rows), sum(int(row[3]) for row in summary_rows)])
        _style_sheet(ws_main)

        for product in products:
            code = str(product.get("productCode", ""))
            grant_count, grant_rows = self._split_grant_value(grants_map.get(code, {}))
            ws = wb.create_sheet(grant_sheet_names[code])
            ws.append(["返回总表"])
            ws["A1"].hyperlink = "#'券总表'!A1"
            ws["A1"].font = Font(color="0563C1", underline="single")
            ws.append(["发券时间", "券码", "会员", "手机号", "发放渠道"])
            if grant_rows:
                for row in grant_rows:
                    ws.append(
                        [
                            self.format_ts(row.get("grantTime") or row.get("createTime") or row.get("openTime")),
                            str(row.get("userCouponCode") or row.get("userCouponGroupCode") or ""),
                            str(row.get("memberName", "")),
                            str(row.get("phone", "")),
                            str(row.get("grantModeDesc") or row.get("grantChannelName") or ""),
                        ]
                    )
                ws.append([f"合计: {grant_count}", "", "", "", ""])
            else:
                ws.append([f"仅统计数量，发券数: {grant_count}", "", "", "", ""])
            _style_sheet(ws)

        for product in products:
            code = str(product.get("productCode", ""))
            cert_rows = certs_map.get(code, [])
            price = float(product.get("payAmount", 0) or 0)
            ws = wb.create_sheet(cert_sheet_names[code])
            ws.append(["返回总表"])
            ws["A1"].hyperlink = "#'券总表'!A1"
            ws["A1"].font = Font(color="0563C1", underline="single")

            if self._is_coupon_package_product(product):
                ws.append(["统计项", "数量"])
                ws.append(["status=3数量", len(cert_rows)])
                _style_sheet(ws)
                continue

            ws.append(["门店名称", "核销数量", "核销金额"])
            store_counts: dict[str, int] = defaultdict(int)
            for row in cert_rows:
                store_counts[str(row.get("storeName", "") or "未知门店")] += 1
            if store_counts:
                for store_name, count in sorted(store_counts.items(), key=lambda item: (-item[1], item[0])):
                    ws.append([store_name, count, round(count * price, 2)])
                total_count = sum(store_counts.values())
                ws.append(["合计", total_count, round(total_count * price, 2)])
            else:
                ws.append(["无数据", "", ""])
            _style_sheet(ws)
        return wb

    def export(self, start_date: str, end_date: str, output_dir: str | Path, *, file_tag: str | None = None) -> ReconciliationJobResult:
        self.log("INFO", f"当前固定账号：{self.login_username}")
        self.log("INFO", "开始查询券商品。")
        products = self.fetch_coupon_products(start_date, end_date)
        if not products:
            raise RuntimeError("当前日期范围内未查询到有效券商品。")

        full_start = f"{start_date} 00:00:00"
        full_end = f"{end_date} 23:59:59"
        grants_map, missing_grant_template_products = self._query_coupon_grant_counts_by_template(products, full_start, full_end)
        certs_map = {str(product.get("productCode", "")): [] for product in products}
        _, missing_template_products = self._query_coupon_cert_rows_by_template_id(products, full_start, full_end, certs_map)
        package_rows_map, package_cert_rows_map, missing_group_products = self._scan_coupon_package_rows(products, full_start, full_end)
        for code, rows in package_rows_map.items():
            grants_map[code] = {"count": len(rows), "rows": rows}
        certs_map.update(package_cert_rows_map)

        if missing_grant_template_products:
            self.log("WARN", f"缺少模板ID，发券数使用兜底字段：{', '.join(missing_grant_template_products[:5])}")
        if missing_template_products:
            self.log("WARN", f"缺少模板ID，无法查询普通券核销：{', '.join(missing_template_products[:5])}")
        if missing_group_products:
            self.log("WARN", f"券包全局扫描未匹配到记录：{', '.join(missing_group_products[:5])}")

        wb = self._build_excel(products, grants_map, certs_map)
        output_root = Path(output_dir)
        output_root.mkdir(parents=True, exist_ok=True)
        filename = f"coupon_tool_{start_date}_{end_date}_{file_tag or 'export'}.xlsx"
        output_file = output_root / filename
        wb.save(output_file)
        return ReconciliationJobResult(output_file=output_file, result_count=len(products))


class ParkingCaptchaClient:
    def __init__(self, project: ParkingProjectConfig, logger=None, session_cookies: dict[str, str] | None = None):
        self.project = project
        self.logger = logger
        self.last_login_error = ""
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"
                ),
                "Referer": project.referer,
                "Origin": "https://szwryportal.aibee.cn",
                "Content-Type": "application/json;charset=UTF-8",
                "accept": "application/json, text/plain, */*",
            }
        )
        if session_cookies:
            self.session.cookies.update(session_cookies)

    def log(self, level: str, message: str) -> None:
        if self.logger:
            self.logger(level, message)

    @staticmethod
    def md5_encrypt(text: str) -> str:
        return hashlib.md5(text.encode("utf-8")).hexdigest()

    def fetch_captcha(self) -> tuple[str, bytes]:
        response = self.session.get(PARKING_CAPTCHA_URL, timeout=10)
        response.raise_for_status()
        payload = response.json()
        if payload.get("status") != 200:
            raise RuntimeError(str(payload.get("message") or "获取停车验证码失败"))
        request_id = str(payload.get("requestId") or "")
        image_text = str(payload.get("data") or "")
        if not request_id or not image_text:
            raise RuntimeError("停车验证码返回为空")
        return request_id, base64.b64decode(image_text)

    def login(self, captcha_code: str, captcha_uuid: str) -> bool:
        self.log("INFO", f"开始登录停车系统，项目={self.project.label}")
        payload = {
            "username": PARKING_USER,
            "password": self.md5_encrypt(PARKING_PWD_RAW),
            "kaptcha": captcha_code,
        }
        try:
            response = self.session.post(PARKING_LOGIN_URL, json=payload, timeout=15)
            response.raise_for_status()
            result = response.json()
        except Exception as exc:
            self.last_login_error = f"停车系统登录请求异常：{exc}"
            self.log("ERROR", self.last_login_error)
            return False

        if result.get("status") != 200:
            self.last_login_error = str(result.get("message") or "停车系统登录失败")
            self.log("ERROR", f"停车登录失败：{self.last_login_error}")
            return False

        token = str(result.get("data", {}).get("access_token") or "")
        if not token:
            self.last_login_error = "停车系统登录成功，但未返回 access_token"
            self.log("ERROR", "停车登录失败：未返回 access_token")
            return False

        self.session.headers.update(
            {
                "Authorization": f"Bearer {token}",
                "x-app-id": "parkingbi",
                "x-brand-id": "railin",
                "x-project-id": self.project.project_id,
            }
        )
        self.last_login_error = ""
        self.log("SUCCESS", f"停车登录成功，项目={self.project.label}")
        return True

    def _fetch_page(self, page_num: int, start_date: str, end_date: str) -> tuple[list[dict[str, Any]], int]:
        payload = {
            "entity_id": "",
            "page_offset": page_num,
            "page_size": PAGE_SIZE,
            "filter_body": {
                "start_order_time": f"{start_date} 00:00:00",
                "end_order_time": f"{end_date} 23:59:59",
                "reservation_number": "",
                "reservation_people": "",
                "phone": "",
                "car_plate": "",
                "parking_space": "",
                "start_arrival_time": "",
                "end_arrival_time": "",
                "order_status": "",
                "order_type": "",
            },
        }
        response = self.session.post(PARKING_DATA_URL, json=payload, timeout=20)
        response.raise_for_status()
        data_json = response.json()
        if data_json.get("status") not in (None, 0, 200, "0", "200"):
            raise RuntimeError(str(data_json.get("message") or "停车记录查询失败"))
        data = data_json.get("data")
        if isinstance(data, dict):
            rows = data.get("list") or data.get("records") or []
            total = int(data.get("total") or 0)
            return rows, total
        return [], 0

    def fetch_all_orders(self, start_date: str, end_date: str) -> list[dict[str, Any]]:
        first_rows, total = self._fetch_page(1, start_date, end_date)
        if not first_rows:
            return []
        all_rows = list(first_rows)
        total_pages = max(1, math.ceil(total / PAGE_SIZE))
        for page in range(2, total_pages + 1):
            rows, _ = self._fetch_page(page, start_date, end_date)
            if not rows:
                break
            all_rows.extend(rows)
        self.log("INFO", f"停车记录抓取完成，共 {len(all_rows)} 条")
        return all_rows


def fetch_parking_captcha(project_key: str | None, logger=None) -> tuple[str, bytes]:
    project = get_parking_project(project_key)
    if not project.enable_parking:
        raise RuntimeError(f"项目 {project.label} 当前不需要停车验证码。")
    client = ParkingCaptchaClient(project=project, logger=logger)
    captcha_uuid, image_bytes = client.fetch_captcha()
    parking_captcha_challenge_store.save(
        project_key=project.key,
        captcha_uuid=captcha_uuid,
        cookies=export_captcha_session_cookies(client.session),
    )
    return captcha_uuid, image_bytes


def create_parking_client_from_captcha(project_key: str | None, captcha_uuid: str, logger=None) -> ParkingCaptchaClient:
    project = get_parking_project(project_key)
    challenge = parking_captcha_challenge_store.get(captcha_uuid.strip())
    if challenge is None:
        raise RuntimeError("停车验证码已过期，请刷新后重试。")
    if challenge.project_key != project.key:
        raise RuntimeError("停车验证码与当前项目不匹配，请刷新后重试。")
    return ParkingCaptchaClient(project=project, logger=logger, session_cookies=challenge.cookies)


def _build_dz_workbook(start_date: str, end_date: str) -> tuple[Workbook, int, dict[str, Any]]:
    with SessionLocal() as session:
        dashboard = BiAnalyticsService(session).build_dashboard(
            start_date=date.fromisoformat(start_date),
            end_date=date.fromisoformat(end_date),
            mode="daily",
        )

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "对账总览"
    ws_summary.append([
        "日期",
        "停车记录数",
        "停车匹配手机号数",
        "停车匹配会员数",
        "交易笔数",
        "交易金额(元)",
        "积分流水笔数",
        "消费金额(元)",
        "正向积分",
        "负向积分",
    ])
    total_rows = 0
    for item in dashboard["daily_series"]:
        ws_summary.append([
            item.get("date", ""),
            item.get("parking_count", 0),
            item.get("matched_mobile_count", 0),
            item.get("matched_member_count", 0),
            item.get("trade_count", 0),
            item.get("trade_amount_yuan", 0),
            item.get("point_flow_count", 0),
            item.get("consume_amount_yuan", 0),
            item.get("positive_points", 0),
            item.get("negative_points", 0),
        ])
        total_rows += 1
    _style_sheet(ws_summary)

    ws_plaza = wb.create_sheet("项目联动汇总")
    ws_plaza.append(["项目", "停车记录数", "交易笔数", "交易金额(元)", "积分流水笔数", "消费金额(元)", "匹配会员数"])
    for item in dashboard["plaza_ranking"]:
        ws_plaza.append([
            item.get("plaza_name", ""),
            item.get("parking_count", 0),
            item.get("trade_count", 0),
            item.get("trade_amount_yuan", 0),
            item.get("point_flow_count", 0),
            item.get("consume_amount_yuan", 0),
            item.get("matched_member_count", 0),
        ])
    _style_sheet(ws_plaza)

    ws_level = wb.create_sheet("会员等级分布")
    ws_level.append(["会员等级", "会员数", "停车会员数", "交易会员数", "积分会员数"])
    for item in dashboard["level_distribution"]:
        ws_level.append([
            item.get("level_name", ""),
            item.get("member_count", 0),
            item.get("parking_members", 0),
            item.get("trade_members", 0),
            item.get("point_members", 0),
        ])
    _style_sheet(ws_level)

    ws_validation = wb.create_sheet("数据验证")
    ws_validation.append(["指标", "结果", "说明"])
    for item in dashboard["validation_metrics"]:
        ws_validation.append([item.get("metric", ""), item.get("value", ""), item.get("description", "")])
    _style_sheet(ws_validation)

    total_rows += len(dashboard["plaza_ranking"]) + len(dashboard["level_distribution"]) + len(dashboard["validation_metrics"])
    return wb, total_rows, dashboard


def _append_parking_sheet(wb: Workbook, parking_rows: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("停车原始记录")
    headers = list(PARKING_HEADER_MAP.values())
    keys = list(PARKING_HEADER_MAP.keys())
    ws.append(headers)
    if not parking_rows:
        ws.append(["暂无数据"])
        _style_sheet(ws)
        return 0
    for row in parking_rows:
        ws.append([_format_export_datetime(row.get(key)) for key in keys])
    _style_sheet(ws)
    return len(parking_rows)


def _append_wechat_summary_sheet(
    wb: Workbook,
    daily_series: list[dict[str, Any]],
    fund_daily: dict[str, dict[tuple[str, str], dict[str, float]]],
    refund_fee_daily: dict[str, float],
    trade_daily: dict[str, float],
) -> int:
    ws = wb.create_sheet("微信账单汇总")
    ws.append(["日期", "微信交易订单金额(元)", "微信资金账单金额(元)", "退款手续费(元)", "停车交易金额(元)"])
    row_count = 0
    for item in daily_series:
        biz_date = str(item.get("date") or "")
        if not biz_date:
            continue
        fund_amount = 0.0
        for summary in fund_daily.get(biz_date, {}).values():
            fund_amount += float(summary.get("amount") or 0)
        ws.append([
            biz_date,
            round(float(trade_daily.get(biz_date) or 0), 2),
            round(fund_amount, 2),
            round(float(refund_fee_daily.get(biz_date) or 0), 2),
            round(float(item.get("trade_amount_yuan") or 0), 2),
        ])
        row_count += 1
    _style_sheet(ws)
    return row_count


def _append_wechat_fund_sheet(wb: Workbook, fund_rows: list[dict[str, Any]]) -> int:
    ws = wb.create_sheet("微信资金账单")
    if not fund_rows:
        ws.append(["暂无数据"])
        _style_sheet(ws)
        return 0
    headers = list(fund_rows[0].keys())
    ws.append(headers)
    for row in fund_rows:
        ws.append([row.get(header, "") for header in headers])
    _style_sheet(ws)
    return len(fund_rows)


def run_new_icsp_dz_export(
    auth_state: dict[str, Any],
    start_date: str,
    end_date: str,
    output_dir: str | Path,
    *,
    project_key: str | None = None,
    captcha_code: str | None = None,
    captcha_uuid: str | None = None,
    upload_root: str | Path,
    upload_session_id: str | None = None,
    file_tag: str | None = None,
    logger=None,
) -> ReconciliationJobResult:
    login_username = _assert_fixed_account(auth_state)
    if logger:
        logger("INFO", f"已验证固定账号：{login_username}")

    fund_csv_path, trade_csv_path = resolve_uploaded_wechat_csvs(upload_root, upload_session_id)
    if logger:
        logger("INFO", f"已加载微信支付资金账单：{fund_csv_path.name}")
        logger("INFO", f"已加载微信支付交易订单：{trade_csv_path.name}")
    fund_rows_raw, fund_daily, refund_fee_daily = parse_wechat_fund_csv(fund_csv_path, logger)
    wechat_trade_daily = parse_wechat_trade_csv(trade_csv_path, logger)

    project = get_parking_project(project_key)
    parking_rows: list[dict[str, Any]] = []
    if project.enable_parking:
        if not captcha_code or not captcha_uuid:
            raise RuntimeError("当前项目需要停车验证码，请先获取并输入验证码。")
        parking_client = create_parking_client_from_captcha(project.key, captcha_uuid.strip(), logger=logger)
        if not parking_client.login(captcha_code.strip(), captcha_uuid.strip()):
            raise RuntimeError(parking_client.last_login_error or "停车系统登录失败，请检查验证码后重试。")
        parking_rows = parking_client.fetch_all_orders(start_date, end_date)
    elif logger:
        logger("INFO", f"项目 {project.label} 当前不启用停车系统对账。")

    wb, result_count, dashboard = _build_dz_workbook(start_date, end_date)
    result_count += _append_parking_sheet(wb, parking_rows)
    result_count += _append_wechat_summary_sheet(wb, dashboard.get("daily_series", []), fund_daily, refund_fee_daily, wechat_trade_daily)
    result_count += _append_wechat_fund_sheet(wb, fund_rows_raw)

    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    filename = f"new_icsp_dz_{start_date}_{end_date}_{file_tag or 'export'}.xlsx"
    output_file = output_root / filename
    wb.save(output_file)
    if logger:
        logger("SUCCESS", f"对账工具导出完成：{output_file.name}")
    return ReconciliationJobResult(output_file=output_file, result_count=result_count)


def run_coupon_tool_export(
    auth_state: dict[str, Any],
    start_date: str,
    end_date: str,
    output_dir: str | Path,
    *,
    file_tag: str | None = None,
    logger=None,
) -> ReconciliationJobResult:
    exporter = CouponToolExporter(auth_state=auth_state, logger=logger)
    return exporter.export(start_date, end_date, output_dir, file_tag=file_tag)
