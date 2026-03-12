from __future__ import annotations

import base64
import json
import math
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Sequence

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ICSP_BASE = "https://icsp.scpgroup.com.cn"
ICSP_CLIENT_ID = "2a5c64fcf8cf475593350a6d11548711"
ICSP_SALT = "d0a8155e8e84e5832c3a908056737c2b"

PLAZA_CODE = "G002Z008C0030"
TENANT_ID = "10000"
ORG_TYPE_CODE = "10003"
PLAZA_BU_ID = 293

POINT_FLOW_URL = ICSP_BASE + "/icsp-point/web/point/water/flow/queryList"
PAGE_SIZE = 100
MAX_PAGE_WORKERS = 8
DEFAULT_FIELD_FILES = ("point_flow_fields.json", "point_flow_fields.txt")
PROJECT_ROOT = Path(__file__).resolve().parents[2]

LoggerCallback = Callable[[str, str], None]
StopChecker = Callable[[], None]


@dataclass(slots=True)
class ExportJobResult:
    output_file: Path
    result_count: int


@dataclass(slots=True)
class ICSPAuthResult:
    username: str
    display_name: str
    user_id: str
    user_code: str
    auth_state: dict[str, Any]


def _read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return path.read_text(encoding=encoding)
        except Exception:
            continue
    return ""


def load_fields_from_sample(base_dir: str | Path | None = None) -> list[str]:
    root = Path(base_dir) if base_dir is not None else PROJECT_ROOT
    candidates = [root / name for name in DEFAULT_FIELD_FILES]

    for path in sorted(root.iterdir()):
        if path.is_file() and path.suffix.lower() in {".json", ".txt"}:
            stem = path.stem.lower()
            if "flow" in stem or "field" in stem:
                candidates.append(path)

    for sample_path in candidates:
        if not sample_path.is_file():
            continue
        try:
            raw = _read_text_with_fallback(sample_path).strip()
            if not raw:
                continue
            sample = json.loads(raw.rstrip(","))
            if isinstance(sample, dict):
                return list(sample.keys())
            if isinstance(sample, list) and all(isinstance(item, str) for item in sample):
                return sample
        except json.JSONDecodeError:
            lines = [line.strip().strip(",") for line in raw.splitlines() if line.strip()]
            if lines:
                return lines
        except Exception:
            continue
    return []


def format_cell_value(key: str, value):
    if value is None:
        return ""
    if key == "consumeAmount":
        try:
            return round(float(value) / 100, 2)
        except Exception:
            return value
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if "time" in key.lower() and isinstance(value, (int, float)) and value > 10**12:
        try:
            return datetime.fromtimestamp(value / 1000).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(value)
    return value


class ICSPClient:
    def __init__(
        self,
        logger: LoggerCallback | None = None,
        stop_checker: StopChecker | None = None,
    ):
        self.logger = logger
        self.stop_checker = stop_checker
        self.session = requests.Session()
        self.user_info = {"userid": "", "usercode": "", "username": ""}
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (X11; Linux x86_64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36"
                ),
                "Accept": "application/json, text/plain, */*",
                "Origin": ICSP_BASE,
                "Referer": ICSP_BASE + "/login.html",
            }
        )

    def log(self, level: str, message: str) -> None:
        if not self.logger:
            return
        try:
            self.logger(level, message)
        except Exception:
            return

    def check_stop(self) -> None:
        if self.stop_checker:
            self.stop_checker()

    @staticmethod
    def _make_password(plain_password: str) -> str:
        combined = (ICSP_SALT + plain_password).encode("utf-8")
        encoded = base64.b64encode(combined).decode()
        return f"{encoded}.{ICSP_SALT}"

    def login(self, username: str, password: str) -> bool:
        self.check_stop()
        payload = {
            "clientId": ICSP_CLIENT_ID,
            "passwd": self._make_password(password),
            "user": username,
        }
        headers = {"Content-Type": "application/x-www-form-urlencoded; charset=UTF-8"}
        self.log("INFO", f"[ICSP] logging in as {username}")

        try:
            auth_resp = self.session.post(
                ICSP_BASE + "/icsp-permission/web/permission/sso/auth/authCode",
                data=payload,
                headers=headers,
                allow_redirects=False,
                timeout=15,
            )
            self.check_stop()
            auth_code = ""
            if auth_resp.status_code == 302 and "Location" in auth_resp.headers:
                auth_code = auth_resp.headers["Location"].split("authCode=")[-1]
            elif auth_resp.status_code == 200:
                body = auth_resp.json()
                if body.get("success") and body.get("data"):
                    auth_code = str(body["data"])

            if not auth_code:
                self.log("ERROR", "[ICSP] authCode not found, login failed")
                return False

            timestamp = str(int(time.time() * 1000))
            self.session.get(f"{ICSP_BASE}/auth.html?authCode={auth_code}", timeout=15)
            self.check_stop()
            self.session.get(
                f"{ICSP_BASE}/icsp-permission/web/wd/login/login/sso?_t={timestamp}&authCode={auth_code}",
                timeout=15,
            )
            self.check_stop()

            user_resp = self.session.get(
                f"{ICSP_BASE}/icsp-employee/web/login/query/v2?_t={timestamp}",
                timeout=15,
            )
            if user_resp.status_code == 200:
                data = user_resp.json().get("data", {})
                self.user_info["userid"] = str(data.get("id", ""))
                self.user_info["usercode"] = str(data.get("loginCode", username))
                self.user_info["username"] = urllib.parse.quote(str(data.get("userName", "")))

            self.log("SUCCESS", "[ICSP] login succeeded")
            return True
        except InterruptedError:
            raise
        except Exception as exc:
            self.log("ERROR", f"[ICSP] login failed: {exc}")
            return False

    def query_current_user(self, login_username: str = "") -> bool:
        self.check_stop()
        timestamp = str(int(time.time() * 1000))

        try:
            response = self.session.get(
                f"{ICSP_BASE}/icsp-employee/web/login/query/v2?_t={timestamp}",
                timeout=15,
            )
            response.raise_for_status()
            payload = response.json()
            data = payload.get("data", {}) if isinstance(payload, dict) else {}
            if not isinstance(data, dict):
                data = {}

            user_id = str(data.get("id", "")).strip()
            user_code = str(data.get("loginCode", login_username or "")).strip()
            user_name = str(data.get("userName", "")).strip()

            if not user_id:
                return False

            self.user_info["userid"] = user_id
            self.user_info["usercode"] = user_code or login_username
            self.user_info["username"] = urllib.parse.quote(user_name or login_username)
            return True
        except InterruptedError:
            raise
        except Exception as exc:
            self.log("WARN", f"[ICSP] failed to query current user info: {exc}")
            return False

    def has_serializable_auth_state(self) -> bool:
        return any(cookie.name and cookie.value for cookie in self.session.cookies)

    def ensure_authenticated_session(self, login_username: str = "") -> bool:
        self.check_stop()
        if self.user_info.get("userid") and self.user_info.get("usercode") and self.user_info.get("username"):
            return True

        self.log("INFO", "[ICSP] resolving authenticated session context")
        if self.query_current_user(login_username):
            return True

        return False

    def validate_authenticated_session(self, login_username: str = "") -> bool:
        self.log("INFO", "[ICSP] validating recovered authenticated session")
        if not self.has_serializable_auth_state():
            self.log("WARN", "[ICSP] no reusable authentication cookies found")
            return False
        return self.ensure_authenticated_session(login_username)

    def get_profile(self, login_username: str) -> dict[str, str]:
        display_name = urllib.parse.unquote(self.user_info.get("username", "")) or login_username
        return {
            "username": login_username,
            "display_name": display_name,
            "user_id": self.user_info.get("userid", ""),
            "user_code": self.user_info.get("usercode", login_username),
        }

    def export_auth_state(self, login_username: str) -> dict[str, Any]:
        serialized_cookies = [
            {
                "name": cookie.name,
                "value": cookie.value,
                "domain": cookie.domain,
                "path": cookie.path,
                "secure": cookie.secure,
                "expires": cookie.expires,
            }
            for cookie in self.session.cookies
        ]
        return {
            "login_username": login_username,
            "authenticated_at": datetime.now().isoformat(),
            "user_info": dict(self.user_info),
            "cookies": serialized_cookies,
        }

    @classmethod
    def from_session(
        cls,
        session: requests.Session,
        user_info: dict[str, Any] | None = None,
        logger: LoggerCallback | None = None,
        stop_checker: StopChecker | None = None,
    ) -> "ICSPClient":
        client = cls(logger=logger, stop_checker=stop_checker)
        client.session.headers.update(session.headers)
        client.session.cookies.update(session.cookies)
        source_user_info = user_info or {}
        client.user_info = {
            "userid": str(source_user_info.get("userid", "")),
            "usercode": str(source_user_info.get("usercode", "")),
            "username": str(source_user_info.get("username", "")),
        }
        return client

    @classmethod
    def from_auth_state(
        cls,
        auth_state: dict[str, Any],
        logger: LoggerCallback | None = None,
        stop_checker: StopChecker | None = None,
    ) -> "ICSPClient":
        client = cls(logger=logger, stop_checker=stop_checker)
        client.user_info = {
            "userid": str((auth_state.get("user_info") or {}).get("userid", "")),
            "usercode": str((auth_state.get("user_info") or {}).get("usercode", "")),
            "username": str((auth_state.get("user_info") or {}).get("username", "")),
        }
        for item in auth_state.get("cookies") or []:
            if not isinstance(item, dict):
                continue
            cookie = requests.cookies.create_cookie(
                name=str(item.get("name", "")),
                value=str(item.get("value", "")),
                domain=item.get("domain"),
                path=str(item.get("path") or "/"),
                secure=bool(item.get("secure", False)),
                expires=item.get("expires"),
            )
            client.session.cookies.set_cookie(cookie)
        return client

    def _api_headers(self) -> dict[str, str]:
        return {
            "plazacode": PLAZA_CODE,
            "orgcode": PLAZA_CODE,
            "orgtypecode": ORG_TYPE_CODE,
            "tenantid": TENANT_ID,
            "groupcode": "G001",
            "internalid": "1",
            "vunioncode": "U001",
            "workingorgcode": PLAZA_CODE,
            "userid": self.user_info["userid"],
            "usercode": self.user_info["usercode"],
            "username": self.user_info["username"],
            "Content-Type": "application/json;charset=utf-8",
            "Referer": ICSP_BASE + "/scpg.html",
            "Accept": "*/*",
            "accept-language": "zh-CN",
        }

    @staticmethod
    def _extract_rows_and_total(payload: dict) -> tuple[list[dict], int]:
        if not isinstance(payload, dict):
            return [], 0

        rows: list[dict] = []
        total = 0

        if isinstance(payload.get("rows"), list):
            rows = payload["rows"]
        elif isinstance(payload.get("list"), list):
            rows = payload["list"]
        elif isinstance(payload.get("resultList"), list):
            rows = payload["resultList"]
        elif isinstance(payload.get("data"), list):
            rows = payload["data"]
        elif isinstance(payload.get("data"), dict):
            nested = payload["data"]
            if isinstance(nested.get("rows"), list):
                rows = nested["rows"]
            elif isinstance(nested.get("list"), list):
                rows = nested["list"]
            elif isinstance(nested.get("resultList"), list):
                rows = nested["resultList"]
            elif isinstance(nested.get("records"), list):
                rows = nested["records"]
            total = int(nested.get("total") or nested.get("totalCount") or nested.get("totalSize") or 0)

        if not total:
            total = int(payload.get("total") or payload.get("totalCount") or payload.get("totalSize") or 0)
        if not total and rows:
            total = len(rows)
        return rows, total

    def _build_worker_session(self) -> requests.Session:
        worker_session = requests.Session()
        worker_session.headers.update(self.session.headers)
        worker_session.cookies.update(self.session.cookies)
        return worker_session

    def _fetch_point_page(
        self,
        page_no: int,
        start_date: str,
        end_date: str,
        session: requests.Session | None = None,
    ) -> tuple[list[dict], int]:
        self.check_stop()
        if session is not None:
            use_session = session
        elif page_no == 1:
            use_session = self.session
        else:
            use_session = self._build_worker_session()

        payload = {
            "pageNo": page_no,
            "pageSize": PAGE_SIZE,
            "plazaBuId": PLAZA_BU_ID,
            "createStartTime": f"{start_date} 00:00:00",
            "createEndTime": f"{end_date} 23:59:59",
            "fromWeb": 1,
        }
        response = use_session.post(
            POINT_FLOW_URL,
            headers=self._api_headers(),
            json=payload,
            timeout=20,
        )
        response.raise_for_status()
        return self._extract_rows_and_total(response.json())

    def fetch_point_flow(self, start_date: str, end_date: str) -> list[dict]:
        first_rows, total = self._fetch_point_page(1, start_date, end_date)
        if not first_rows:
            self.log("WARN", "[points-flow] no data returned")
            return []

        all_rows: list[dict] = list(first_rows)
        self.log("INFO", f"[points-flow] first page rows={len(first_rows)}, total={total}")

        if total and total > PAGE_SIZE:
            total_pages = math.ceil(total / PAGE_SIZE)
            workers = min(MAX_PAGE_WORKERS, max(1, total_pages - 1))
            self.log(
                "INFO",
                f"[points-flow] fetching remaining pages concurrently, pages={total_pages}, workers={workers}",
            )

            tasks = list(range(2, total_pages + 1))
            with ThreadPoolExecutor(max_workers=workers) as executor:
                futures = {
                    executor.submit(self._fetch_point_page, page, start_date, end_date): page for page in tasks
                }
                done = 0
                for future in as_completed(futures):
                    self.check_stop()
                    page = futures[future]
                    try:
                        rows, _ = future.result()
                        if rows:
                            all_rows.extend(rows)
                    except Exception as exc:
                        self.log("WARN", f"[points-flow] page {page} failed: {exc}")
                    done += 1
                    if done % 5 == 0 or done == len(futures):
                        self.log(
                            "INFO",
                            f"[points-flow] completed pages={done}/{len(futures)}, rows={len(all_rows)}",
                        )
        else:
            page_no = 2
            last_size = len(first_rows)
            while last_size >= PAGE_SIZE and page_no <= 2000:
                self.check_stop()
                rows, _ = self._fetch_point_page(page_no, start_date, end_date)
                if not rows:
                    break
                all_rows.extend(rows)
                last_size = len(rows)
                self.log("INFO", f"[points-flow] fetched page {page_no}, rows={len(all_rows)}")
                page_no += 1

        self.log("SUCCESS", f"[points-flow] fetch completed, total rows={len(all_rows)}")
        return all_rows


def export_to_excel(
    rows: list[dict],
    fields: Sequence[str],
    start_date: str,
    end_date: str,
    output_dir: str | Path,
    file_tag: str | None = None,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "points_flow"

    field_list = list(fields)
    if not field_list:
        keys: set[str] = set()
        for row in rows:
            keys.update(row.keys())
        field_list = sorted(keys)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    worksheet.append(field_list)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    for row in rows:
        worksheet.append([format_cell_value(key, row.get(key)) for key in field_list])

    for row_cells in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(field_list),
    ):
        for cell in row_cells:
            cell.border = thin
            cell.alignment = center

    worksheet.freeze_panes = "A2"
    for col_idx, header in enumerate(field_list, start=1):
        column = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_num in range(2, worksheet.max_row + 1):
            value = worksheet[f"{column}{row_num}"].value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        worksheet.column_dimensions[column].width = min(max_len * 1.2 + 4, 60)

    export_root = Path(output_dir)
    export_root.mkdir(parents=True, exist_ok=True)

    suffix = f"_{file_tag}" if file_tag else ""
    file_name = f"points_flow_{start_date}_{end_date}{suffix}.xlsx"
    output_path = export_root / file_name
    workbook.save(output_path)
    return output_path


def run_points_flow_export(
    username: str,
    password: str,
    start_date: str,
    end_date: str,
    output_dir: str | Path,
    logger: LoggerCallback | None = None,
    stop_checker: StopChecker | None = None,
    file_tag: str | None = None,
) -> ExportJobResult:
    if logger:
        logger("INFO", "Starting login to ICSP.")
    client = ICSPClient(logger=logger, stop_checker=stop_checker)
    if not client.login(username, password):
        raise RuntimeError("ICSP \u767b\u5f55\u5931\u8d25\uff0c\u8bf7\u68c0\u67e5\u8d26\u53f7\u6216\u5bc6\u7801\u3002")

    if logger:
        logger("INFO", "Starting points flow data fetch.")
    rows = client.fetch_point_flow(start_date, end_date)

    if logger:
        logger("INFO", "Starting Excel export.")
    fields = load_fields_from_sample()
    output_file = export_to_excel(
        rows=rows,
        fields=fields,
        start_date=start_date,
        end_date=end_date,
        output_dir=output_dir,
        file_tag=file_tag,
    )
    if logger:
        logger("SUCCESS", f"Excel export completed: {output_file.name}")
    return ExportJobResult(output_file=output_file, result_count=len(rows))


def authenticate_icsp_user(
    username: str,
    password: str,
    logger: LoggerCallback | None = None,
) -> ICSPAuthResult:
    client = ICSPClient(logger=logger)
    if not client.login(username, password):
        raise RuntimeError("ICSP \u767b\u5f55\u5931\u8d25\uff0c\u8bf7\u68c0\u67e5\u8d26\u53f7\u6216\u5bc6\u7801\u3002")
    if not client.has_serializable_auth_state():
        raise RuntimeError(
            "ICSP \u767b\u5f55\u6210\u529f\uff0c\u4f46\u672a\u83b7\u53d6\u5230\u53ef\u590d\u7528\u7684\u8ba4\u8bc1\u4fe1\u606f\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5\u3002"
        )

    serialized_auth_state = client.export_auth_state(username)
    recovered_client = ICSPClient.from_auth_state(serialized_auth_state, logger=logger)
    if not recovered_client.validate_authenticated_session(username):
        raise RuntimeError(
            "ICSP \u767b\u5f55\u6210\u529f\uff0c\u4f46\u672a\u80fd\u5efa\u7acb\u53ef\u590d\u7528\u4f1a\u8bdd\uff0c\u8bf7\u7a0d\u540e\u91cd\u8bd5\u3002"
        )

    profile = recovered_client.get_profile(username)
    return ICSPAuthResult(
        username=profile["username"],
        display_name=profile["display_name"],
        user_id=profile["user_id"],
        user_code=profile["user_code"],
        auth_state=recovered_client.export_auth_state(profile["username"]),
    )


def run_points_flow_export_with_auth_state(
    auth_state: dict[str, Any],
    start_date: str,
    end_date: str,
    output_dir: str | Path,
    logger: LoggerCallback | None = None,
    stop_checker: StopChecker | None = None,
    file_tag: str | None = None,
) -> ExportJobResult:
    client = ICSPClient.from_auth_state(auth_state, logger=logger, stop_checker=stop_checker)
    login_username = str(auth_state.get("login_username", "")).strip()
    if not client.validate_authenticated_session(login_username):
        raise RuntimeError("\u767b\u5f55\u5931\u6548\uff0c\u8bf7\u91cd\u65b0\u767b\u5f55")

    if logger:
        logger("INFO", "Using authenticated ICSP session.")
        logger("INFO", "Starting points flow data fetch.")
    rows = client.fetch_point_flow(start_date, end_date)

    if logger:
        logger("INFO", "Starting Excel export.")
    fields = load_fields_from_sample()
    output_file = export_to_excel(
        rows=rows,
        fields=fields,
        start_date=start_date,
        end_date=end_date,
        output_dir=output_dir,
        file_tag=file_tag,
    )
    if logger:
        logger("SUCCESS", f"Excel export completed: {output_file.name}")
    return ExportJobResult(output_file=output_file, result_count=len(rows))


def refresh_icsp_auth_state(
    auth_state: dict[str, Any],
    logger: LoggerCallback | None = None,
) -> ICSPAuthResult:
    login_username = str(auth_state.get("login_username", "")).strip()
    client = ICSPClient.from_auth_state(auth_state, logger=logger)
    if not client.validate_authenticated_session(login_username):
        raise RuntimeError("\u767b\u5f55\u5931\u6548\uff0c\u8bf7\u91cd\u65b0\u767b\u5f55")

    profile = client.get_profile(login_username or client.user_info.get("usercode", ""))
    refreshed_auth_state = client.export_auth_state(profile["username"])
    return ICSPAuthResult(
        username=profile["username"],
        display_name=profile["display_name"],
        user_id=profile["user_id"],
        user_code=profile["user_code"],
        auth_state=refreshed_auth_state,
    )


class PointsFlowExportService:
    def authenticate_user(
        self,
        username: str,
        password: str,
        log_callback: LoggerCallback | None = None,
    ) -> ICSPAuthResult:
        if log_callback:
            log_callback("INFO", "Starting ICSP login verification.")
        return authenticate_icsp_user(
            username=username,
            password=password,
            logger=log_callback,
        )

    def refresh_authenticated_session(
        self,
        auth_state: dict[str, Any],
        log_callback: LoggerCallback | None = None,
    ) -> ICSPAuthResult:
        return refresh_icsp_auth_state(
            auth_state=auth_state,
            logger=log_callback,
        )

    def run_export(
        self,
        task_id: str,
        auth_state: dict[str, Any],
        start_date: str,
        end_date: str,
        export_dir: str | Path,
        log_callback: LoggerCallback | None = None,
    ) -> ExportJobResult:
        file_tag = task_id[:8] if task_id else None
        return run_points_flow_export_with_auth_state(
            auth_state=auth_state,
            start_date=start_date,
            end_date=end_date,
            output_dir=export_dir,
            logger=log_callback,
            file_tag=file_tag,
        )
