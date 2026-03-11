#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import base64
import os
import json
import re
import shlex
import sys
import time
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import aiohttp
from openpyxl import load_workbook


ENCRYPTION_KEY = "c67b8c67558a2e609912a532e45ebb30"
DEFAULT_CURL_FILE = "车牌查会员.txt"
DEFAULT_CACHE_FILE = ".member_lookup_cache.json"
DEFAULT_CONCURRENCY = 0
DEFAULT_AUTO_CONCURRENCY_CAP = 16
DEFAULT_TIMEOUT = 20.0
DEFAULT_RETRIES = 5
RETRYABLE_HTTP_STATUSES = {429, 500, 502, 503, 504}
AUTH_ERROR_PATTERNS = (
    "HTTP 401:",
    '"code":"401"',
    '"code": "401"',
    "登录状态已失效",
    "401 未授权",
)
PLATE_HEADER_CANDIDATES = ("车牌号", "车牌", "plate", "plate_no", "plateNo")


@dataclass
class LookupResult:
    plate: str
    mobile: str = ""
    level_name: str = ""
    matched_plate: str = ""
    status: str = ""
    error: str = ""

    def cacheable(self) -> bool:
        return self.status in {"success", "not_found"}


class TokenExpiredError(RuntimeError):
    pass


@dataclass
class PlateExtraction:
    unique_plates: list[str]
    plate_to_rows: dict[str, list[int]]
    empty_row_indexes: list[int]
    total_rows: int


@dataclass
class MemberLookupTemplate:
    scheme: str
    netloc: str
    path: str
    fragment: str
    headers: dict[str, str]
    params: list[tuple[str, str]]
    matching_key: str
    token_expire_at: int | None = None

    @property
    def base_url(self) -> str:
        return urlunsplit((self.scheme, self.netloc, self.path, "", self.fragment))

    def build_url(self, plate: str) -> str:
        encrypted_pairs: list[tuple[str, str]] = []
        for key, value in self.params:
            actual_value = plate if key == self.matching_key else value
            encrypted_pairs.append((encrypt_token(key), encrypt_token(str(actual_value))))
        query = urlencode(encrypted_pairs)
        return urlunsplit((self.scheme, self.netloc, self.path, query, self.fragment))


@dataclass
class ProgressPrinter:
    label: str
    total: int
    unit: str = "项"
    min_interval: float = 0.5
    current: int = 0
    started_at: float = 0.0
    last_print_at: float = 0.0

    def __post_init__(self) -> None:
        self.started_at = time.perf_counter()

    def update(self, step: int = 1, *, force: bool = False, extra: str = "") -> None:
        self.current += step
        self._print(force=force, extra=extra)

    def finish(self, extra: str = "") -> None:
        self.current = self.total
        self._print(force=True, extra=extra)

    def _print(self, *, force: bool, extra: str) -> None:
        now = time.perf_counter()
        if not force and self.current < self.total and now - self.last_print_at < self.min_interval:
            return
        self.last_print_at = now
        elapsed = max(now - self.started_at, 1e-6)
        done = min(self.current, self.total) if self.total > 0 else self.current
        if self.total > 0:
            percent = done / self.total * 100
            rate = done / elapsed
            remain = max(self.total - done, 0) / rate if rate > 0 else 0
            message = (
                f"\r[{self.label}] {done}/{self.total} ({percent:5.1f}%)"
                f" | {rate:6.1f}{self.unit}/s"
                f" | 已耗时 {format_duration(elapsed)}"
                f" | 预计剩余 {format_duration(remain)}"
            )
        else:
            message = f"\r[{self.label}] 0/{self.total} | 已耗时 {format_duration(elapsed)}"
        if extra:
            message += f" | {extra}"
        print(message, file=sys.stderr, end="", flush=True)
        if force or (self.total > 0 and done >= self.total):
            print(file=sys.stderr, flush=True)


def format_duration(seconds: float) -> str:
    total_seconds = max(int(seconds), 0)
    minutes, secs = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours:d}h{minutes:02d}m{secs:02d}s"
    if minutes:
        return f"{minutes:d}m{secs:02d}s"
    return f"{secs:d}s"


def resolve_concurrency(requested: int, total: int) -> int:
    if total <= 0:
        return 1
    if requested > 0:
        return min(requested, total)
    cpu_count = os.cpu_count() or 4
    auto_concurrency = max(8, cpu_count * 2)
    return min(auto_concurrency, DEFAULT_AUTO_CONCURRENCY_CAP, total)


def is_auth_error_message(message: str) -> bool:
    return any(pattern in message for pattern in AUTH_ERROR_PATTERNS)


def encrypt_token(text: str) -> str:
    raw = text.encode("utf-8")
    size = len(raw)
    key_size = len(ENCRYPTION_KEY)
    output = bytearray(size)
    for index, byte_value in enumerate(raw):
        key_index = (size - index) % key_size
        key_value = ord(ENCRYPTION_KEY[key_index])
        output[index] = (key_value ^ (255 - byte_value)) & 0xFF
    return base64.b64encode(bytes(output)).decode("ascii")


def decrypt_token(text: str) -> str:
    raw = base64.b64decode(text)
    size = len(raw)
    key_size = len(ENCRYPTION_KEY)
    output = bytearray(size)
    for index, byte_value in enumerate(raw):
        key_index = (size - index) % key_size
        key_value = ord(ENCRYPTION_KEY[key_index])
        output[index] = 255 - (key_value ^ byte_value)
    return output.decode("utf-8")


def parse_curl_command(curl_path: Path) -> MemberLookupTemplate:
    raw = curl_path.read_text(encoding="utf-8").strip()
    compact = re.sub(r"\\\r?\n", " ", raw)
    tokens = shlex.split(compact, posix=True)
    if not tokens or tokens[0] != "curl":
        raise ValueError(f"{curl_path} 不是合法的 curl 命令文件")

    url = ""
    headers: dict[str, str] = {}
    i = 1
    while i < len(tokens):
        token = tokens[i]
        if token in {"-H", "--header"}:
            if i + 1 >= len(tokens):
                raise ValueError("curl 头部参数缺少值")
            header = tokens[i + 1]
            if ":" not in header:
                raise ValueError(f"无法解析 header: {header}")
            name, value = header.split(":", 1)
            headers[name.strip()] = value.strip()
            i += 2
            continue
        if token == "--url":
            if i + 1 >= len(tokens):
                raise ValueError("--url 缺少值")
            url = tokens[i + 1]
            i += 2
            continue
        if not token.startswith("-") and not url:
            url = token
        i += 1

    if not url:
        raise ValueError("curl 中未找到请求 URL")

    headers = {
        key: value
        for key, value in headers.items()
        if key.lower() not in {"host", "content-length"}
    }

    parts = urlsplit(url)
    encrypted_pairs = parse_qsl(parts.query, keep_blank_values=True)
    params: list[tuple[str, str]] = []
    for encrypted_key, encrypted_value in encrypted_pairs:
        params.append((decrypt_token(encrypted_key), decrypt_token(encrypted_value)))

    plain_keys = {key for key, _ in params}
    if "matching" not in plain_keys:
        raise ValueError(f"未在 curl query 中找到 matching 参数，实际参数: {sorted(plain_keys)}")

    return MemberLookupTemplate(
        scheme=parts.scheme,
        netloc=parts.netloc,
        path=parts.path,
        fragment=parts.fragment,
        headers=headers,
        params=params,
        matching_key="matching",
        token_expire_at=extract_jwt_expire_at(headers),
    )


def extract_jwt_expire_at(headers: dict[str, str]) -> int | None:
    authorization = next(
        (value for key, value in headers.items() if key.lower() == "authorization"),
        "",
    )
    if not authorization.startswith("Bearer "):
        return None
    token = authorization.split(" ", 1)[1].strip()
    parts = token.split(".")
    if len(parts) != 3:
        return None
    payload = parts[1]
    payload += "=" * (-len(payload) % 4)
    try:
        decoded = json.loads(base64.urlsafe_b64decode(payload))
    except Exception:
        return None
    exp = decoded.get("exp")
    return int(exp) if isinstance(exp, (int, float)) else None


def find_default_excel(cwd: Path) -> Path:
    candidates = sorted(cwd.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("当前目录没有找到 .xlsx 文件")
    return candidates[0]


def normalize_plate(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    text = re.sub(r"\s+", "", text)
    if text in {"", "-", "NONE", "NULL"}:
        return ""
    if "无牌" in text:
        return ""
    return text


def find_plate_column(worksheet) -> int:
    headers = [cell.value for cell in worksheet[1]]
    for candidate in PLATE_HEADER_CANDIDATES:
        for index, header in enumerate(headers, start=1):
            if str(header).strip() == candidate:
                return index
    raise ValueError(f"未找到车牌列，首行表头为: {headers}")


def extract_plate_details(
    worksheet,
    plate_column: int,
    limit: int | None,
    progress: ProgressPrinter | None = None,
) -> PlateExtraction:
    plate_to_rows: dict[str, list[int]] = defaultdict(list)
    empty_row_indexes: list[int] = []
    unique_plates: list[str] = []
    plate_col_index = plate_column - 1
    max_row = worksheet.max_row
    if limit is not None:
        max_row = min(max_row, limit + 1)

    for row_index, row_values in enumerate(
        worksheet.iter_rows(min_row=2, max_row=max_row, values_only=True),
        start=2,
    ):
        plate = normalize_plate(row_values[plate_col_index] if plate_col_index < len(row_values) else "")
        if plate:
            if plate not in plate_to_rows:
                unique_plates.append(plate)
            plate_to_rows[plate].append(row_index)
        else:
            empty_row_indexes.append(row_index)
        if progress is not None:
            progress.update(extra=f"唯一车牌 {len(unique_plates)}")

    total_rows = max_row - 1 if max_row >= 2 else 0
    if progress is not None:
        progress.finish(extra=f"唯一车牌 {len(unique_plates)}")
    return PlateExtraction(
        unique_plates=unique_plates,
        plate_to_rows=dict(plate_to_rows),
        empty_row_indexes=empty_row_indexes,
        total_rows=total_rows,
    )


def parse_lookup_response(plate: str, payload: dict) -> LookupResult:
    code = str(payload.get("code", ""))
    data = payload.get("payload") or {}
    rows = data.get("row") or []
    if not rows:
        if code not in {"1001", ""}:
            message = payload.get("message") or payload.get("hint") or f"接口返回 code={code}"
            return LookupResult(plate=plate, status="error", error=str(message))
        return LookupResult(plate=plate, status="not_found")

    normalized_plate = normalize_plate(plate)
    matched_row = None
    for row in rows:
        if normalize_plate(row.get("plate")) == normalized_plate:
            matched_row = row
            break
    if matched_row is None and len(rows) == 1:
        matched_row = rows[0]
    if matched_row is None:
        return LookupResult(plate=plate, status="error", error="接口返回多条结果但未命中精确车牌")

    member = matched_row.get("member") or {}
    member_level = member.get("member_level") or {}
    mobile = str(member.get("mobile") or matched_row.get("mobile") or "").strip()
    level_name = str(
        member.get("level_name")
        or member_level.get("name")
        or matched_row.get("level_name")
        or ""
    ).strip()
    matched_plate = normalize_plate(matched_row.get("plate"))
    return LookupResult(
        plate=plate,
        mobile=mobile,
        level_name=level_name,
        matched_plate=matched_plate,
        status="success",
    )


async def fetch_member_info(
    session: aiohttp.ClientSession,
    plate: str,
    url: str,
    retries: int,
) -> LookupResult:
    last_error = ""
    for attempt in range(1, retries + 1):
        try:
            async with session.get(url) as response:
                body_text = await response.text(encoding="utf-8", errors="ignore")
                if response.status == 401:
                    return LookupResult(plate=plate, status="error", error="401 未授权，token 可能已过期")
                if response.status in RETRYABLE_HTTP_STATUSES:
                    last_error = f"HTTP {response.status}: {body_text[:200]}"
                    if attempt >= retries:
                        return LookupResult(plate=plate, status="error", error=last_error)
                    await asyncio.sleep(min(2 ** (attempt - 1), 8))
                    continue
                if response.status >= 400:
                    return LookupResult(plate=plate, status="error", error=f"HTTP {response.status}: {body_text[:200]}")
                try:
                    payload = json.loads(body_text)
                except json.JSONDecodeError as exc:
                    raise RuntimeError(f"响应不是合法 JSON: {body_text[:200]}") from exc
                result = parse_lookup_response(plate, payload)
                if result.status == "error" and is_auth_error_message(result.error):
                    return result
                return result
        except (aiohttp.ClientError, asyncio.TimeoutError, RuntimeError) as exc:
            last_error = str(exc)
            if attempt >= retries:
                break
            await asyncio.sleep(min(2 ** (attempt - 1), 8))
    return LookupResult(plate=plate, status="error", error=last_error or "请求失败")


async def fetch_all_member_info(
    template: MemberLookupTemplate,
    plates: Iterable[str],
    concurrency: int,
    timeout_seconds: float,
    retries: int,
    cached_results: dict[str, LookupResult],
) -> dict[str, LookupResult]:
    pending_requests = [
        (plate, template.build_url(plate))
        for plate in plates
        if plate not in cached_results
    ]
    results = dict(cached_results)
    total = len(pending_requests)
    if not pending_requests:
        print("[会员查询] 无需发起请求，全部命中本地缓存", file=sys.stderr)
        return results

    queue: asyncio.Queue[tuple[str, str] | None] = asyncio.Queue()
    for request in pending_requests:
        queue.put_nowait(request)

    connector = aiohttp.TCPConnector(
        limit=concurrency,
        limit_per_host=concurrency,
        ttl_dns_cache=600,
        enable_cleanup_closed=True,
    )
    session_timeout = aiohttp.ClientTimeout(
        total=None,
        connect=timeout_seconds,
        sock_connect=timeout_seconds,
        sock_read=timeout_seconds,
    )
    completed = 0
    progress = ProgressPrinter(label="会员查询", total=total, unit="牌")
    success_count = 0
    not_found_count = 0
    error_count = 0
    auth_error_message = ""
    stop_event = asyncio.Event()

    async with aiohttp.ClientSession(
        connector=connector,
        headers=template.headers,
        timeout=session_timeout,
    ) as session:
        async def worker() -> None:
            nonlocal completed, success_count, not_found_count, error_count, auth_error_message
            while True:
                request = await queue.get()
                try:
                    if request is None:
                        return
                    if stop_event.is_set():
                        continue
                    plate, url = request
                    result = await fetch_member_info(session, plate, url, retries)
                    results[plate] = result
                    completed += 1
                    if result.status == "success":
                        success_count += 1
                    elif result.status == "not_found":
                        not_found_count += 1
                    elif result.status == "error":
                        error_count += 1
                        if is_auth_error_message(result.error):
                            auth_error_message = result.error
                            stop_event.set()
                    progress.update(
                        extra=(
                            f"成功 {success_count}"
                            f" | 未匹配 {not_found_count}"
                            f" | 失败 {error_count}"
                        ),
                    )
                finally:
                    queue.task_done()

        worker_count = min(concurrency, max(total, 1))
        workers = [asyncio.create_task(worker()) for _ in range(worker_count)]
        await queue.join()
        for _ in workers:
            queue.put_nowait(None)
        await asyncio.gather(*workers)
    if auth_error_message:
        progress.finish(
            extra=(
                f"成功 {success_count}"
                f" | 未匹配 {not_found_count}"
                f" | 失败 {error_count}"
                f" | {auth_error_message}"
            ),
        )
        raise TokenExpiredError(f"查询过程中鉴权失效: {auth_error_message}")
    progress.finish(
        extra=(
            f"成功 {success_count}"
            f" | 未匹配 {not_found_count}"
            f" | 失败 {error_count}"
        ),
    )
    return results


def load_cache(cache_path: Path) -> dict[str, LookupResult]:
    if not cache_path.exists():
        return {}
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    cache: dict[str, LookupResult] = {}
    for plate, value in payload.items():
        if not isinstance(value, dict):
            continue
        cache[plate] = LookupResult(
            plate=plate,
            mobile=str(value.get("mobile", "")),
            level_name=str(value.get("level_name", "")),
            matched_plate=str(value.get("matched_plate", "")),
            status=str(value.get("status", "")),
            error=str(value.get("error", "")),
        )
    return cache


def save_cache(cache_path: Path, results: dict[str, LookupResult]) -> None:
    payload = {
        plate: asdict(result)
        for plate, result in results.items()
        if result.cacheable()
    }
    cache_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def add_or_get_column(worksheet, header_name: str) -> int:
    for index, cell in enumerate(worksheet[1], start=1):
        if str(cell.value).strip() == header_name:
            return index
    column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=column_index, value=header_name)
    return column_index


def build_output_path(input_path: Path, output_path: str | None) -> Path:
    if output_path:
        return Path(output_path)
    return input_path


def apply_results_to_workbook(
    input_path: Path,
    output_path: Path,
    plate_to_rows: dict[str, list[int]],
    empty_row_indexes: list[int],
    results: dict[str, LookupResult],
    sheet_name: str | None,
) -> None:
    workbook = load_workbook(input_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    total_rows = sum(len(row_indexes) for row_indexes in plate_to_rows.values()) + len(empty_row_indexes)
    progress = ProgressPrinter(label="回写Excel", total=total_rows, unit="行")

    mobile_col = add_or_get_column(worksheet, "会员手机号")
    level_col = add_or_get_column(worksheet, "会员等级")
    matched_plate_col = add_or_get_column(worksheet, "会员匹配车牌")
    status_col = add_or_get_column(worksheet, "会员查询状态")
    error_col = add_or_get_column(worksheet, "会员查询错误")

    for plate, row_indexes in plate_to_rows.items():
        result = results.get(plate)
        mobile = result.mobile if result else ""
        level_name = result.level_name if result else ""
        matched_plate = result.matched_plate if result else ""
        status = result.status if result else ""
        error_text = result.error if result else ""
        for row_index in row_indexes:
            worksheet.cell(row=row_index, column=mobile_col, value=mobile)
            worksheet.cell(row=row_index, column=level_col, value=level_name)
            worksheet.cell(row=row_index, column=matched_plate_col, value=matched_plate)
            worksheet.cell(row=row_index, column=status_col, value=status)
            worksheet.cell(row=row_index, column=error_col, value=error_text)
            progress.update(extra=f"当前车牌 {plate}")

    for row_index in empty_row_indexes:
        worksheet.cell(row=row_index, column=mobile_col, value="")
        worksheet.cell(row=row_index, column=level_col, value="")
        worksheet.cell(row=row_index, column=matched_plate_col, value="")
        worksheet.cell(row=row_index, column=status_col, value="")
        worksheet.cell(row=row_index, column=error_col, value="")
        progress.update(extra="空车牌")

    progress.finish(extra="正在保存文件")
    workbook.save(output_path)
    print("[回写Excel] 文件保存完成", file=sys.stderr)


def validate_token_expiration(expire_at: int | None) -> int | None:
    if expire_at is None:
        return None
    now = int(time.time())
    remain = expire_at - now
    expire_text = datetime.fromtimestamp(expire_at, tz=timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S")
    if remain <= 0:
        raise TokenExpiredError(f"curl 中的 Bearer token 已过期，过期时间: {expire_text}")
    if remain < 300:
        print(f"[警告] curl 中的 Bearer token 将在 5 分钟内过期，过期时间: {expire_text}", file=sys.stderr)
    return remain


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据停车场进出记录中的车牌，高并发查询会员手机号和会员等级并回写 Excel。",
    )
    parser.add_argument("--input", help="输入 Excel 文件路径，默认取当前目录第一个 .xlsx")
    parser.add_argument("--output", help="输出 Excel 文件路径；不传时直接回写原 Excel")
    parser.add_argument("--sheet", help="指定工作表名称，默认使用活动工作表")
    parser.add_argument("--curl-file", default=DEFAULT_CURL_FILE, help=f"会员查询 curl 文件，默认 {DEFAULT_CURL_FILE}")
    parser.add_argument("--cache-file", default=DEFAULT_CACHE_FILE, help=f"本地缓存文件，默认 {DEFAULT_CACHE_FILE}")
    parser.add_argument(
        "--concurrency",
        type=int,
        default=DEFAULT_CONCURRENCY,
        help="并发数，默认 0 表示自动高并发",
    )
    parser.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT, help=f"单次请求超时秒数，默认 {DEFAULT_TIMEOUT}")
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES, help=f"失败重试次数，默认 {DEFAULT_RETRIES}")
    parser.add_argument("--limit", type=int, help="仅处理前 N 条数据，便于测试")
    parser.add_argument("--no-cache", action="store_true", help="不读取也不写入本地缓存")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    cwd = Path.cwd()
    input_path = Path(args.input) if args.input else find_default_excel(cwd)
    curl_path = Path(args.curl_file)
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")
    if not curl_path.exists():
        raise FileNotFoundError(f"curl 文件不存在: {curl_path}")
    output_path = build_output_path(input_path, args.output)
    cache_path = Path(args.cache_file)

    template = parse_curl_command(curl_path)
    validate_token_expiration(template.token_expire_at)

    workbook = load_workbook(input_path, read_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook.active
    plate_column = find_plate_column(worksheet)
    total_rows = max(worksheet.max_row - 1, 0)
    if args.limit is not None:
        total_rows = min(total_rows, max(args.limit, 0))
    scan_progress = ProgressPrinter(label="读取车牌", total=total_rows, unit="行")
    extraction = extract_plate_details(worksheet, plate_column, args.limit, progress=scan_progress)
    workbook.close()

    cache_results = {} if args.no_cache else load_cache(cache_path)
    unique_plates = extraction.unique_plates
    cached_hits = sum(1 for plate in unique_plates if plate in cache_results)
    resolved_concurrency = resolve_concurrency(args.concurrency, max(len(unique_plates) - cached_hits, 0))
    print(
        f"待处理行数: {extraction.total_rows}，唯一车牌数: {len(unique_plates)}，缓存命中: {cached_hits}，并发数: {resolved_concurrency}",
        file=sys.stderr,
    )

    results = asyncio.run(
        fetch_all_member_info(
            template=template,
            plates=unique_plates,
            concurrency=resolved_concurrency,
            timeout_seconds=max(args.timeout, 1),
            retries=max(args.retries, 1),
            cached_results=cache_results,
        )
    )

    if not args.no_cache:
        save_cache(cache_path, results)

    apply_results_to_workbook(
        input_path=input_path,
        output_path=output_path,
        plate_to_rows=extraction.plate_to_rows,
        empty_row_indexes=extraction.empty_row_indexes,
        results=results,
        sheet_name=args.sheet,
    )

    success_count = sum(1 for plate in unique_plates if results.get(plate) and results[plate].status == "success")
    not_found_count = sum(1 for plate in unique_plates if results.get(plate) and results[plate].status == "not_found")
    error_count = sum(1 for plate in unique_plates if results.get(plate) and results[plate].status == "error")
    error_reasons = Counter(
        result.error
        for result in results.values()
        if result.status == "error" and result.error
    )
    print(f"输出文件: {output_path}")
    print(f"成功: {success_count}，未匹配: {not_found_count}，失败: {error_count}")
    if error_reasons:
        print("错误Top:")
        for reason, count in error_reasons.most_common(5):
            print(f"{count} x {reason}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except TokenExpiredError as exc:
        print(f"鉴权失败: {exc}", file=sys.stderr)
        raise SystemExit(2)
    except KeyboardInterrupt:
        print("已中断", file=sys.stderr)
        raise SystemExit(130)
